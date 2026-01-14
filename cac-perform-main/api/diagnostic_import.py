#!/usr/bin/env python3
"""
Script de diagnostic pour l'import des balances
"""

from pymongo import MongoClient
from bson import ObjectId
import openpyxl
import os

def diagnostic_import():
    """Diagnostique les problèmes d'import des balances"""
    
    client = MongoClient('mongodb://localhost:27017/')
    db = client['cac_perform']
    
    print("🔍 Diagnostic de l'Import des Balances")
    print("=" * 50)
    
    # Trouver la mission la plus récente
    missions = list(db.Mission1.find({}, {'_id': 1, 'nom': 1, 'balances': 1}).sort('_id', -1).limit(1))
    
    if not missions:
        print("❌ Aucune mission trouvée")
        return
    
    mission = missions[0]
    mission_id = mission['_id']
    balances = mission.get('balances', [])
    
    print(f"📋 Mission: {mission_id}")
    print(f"📊 Balances: {len(balances)}")
    
    if len(balances) < 2:
        print("❌ Moins de 2 balances")
        return
    
    # Analyser chaque balance
    for i, balance_id in enumerate(balances):
        print(f"\n🔍 Balance {i+1} (ID: {balance_id}):")
        
        balance = db.Balance.find_one({'_id': ObjectId(balance_id)})
        if not balance:
            print("  ❌ Balance introuvable")
            continue
        
        balance_data = balance.get('balance', [])
        print(f"  📊 Lignes importées: {len(balance_data)}")
        
        if balance_data:
            # Analyser la première ligne
            premiere_ligne = balance_data[0]
            print(f"  📋 Première ligne:")
            print(f"    - Numéro compte: {premiere_ligne.get('numero_compte')}")
            print(f"    - Libellé: {premiere_ligne.get('libelle', 'N/A')[:50]}...")
            print(f"    - Débit initial: {premiere_ligne.get('debit_initial')}")
            print(f"    - Crédit initial: {premiere_ligne.get('credit_initial')}")
            print(f"    - Débit final: {premiere_ligne.get('debit_fin')}")
            print(f"    - Crédit final: {premiere_ligne.get('credit_fin')}")
            
            # Compter les comptes de bilan
            comptes_bilan = [l for l in balance_data 
                           if str(l.get('numero_compte', '')).startswith(('1', '2', '3', '4', '5'))]
            print(f"  🏦 Comptes de bilan: {len(comptes_bilan)}")
            
            # Vérifier les comptes problématiques
            compte_101300 = next((l for l in balance_data 
                                if str(l.get('numero_compte')) == '101300'), None)
            if compte_101300:
                print(f"  ✅ Compte 101300 trouvé:")
                print(f"    - Débit initial: {compte_101300.get('debit_initial')}")
                print(f"    - Crédit initial: {compte_101300.get('credit_initial')}")
                print(f"    - Débit final: {compte_101300.get('debit_fin')}")
                print(f"    - Crédit final: {compte_101300.get('credit_fin')}")
            else:
                print(f"  ❌ Compte 101300 non trouvé")
        else:
            print("  ❌ Aucune donnée importée")
    
    # Vérifier les fichiers sources
    print(f"\n📁 Vérification des fichiers sources:")
    docs_dir = "../docs"
    if os.path.exists(docs_dir):
        fichiers = [f for f in os.listdir(docs_dir) if f.endswith(('.xlsx', '.xls'))]
        print(f"  📊 Fichiers Excel trouvés: {len(fichiers)}")
        for fichier in fichiers:
            print(f"    - {fichier}")
            
            # Analyser le fichier Excel
            try:
                file_path = os.path.join(docs_dir, fichier)
                workbook = openpyxl.load_workbook(file_path)
                print(f"      Feuilles: {workbook.sheetnames}")
                
                # Analyser la première feuille
                sheet = workbook[workbook.sheetnames[0]]
                print(f"      Lignes: {sheet.max_row}, Colonnes: {sheet.max_column}")
                
                # Vérifier les en-têtes
                if sheet.max_row > 0:
                    headers = [cell.value for cell in sheet[1]]
                    print(f"      En-têtes: {headers[:5]}...")
                
            except Exception as e:
                print(f"      ❌ Erreur lecture: {e}")
    else:
        print(f"  ❌ Dossier docs non trouvé: {docs_dir}")

if __name__ == "__main__":
    diagnostic_import()

