from src import mongo
from genericpath import exists
from mongoengine import Document, StringField, EmbeddedDocument, EmbeddedDocumentField, ListField, DictField
import bcrypt
import openpyxl
from pymongo import MongoClient
from bson import ObjectId
import json
from io import BytesIO
import os

# Connexion directe MongoDB en cas de problème avec mongo.get_db
try:
    db = mongo.get_db
    if db is None:
        raise Exception("mongo.get_db est None")
except:
    print("⚠️  Utilisation de la connexion MongoDB directe")
    client = MongoClient('mongodb://localhost:27017/')
    db = client['cac_perform']


# ==============================
#  Clients
# ==============================
class Client(Document):

    @classmethod
    def afficher_clients(cls):
        try:
            cust_data = list(db.Client.find().sort({"_id": -1}))
            for client in cust_data:
                client['_id'] = str(client['_id'])
            return cust_data
        except Exception as e:
            print(f"An exception : {str(e)}")

    @classmethod
    def afficher_missions(cls, id):
        results = []
        final = []
        
        # Normaliser l'id (s'assurer que c'est une string)
        id_str = str(id).strip()
        
        print(f"🔍 Recherche des missions pour client ID: '{id_str}'")
        
        # Chercher avec l'id tel quel
        query = {"id_client": id_str}
        results = db.Mission1.find(query)
        
        count = db.Mission1.count_documents(query)
        print(f"📊 Missions trouvées avec id_client='{id_str}': {count}")
        
        # Si aucune mission trouvée, essayer avec ObjectId si l'id ressemble à un ObjectId
        if count == 0:
            try:
                # Essayer avec ObjectId
                query_objid = {"id_client": ObjectId(id_str)}
                count_objid = db.Mission1.count_documents(query_objid)
                print(f"📊 Tentative avec ObjectId: {count_objid} missions")
                if count_objid > 0:
                    results = db.Mission1.find(query_objid)
                    count = count_objid
            except:
                pass
        
        # Récupérer toutes les missions pour debug
        total_missions = db.Mission1.count_documents({})
        print(f"📊 Total de missions dans la base: {total_missions}")
        
        if total_missions > 0 and count == 0:
            # Afficher quelques exemples pour debug
            sample_missions = list(db.Mission1.find({}).limit(3))
            print(f"⚠️  Exemples d'id_client dans la base:")
            for sample in sample_missions:
                sample_id_client = sample.get('id_client', 'NON DÉFINI')
                print(f"   - Mission {sample['_id']}: id_client='{sample_id_client}' (type: {type(sample_id_client)})")
        
        for result in results:
            result['_id'] = str(result['_id'])
            final.append(result)
        
        print(f"✅ Retour de {len(final)} missions pour le client {id_str}")
        return final

    @classmethod
    def ajouter_client(cls, data):
        try:
            cust_data = db.Client.insert_one(data)
            inserted_id = str(cust_data.inserted_id)
            return inserted_id
        except Exception as e:
            print(f'An exception occurred: {str(e)}')

    @classmethod
    def modifier_client(cls, data):
        try:
            _id = str(data['_id'])
            id_client = ObjectId(_id)

            nom = data['nom']
            activite = data['activite']
            referentiel = data['referentiel']
            forme_juridique = data['forme_juridique']
            capital = data['capital']
            siege_social = data['siege_social']
            adresse = data['adresse']
            n_cc = data['n_cc']

            if id_client:
                db.Client.update_one(
                    {"_id": id_client},
                    {"$set": {
                        "nom": nom,
                        "activite": activite,
                        "referentiel": referentiel,
                        "forme_juridique": forme_juridique,
                        "capital": capital,
                        "siege_social": siege_social,
                        "adresse": adresse,
                        "n_cc": n_cc,
                    }}
                )
                return "Updated succeeded"
        except Exception as e:
            print(f'An exception occurred: {str(e)}')

    @classmethod
    def info_client(cls, id_clit):
        try:
            id_client = id_clit
            info = db.Client.find_one({"_id": id_client})

            valeurs = {}
            if info:
                valeurs = {
                    "_id": str(id_client),
                    "nom": info['nom'],
                    "activite": info['activite'],
                    "siege_social": info['siege_social'],
                    "adresse": info['adresse'],
                    "referentiel": info['referentiel'],
                    "forme_juridique": info['forme_juridique'],
                    "capital": info['capital'],
                    "n_cc": info['n_cc'],
                }
            return valeurs
        except Exception as e:
            print(f"An exception : {str(e)}")
            return None

    @classmethod
    def supprimer_client(cls, id_client):
        """Supprime un client et toutes ses missions associées"""
        try:
            # Convertir l'ID en ObjectId
            client_id = ObjectId(str(id_client))
            
            # Vérifier que le client existe
            client_info = db.Client.find_one({"_id": client_id})
            if not client_info:
                return {"success": False, "message": "Client non trouvé"}
            
            # Supprimer toutes les missions associées au client
            missions_deleted = db.Mission1.delete_many({"id_client": str(id_client)})
            
            # Supprimer le client
            client_deleted = db.Client.delete_one({"_id": client_id})
            
            if client_deleted.deleted_count > 0:
                return {
                    "success": True, 
                    "message": f"Client supprimé avec succès. {missions_deleted.deleted_count} mission(s) supprimée(s).",
                    "client_name": client_info.get('nom', 'Inconnu')
                }
            else:
                return {"success": False, "message": "Erreur lors de la suppression du client"}
                
        except Exception as e:
            print(f"Erreur lors de la suppression du client: {str(e)}")
            return {"success": False, "message": f"Erreur: {str(e)}"}

    def choix_referentiel(self):
        try:
            ref = db.Grouping.find()
            return ref
        except Exception as e:
            print(f"An exception as occured: {str(e)}")


# ==============================
#  Manager
# ==============================
class Manager(Document):
    _id = StringField()
    email = StringField(required=True)
    mot_de_passe = StringField(required=True)

    def sign_in(self, data):
        try:
            email = data['mail']
            password = data['pwd']
            user = db.Manager.find_one({"email": email})
            if not user:
                return None

            stored = user.get("mot_de_passe", "")
            ok = False
            try:
                # tenter bcrypt si le hash semble valide
                if isinstance(stored, str):
                    stored_bytes = stored.encode('utf-8')
                else:
                    stored_bytes = stored
                ok = bcrypt.checkpw(password.encode('utf-8'), stored_bytes)
            except Exception:
                ok = False

            # fallback: mot de passe en clair (pour compat rétro)
            if not ok and stored == password:
                ok = True

            if ok:
                return {"email": user.get("email"), "_id": str(user.get("_id"))}
            return None
        except Exception as e:
            print(f"An exception occurred : {str(e)}")
            return None


# ==============================
#  Modèles embarqués
# ==============================
class LigneComptable(EmbeddedDocument):
    num_compte = StringField(required=True, unique=True)
    libelle = StringField()
    solde = DictField()


class BalanceComptable(EmbeddedDocument):
    lignes = ListField(EmbeddedDocumentField(LigneComptable))


# ==============================
#  Mission
# ==============================
class Mission(Document):
    # champs (non utilisés par mongoengine ici mais conservés)
    balance = EmbeddedDocumentField(BalanceComptable)
    annee_auditee = ListField()

    # ---------- Revue analytique ----------
    def revue_analytique(self, id_mission):
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        variations = mission.get("balance_variation", [])

        # mapping_efi.json (un niveau au-dessus de ce fichier)
        mapping_path = os.path.join(os.path.dirname(__file__), "..", "mapping_efi.json")
        with open(mapping_path, "r", encoding="utf-8") as f:
            mapping_root = json.load(f)
            mapping = mapping_root["structure"]

        # seuils (matérialité choisie si existe)
        materiality = mission.get("materiality", [])
        choice = next((m for m in materiality if m.get("choice")), None)
        perf_mat = abs(int(choice.get("performance_materiality", 0))) if choice else 0

        def collect_prefixes(bloc):
            prefixes = []
            for key in ("brut_cpt", "amor_cpt", "net_cpt", "brut_except_cpt", "amor_except_cpt", "net_except_cpt"):
                if key in bloc and bloc[key]:
                    if isinstance(bloc[key], str):
                        prefixes += [p.strip() for p in bloc[key].split(",") if p.strip()]
                    elif isinstance(bloc[key], list):
                        prefixes += [str(p).strip() for p in bloc[key] if str(p).strip()]
            return prefixes

        def map_efi(numero):
            refs = []
            for bloc in mapping:
                prefixes = collect_prefixes(bloc)
                if any(str(numero).startswith(p) for p in prefixes):
                    refs.append(bloc["ref"])
            return sorted(list(set(refs)))

        out = []
        for row in variations:
            n = int(row.get("solde_n", 0) or 0)
            n1 = int(row.get("solde_n1", 0) or 0)
            delta = n - n1
            delta_pct = (delta / (abs(n1) if n1 else 1.0))
            efi_refs = map_efi(row.get("numero_compte", ""))

            # Commentaire automatique (EFI + Matérialité)
            commentaire_auto = (
                f"EFI: {', '.join(efi_refs) if efi_refs else 'Aucun'} — "
                f"Matérialité: {'Oui' if (perf_mat and abs(delta) >= perf_mat) else 'Non'}"
            )

            out.append({
                "numero_compte": row.get("numero_compte"),
                "libelle": row.get("libelle"),
                "solde_n": n,
                "solde_n1": n1,
                "delta_abs": delta,
                "delta_pct": delta_pct,
                "commentaire_auto": commentaire_auto,
                "commentaire_perso": ""  # Commentaire personnalisé vide par défaut
            })

        db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"revue_analytique": out}})
        return out

    def update_commentaire_perso(self, id_mission, numero_compte, commentaire_perso):
        """
        Met à jour (ou initialise) le champ 'commentaire_perso' pour une ligne
        de la revue analytique correspondant au numero_compte.
        Retourne True si l'opération aboutit, False sinon.
        """
        try:
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return False
            revue = mission.get("revue_analytique", [])
            if not revue:
                return False
            index = next((i for i, it in enumerate(revue) if it.get("numero_compte") == numero_compte), None)
            if index is None:
                return False
            path = f"revue_analytique.{index}.commentaire_perso"
            result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {path: commentaire_perso or ""}})
            return result.acknowledged
        except Exception:
            return False

    def update_commentaire_perso(self, id_mission, numero_compte, commentaire_perso):
        """
        Met à jour le commentaire personnalisé pour un compte spécifique
        """
        try:
            print(f"Début de mise à jour du commentaire pour le compte {numero_compte}")
            
            # Vérifier d'abord que la mission existe
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                print(f"Mission {id_mission} non trouvée")
                return False
            
            # Vérifier que la revue analytique existe
            if "revue_analytique" not in mission:
                print(f"Revue analytique non trouvée pour la mission {id_mission}")
                return False
            
            # Trouver l'index du compte dans la revue analytique
            revue_analytique = mission["revue_analytique"]
            compte_index = None
            
            for i, item in enumerate(revue_analytique):
                if item.get("numero_compte") == numero_compte:
                    compte_index = i
                    break
            
            if compte_index is None:
                print(f"Compte {numero_compte} non trouvé dans la revue analytique")
                return False
            
            # Vérifier si le commentaire a réellement changé
            current_comment = revue_analytique[compte_index].get("commentaire_perso", "")
            if current_comment == commentaire_perso:
                print(f"Commentaire identique pour le compte {numero_compte}, aucune modification nécessaire")
                return True  # Retourner True car c'est un succès (pas d'erreur)
            
            print(f"Mise à jour du commentaire: '{current_comment}' -> '{commentaire_perso}'")
            
            # Mettre à jour le commentaire personnalisé
            result = db.Mission1.update_one(
                {"_id": ObjectId(id_mission)},
                {"$set": {f"revue_analytique.{compte_index}.commentaire_perso": commentaire_perso}}
            )
            
            success = result.modified_count > 0
            if success:
                print(f"Commentaire mis à jour avec succès pour le compte {numero_compte}")
            else:
                print(f"Aucune modification effectuée pour le compte {numero_compte} (modified_count: {result.modified_count})")
            
            return success
            
        except Exception as e:
            print(f"Erreur lors de la mise à jour du commentaire: {e}")
            import traceback
            traceback.print_exc()
            return False

    # ---------- Nouvelle mission ----------
    def nouvelle_mission(self, balances, annee_auditee, id_client, date_debut, date_fin):
        balance_ids = []
        les_balance_n_n1 = []
        annee_balance = annee_auditee

        for balance in balances:
            balance_created = self.creation_balance(balance, int(annee_auditee), id_client)
            annee_auditee = int(annee_auditee) - 1

            tuple_en_tableau = list(*[balance_created])
            balance_ids.append(tuple_en_tableau[0])
            les_balance_n_n1.append(tuple_en_tableau[1])

        balance_variation = self.rapprochement_des_balances(
            les_balance_n_n1[0], les_balance_n_n1[1]
        )

        # Récupérer le référentiel (par défaut "syscohada")
        referentiel = "syscohada"  # Vous pouvez le récupérer depuis les données si nécessaire
        grouping_model = self.create_grouping(balance_variation, referentiel)
        etats = self.prod_efi(les_balance_n_n1[0], les_balance_n_n1[1], balance_variation)

        # S'assurer que id_client est bien une string
        id_client_str = str(id_client).strip()
        
        print(f"💾 Création de mission avec id_client='{id_client_str}'")
        
        result = db.Mission1.insert_one({
            "id_client": id_client_str,
            "annee_auditee": str(annee_balance),
            "date_debut": date_debut,
            "date_fin": date_fin,
            "balances": balance_ids,
            "balance_variation": balance_variation,
            "grouping": grouping_model,
            "efi": etats,
            "materiality": []
        })

        insert_id = str(result.inserted_id)
        print(f"✅ Mission créée avec ID: {insert_id}")
        
        # Vérifier que la mission a bien été sauvegardée
        mission_saved = db.Mission1.find_one({"_id": result.inserted_id})
        if mission_saved:
            print(f"✅ Mission vérifiée en base: id_client='{mission_saved.get('id_client')}'")
        else:
            print(f"⚠️  ATTENTION: Mission créée mais non trouvée en base!")
        res = {
            "id_client": id_client,
            "annee_auditee": str(annee_balance),
            "date_debut": date_debut,
            "date_fin": date_fin,
            "balances": balance_ids,
            "balance_variation": balance_variation,
            "grouping": grouping_model,
            "efi": etats,
            "materiality": []
        }

        format_id = {"_id": insert_id, "mission": res}
        self.audit_trail(format_id['_id'])
        return format_id

    # ---------- Lecture d'une balance Excel -> Mongo (Excel uniquement) ----------
    def creation_balance(self, balance_data, annee_auditee, id_client):
        balance = balance_data
        data = []

        try:
            # Traitement des fichiers Excel uniquement
            workbook = openpyxl.load_workbook(balance)
            
            # Détection automatique de la feuille de balance
            sheet = None
            sheet_name = None
            
            # Noms de feuilles acceptés (par ordre de priorité)
            accepted_sheet_names = [
                'Balance des comptes',  # Nom avec espaces (détecté dans vos fichiers)
                'Balance_des_comptes',  # Nom standard
                'BALANCE_2023',         # Votre format
                'BALANCE__2024',        # Votre format
                'Sage',                 # Format Sage
                'Sheet1',               # Format générique
                'Feuil1',               # Format français Excel (détecté dans vos fichiers)
                'Balance',              # Format court
                'Comptes',              # Format alternatif
            ]
            
            # Chercher la première feuille disponible
            for name in accepted_sheet_names:
                if name in workbook.sheetnames:
                    sheet = workbook[name]
                    sheet_name = name
                    print(f"✅ Feuille trouvée: '{name}' dans {balance.filename if hasattr(balance, 'filename') else 'fichier'}")
                    break
            
            # Si aucune feuille standard n'est trouvée, prendre la première
            if sheet is None and workbook.sheetnames:
                sheet = workbook[workbook.sheetnames[0]]
                sheet_name = workbook.sheetnames[0]
                print(f"⚠️  Feuille non standard utilisée: '{sheet_name}' dans {balance.filename if hasattr(balance, 'filename') else 'fichier'}")
            
            if sheet is None:
                raise Exception("Aucune feuille trouvée dans le fichier Excel")

            print(f"📊 Traitement de la feuille '{sheet_name}' avec {sheet.max_row} lignes et {sheet.max_column} colonnes")
            
            # Détection automatique du format et des colonnes
            print(f"🔍 Analyse du format de la feuille '{sheet_name}'...")
            
            # Analyser les premières lignes pour détecter le format
            format_detecte = None
            colonnes_map = {}
            
            # Formats supportés
            formats_supportes = {
                'standard': {
                    'pattern': ['numero', 'libelle', 'debit_initial', 'credit_initial', 'debit_mvt', 'credit_mvt', 'debit_fin', 'credit_fin'],
                    'detection': lambda row: len(row) >= 8 and any('compte' in str(cell).lower() for cell in row if cell)
                },
                'balance_simple': {
                    'pattern': ['numero', 'libelle', 'debit_initial', 'credit_initial', 'debit_fin', 'credit_fin'],
                    'detection': lambda row: len(row) >= 6 and (
                        any('compte' in str(cell).lower() for cell in row if cell) or
                        (str(row[0]).isdigit() and len(str(row[0])) >= 3)
                    )
                },
                'sage': {
                    'pattern': ['numero', 'libelle', 'solde_n1', 'mouvement_debit', 'mouvement_credit', 'solde_n'],
                    'detection': lambda row: len(row) >= 6 and str(row[0]).isdigit() and len(str(row[0])) >= 3
                },
                'simple': {
                    'pattern': ['numero', 'libelle', 'montant'],
                    'detection': lambda row: len(row) >= 3 and str(row[0]).isdigit()
                }
            }
            
            # Détecter le format (par ordre de priorité)
            # IMPORTANT: Si le fichier a 8 colonnes avec mouvements, utiliser "standard" en priorité
            # Ordre de détection: standard avant balance_simple si 8 colonnes détectées
            if sheet.max_column >= 8:
                format_priority = ['standard', 'balance_simple', 'sage', 'simple']
            else:
                format_priority = ['balance_simple', 'standard', 'sage', 'simple']
            
            for row_idx in range(1, min(10, sheet.max_row + 1)):
                row_data = [sheet.cell(row=row_idx, column=col).value for col in range(1, min(10, sheet.max_column + 1))]
                
                # Tester les formats par ordre de priorité
                for format_name in format_priority:
                    if format_name in formats_supportes:
                        format_info = formats_supportes[format_name]
                        if format_info['detection'](row_data):
                            format_detecte = format_name
                            print(f"✅ Format détecté: {format_name} à la ligne {row_idx}")
                            break
                
                if format_detecte:
                    break
            
            if not format_detecte:
                print("⚠️  Format non reconnu, utilisation du format par défaut")
                format_detecte = 'standard'
            
            print(f"\n{'='*60}")
            print(f"🔍 FORMAT FINAL DÉTECTÉ: {format_detecte}")
            print(f"   Nombre de colonnes dans le fichier: {sheet.max_column}")
            print(f"{'='*60}\n")
            
            # Traitement selon le format détecté
            if format_detecte == 'sage':
                # Format Sage: [numero, libelle, solde_n1, mouvement_debit, mouvement_credit, solde_n]
                print("📊 Traitement format Sage...")
                
                # Compteurs pour le diagnostic
                lignes_traitees = 0
                lignes_ignorees_colonnes = 0
                lignes_ignorees_vides = 0
                lignes_ignorees_numero_invalide = 0
                lignes_erreur = 0
                
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
                    # Vérifier que la ligne a au moins 6 colonnes
                    if len(row) < 6:
                        lignes_ignorees_colonnes += 1
                        continue
                        
                    if not row[0] or str(row[0]).strip() == '':
                        lignes_ignorees_vides += 1
                        continue
                    
                    try:
                        numero = str(row[0]).strip()
                        if not numero.isdigit() or len(numero) < 3:
                            lignes_ignorees_numero_invalide += 1
                            continue
                            
                        ligne = {}
                        ligne['numero_compte'] = numero
                        ligne['libelle'] = str(row[1] or '').strip()
                        
                        # Format Sage: solde_n1, mouvement_debit, mouvement_credit, solde_n
                        solde_n1 = float(row[2] or 0)
                        mouvement_debit = float(row[3] or 0)
                        mouvement_credit = float(row[4] or 0)
                        solde_n = float(row[5] or 0)
                        
                        # Calculer les soldes d'ouverture et de clôture
                        ligne['debit_initial'] = max(0, solde_n1) if solde_n1 > 0 else 0
                        ligne['credit_initial'] = abs(solde_n1) if solde_n1 < 0 else 0
                        ligne['debit_mvt'] = mouvement_debit
                        ligne['credit_mvt'] = mouvement_credit
                        ligne['debit_fin'] = max(0, solde_n) if solde_n > 0 else 0
                        ligne['credit_fin'] = abs(solde_n) if solde_n < 0 else 0
                        
                        ligne['solde_reel'] = ligne['debit_fin'] - ligne['credit_fin']
                        ligne['solde'] = abs(ligne['solde_reel'])
                        ligne['sign_solde'] = "D" if ligne['debit_fin'] >= ligne['credit_fin'] else "C"
                        
                        data.append(ligne)
                        lignes_traitees += 1
                        
                    except Exception as e:
                        lignes_erreur += 1
                        print(f"⚠️  Erreur ligne {row_idx} (Sage): {e}")
                        continue
                
                # Afficher le résumé du traitement
                print(f"📊 Résumé du traitement (format Sage):")
                print(f"   ✅ Lignes traitées avec succès: {lignes_traitees}")
                if lignes_ignorees_colonnes > 0:
                    print(f"   ⚠️  Lignes ignorées (moins de 6 colonnes): {lignes_ignorees_colonnes}")
                if lignes_ignorees_vides > 0:
                    print(f"   ⚠️  Lignes ignorées (numéro de compte vide): {lignes_ignorees_vides}")
                if lignes_ignorees_numero_invalide > 0:
                    print(f"   ⚠️  Lignes ignorées (numéro de compte invalide, doit être numérique avec 3+ chiffres): {lignes_ignorees_numero_invalide}")
                if lignes_erreur > 0:
                    print(f"   ❌ Lignes avec erreur: {lignes_erreur}")
            
            elif format_detecte == 'balance_simple':
                # Format balance simple: [numero, libelle, debit_initial, credit_initial, debit_fin, credit_fin]
                # Sans les colonnes de mouvement
                print(f"\n{'='*60}")
                print("📊 Traitement format balance simple...")
                print(f"{'='*60}\n")
                
                # Compteurs pour le diagnostic
                lignes_traitees = 0
                lignes_ignorees_colonnes = 0
                lignes_ignorees_vides = 0
                lignes_ignorees_numero_vide = 0
                lignes_erreur = 0
                
                header_row = 1
                
                # Trouver la ligne d'en-tête
                for row_idx in range(1, min(5, sheet.max_row + 1)):
                    row_data = [sheet.cell(row=row_idx, column=col).value for col in range(1, min(9, sheet.max_column + 1))]
                    if any('compte' in str(cell).lower() for cell in row_data if cell):
                        header_row = row_idx
                        print(f"📝 En-tête détecté à la ligne {row_idx}: {row_data}")
                        break
                
                # DÉTECTION DE L'ORDRE DES COLONNES (si 8 colonnes)
                colonnes_inversees_balance_simple = False
                if sheet.max_column >= 8:
                    header_row_data = [sheet.cell(row=header_row, column=col).value for col in range(1, min(9, sheet.max_column + 1))]
                    col7_header = str(header_row_data[6] if len(header_row_data) > 6 else '').strip()
                    col8_header = str(header_row_data[7] if len(header_row_data) > 7 else '').strip()
                    col7_val = col7_header.lower()
                    col8_val = col8_header.lower()
                    
                    print(f"📝 DÉTECTION DE L'ORDRE DES COLONNES (format balance_simple):")
                    print(f"   Colonne 7 (index 6): '{col7_header}'")
                    print(f"   Colonne 8 (index 7): '{col8_header}'")
                    
                    has_solde_credit_col7 = (
                        ('solde' in col7_val and 'credit' in col7_val) or
                        ('solde' in col7_val and 'crédit' in col7_val) or
                        (col7_val and 'credit' in col7_val and 'debit' not in col7_val and 'débit' not in col7_val)
                    )
                    
                    has_solde_debit_col8 = (
                        ('solde' in col8_val and 'debit' in col8_val) or
                        ('solde' in col8_val and 'débit' in col8_val) or
                        (col8_val and 'debit' in col8_val and 'credit' not in col8_val and 'crédit' not in col8_val)
                    )
                    
                    if has_solde_credit_col7 and has_solde_debit_col8:
                        colonnes_inversees_balance_simple = True
                        print(f"✅ Format détecté: Solde crédit (col 7) puis Solde débit (col 8) - COLONNES INVERSÉES")
                    else:
                        print(f"✅ Format détecté: Débit (col 7) puis Crédit (col 8) - FORMAT STANDARD")
                
                # Traitement des lignes de données
                for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                    # Vérifier que la ligne a au moins 6 colonnes
                    if len(row) < 6:
                        lignes_ignorees_colonnes += 1
                        continue
                        
                    if row[0] is None and row[1] is None:
                        lignes_ignorees_vides += 1
                        break  # Fin des données
                        
                    # Vérifier que la ligne contient des données valides
                    if not row[0] or str(row[0]).strip() == '':
                        lignes_ignorees_numero_vide += 1
                        continue
                    
                    try:
                        ligne = {}
                        # Convertir le numéro de compte en string, en gérant les cas None, 0, etc.
                        num_compte_raw = row[0]
                        if num_compte_raw is None:
                            raise ValueError("Numéro de compte est None")
                        num_compte_str = str(num_compte_raw).strip()
                        if not num_compte_str or num_compte_str.lower() == 'none' or num_compte_str.lower() == 'nan':
                            raise ValueError(f"Numéro de compte invalide: '{num_compte_raw}'")
                        ligne['numero_compte'] = num_compte_str
                        ligne['libelle'] = row[1] or ''
                        
                        # Format balance : on vérifie d'abord si les mouvements sont dans le fichier
                        ligne['debit_initial'] = float(row[2] or 0)
                        ligne['credit_initial'] = float(row[3] or 0)
                        
                        # Vérifier si le fichier Excel contient les colonnes de mouvements
                        # Format 1 (avec mouvements) : debit_initial, credit_initial, debit_mvt, credit_mvt, debit_fin, credit_fin
                        # Format 2 (sans mouvements) : debit_initial, credit_initial, debit_fin, credit_fin
                        try:
                            # Essayer de lire les mouvements depuis le fichier (colonnes 4 et 5)
                            if len(row) > 5 and (row[4] is not None or row[5] is not None):
                                ligne['debit_mvt'] = float(row[4] or 0)
                                ligne['credit_mvt'] = float(row[5] or 0)
                                # Utiliser l'ordre détecté pour les colonnes finales
                                if colonnes_inversees_balance_simple and len(row) > 7:
                                    ligne['credit_fin'] = float(row[6] or 0)
                                    ligne['debit_fin'] = float(row[7] or 0)
                                else:
                                    ligne['debit_fin'] = float(row[6] or 0)
                                    ligne['credit_fin'] = float(row[7] or 0)
                            else:
                                # Format sans mouvements : utiliser solde_final = solde_initial + mouvements
                                # Pour la vérification d'identité, on utilise l'équation : 
                                # (debit_fin - credit_fin) = (debit_initial - credit_initial) + (debit_mvt - credit_mvt)
                                ligne['debit_fin'] = float(row[4] or 0)
                                ligne['credit_fin'] = float(row[5] or 0)
                                
                                # Calculer le solde initial et final
                                solde_initial = ligne['debit_initial'] - ligne['credit_initial']
                                solde_final = ligne['debit_fin'] - ligne['credit_fin']
                                mouvement_net = solde_final - solde_initial
                                
                                # Approximer les mouvements débit/crédit à partir du mouvement net
                                # Note: Cette approximation n'est pas parfaite car on ne connaît pas
                                # le détail des mouvements débit/crédit séparément
                                if mouvement_net > 0:
                                    # Le solde a augmenté, donc plus de débits que de crédits
                                    ligne['debit_mvt'] = abs(mouvement_net)
                                    ligne['credit_mvt'] = 0
                                elif mouvement_net < 0:
                                    # Le solde a diminué, donc plus de crédits que de débits
                                    ligne['debit_mvt'] = 0
                                    ligne['credit_mvt'] = abs(mouvement_net)
                                else:
                                    # Pas de changement, mouvements à zéro ou équilibrés
                                    # Approximation : répartir selon la variation des totaux débit/crédit
                                    debit_mvt_calc = max(0, ligne['debit_fin'] - ligne['debit_initial'])
                                    credit_mvt_calc = max(0, ligne['credit_fin'] - ligne['credit_initial'])
                                    ligne['debit_mvt'] = debit_mvt_calc
                                    ligne['credit_mvt'] = credit_mvt_calc
                        except (IndexError, ValueError, TypeError):
                            # Si erreur, utiliser le format simple sans mouvements
                            ligne['debit_mvt'] = 0
                            ligne['credit_mvt'] = 0
                            ligne['debit_fin'] = float(row[4] or 0)
                            ligne['credit_fin'] = float(row[5] or 0)
                        
                        ligne['solde_reel'] = ligne['debit_fin'] - ligne['credit_fin']
                        ligne['solde'] = abs(ligne['solde_reel'])
                        ligne['sign_solde'] = "D" if ligne['debit_fin'] >= ligne['credit_fin'] else "C"
                        
                        data.append(ligne)
                        lignes_traitees += 1
                    except Exception as e:
                        lignes_erreur += 1
                        print(f"⚠️  Erreur ligne {row_idx} (balance simple): {e}")
                        continue
                
                # Afficher le résumé du traitement
                print(f"📊 Résumé du traitement (format balance simple):")
                print(f"   ✅ Lignes traitées avec succès: {lignes_traitees}")
                if lignes_ignorees_colonnes > 0:
                    print(f"   ⚠️  Lignes ignorées (moins de 6 colonnes): {lignes_ignorees_colonnes}")
                if lignes_ignorees_numero_vide > 0:
                    print(f"   ⚠️  Lignes ignorées (numéro de compte vide): {lignes_ignorees_numero_vide}")
                if lignes_ignorees_vides > 0:
                    print(f"   ⚠️  Lignes vides détectées: {lignes_ignorees_vides}")
                if lignes_erreur > 0:
                    print(f"   ❌ Lignes avec erreur: {lignes_erreur}")
                        
            else:
                # Format standard ou autre
                print("📊 Traitement format standard...")
                header_row = 1
                
                # Trouver la ligne d'en-tête
                for row_idx in range(1, min(5, sheet.max_row + 1)):
                    row_data = [sheet.cell(row=row_idx, column=col).value for col in range(1, min(8, sheet.max_column + 1))]
                    if any('compte' in str(cell).lower() for cell in row_data if cell):
                        header_row = row_idx
                        print(f"📝 En-tête détecté à la ligne {row_idx}: {row_data}")
                        break
                
                # DÉTECTION DE L'ORDRE DES COLONNES (une seule fois avant la boucle)
                # Format 1 (standard) : debit_fin, credit_fin (colonnes 7, 8)
                # Format 2 (balance avec solde crédit/débit) : credit_fin, debit_fin (colonnes 7, 8)
                header_row_data = [sheet.cell(row=header_row, column=col).value for col in range(1, min(9, sheet.max_column + 1))]
                
                # Vérifier les colonnes 7 et 8 (index 6 et 7) pour détecter l'ordre
                col7_header = str(header_row_data[6] if len(header_row_data) > 6 else '').strip()
                col8_header = str(header_row_data[7] if len(header_row_data) > 7 else '').strip()
                col7_val = col7_header.lower()
                col8_val = col8_header.lower()
                
                print(f"📝 DÉTECTION DE L'ORDRE DES COLONNES:")
                print(f"   Colonne 7 (index 6): '{col7_header}'")
                print(f"   Colonne 8 (index 7): '{col8_header}'")
                
                # Détecter si "solde crédit" ou "credit" est en colonne 7 et "solde débit" ou "debit" en colonne 8
                # Patterns possibles : "Solde crédit", "Solde Crédit", "SOLDE CREDIT", "Crédit", etc.
                has_solde_credit_col7 = (
                    ('solde' in col7_val and 'credit' in col7_val) or
                    ('solde' in col7_val and 'crédit' in col7_val) or
                    (col7_val and 'credit' in col7_val and 'debit' not in col7_val and 'débit' not in col7_val)
                )
                
                has_solde_debit_col8 = (
                    ('solde' in col8_val and 'debit' in col8_val) or
                    ('solde' in col8_val and 'débit' in col8_val) or
                    (col8_val and 'debit' in col8_val and 'credit' not in col8_val and 'crédit' not in col8_val)
                )
                
                # Détecter aussi le format inverse (débit puis crédit)
                has_debit_col7 = (
                    ('debit' in col7_val and 'credit' not in col7_val and 'crédit' not in col7_val) or
                    ('débit' in col7_val and 'credit' not in col7_val and 'crédit' not in col7_val)
                )
                
                has_credit_col8 = (
                    ('credit' in col8_val and 'debit' not in col8_val and 'débit' not in col8_val) or
                    ('crédit' in col8_val and 'debit' not in col8_val and 'débit' not in col8_val)
                )
                
                # Déterminer l'ordre
                colonnes_inversees = False
                if has_solde_credit_col7 and has_solde_debit_col8:
                    colonnes_inversees = True
                    print(f"✅ Format détecté: Solde crédit (col 7) puis Solde débit (col 8) - COLONNES INVERSÉES")
                    print(f"   Mapping: credit_fin = row[6] (col 7), debit_fin = row[7] (col 8)")
                elif has_debit_col7 and has_credit_col8:
                    colonnes_inversees = False
                    print(f"✅ Format détecté: Débit (col 7) puis Crédit (col 8) - FORMAT STANDARD")
                    print(f"   Mapping: debit_fin = row[6] (col 7), credit_fin = row[7] (col 8)")
                else:
                    # Par défaut, utiliser le format standard
                    colonnes_inversees = False
                    print(f"⚠️  Format non reconnu, utilisation du format standard par défaut")
                    print(f"   Mapping: debit_fin = row[6] (col 7), credit_fin = row[7] (col 8)")
                
                # Compteurs pour le diagnostic
                lignes_traitees = 0
                lignes_ignorees_colonnes = 0
                lignes_ignorees_vides = 0
                lignes_ignorees_numero_vide = 0
                lignes_erreur = 0
                
                # Traitement des lignes de données
                for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                    # Vérifier que la ligne a au moins 8 colonnes
                    if len(row) < 8:
                        lignes_ignorees_colonnes += 1
                        continue
                        
                    if row[0] is None and row[1] is None:
                        lignes_ignorees_vides += 1
                        break  # Fin des données
                        
                    # Vérifier que la ligne contient des données valides
                    if not row[0] or str(row[0]).strip() == '':
                        lignes_ignorees_numero_vide += 1
                        continue
                    
                    try:
                        ligne = {}
                        ligne['numero_compte'] = str(row[0])
                        ligne['libelle'] = row[1] or ''
                        ligne['debit_initial'] = float(row[2] or 0)
                        ligne['credit_initial'] = float(row[3] or 0)
                        ligne['debit_mvt'] = float(row[4] or 0)
                        ligne['credit_mvt'] = float(row[5] or 0)
                        
                        # Utiliser l'ordre détecté précédemment
                        if colonnes_inversees:
                            # Format avec "Solde crédit" (col 7) puis "Solde débit" (col 8)
                            ligne['credit_fin'] = float(row[6] or 0)
                            ligne['debit_fin'] = float(row[7] or 0)
                        else:
                            # Format standard avec "Débit fin" puis "Crédit fin"
                            ligne['debit_fin'] = float(row[6] or 0)
                            ligne['credit_fin'] = float(row[7] or 0)
                        
                        ligne['solde_reel'] = ligne['debit_fin'] - ligne['credit_fin']
                        ligne['solde'] = abs(ligne['solde_reel'])
                        ligne['sign_solde'] = "D" if ligne['debit_fin'] >= ligne['credit_fin'] else "C"
                        data.append(ligne)
                        lignes_traitees += 1
                    except Exception as e:
                        lignes_erreur += 1
                        print(f"⚠️  Erreur ligne {row_idx} (standard): {e}")
                        continue
                
                # Afficher le résumé du traitement
                print(f"📊 Résumé du traitement:")
                print(f"   ✅ Lignes traitées avec succès: {lignes_traitees}")
                if lignes_ignorees_colonnes > 0:
                    print(f"   ⚠️  Lignes ignorées (moins de 8 colonnes): {lignes_ignorees_colonnes}")
                if lignes_ignorees_numero_vide > 0:
                    print(f"   ⚠️  Lignes ignorées (numéro de compte vide): {lignes_ignorees_numero_vide}")
                if lignes_ignorees_vides > 0:
                    print(f"   ⚠️  Lignes vides détectées: {lignes_ignorees_vides}")
                if lignes_erreur > 0:
                    print(f"   ❌ Lignes avec erreur: {lignes_erreur}")
            
            print(f"✅ {len(data)} lignes de données extraites de la feuille '{sheet_name}'")
            
            # Vérifier qu'au moins une ligne de données a été extraite
            if len(data) == 0:
                filename = balance.filename if hasattr(balance, 'filename') else 'fichier inconnu'
                format_info = ""
                if format_detecte == 'sage':
                    format_info = "Format Sage détecté: attendu 6 colonnes minimum (numero, libelle, solde_n1, mouvement_debit, mouvement_credit, solde_n)"
                elif format_detecte == 'balance_simple':
                    format_info = "Format balance simple détecté: attendu 6 colonnes (numero, libelle, debit_initial, credit_initial, debit_fin, credit_fin)"
                else:
                    format_info = "Format standard détecté: attendu 8 colonnes minimum (numero, libelle, debit_initial, credit_initial, debit_mvt, credit_mvt, debit_fin, credit_fin)"
                
                error_msg = (
                    f"Aucune donnée extraite du fichier '{filename}' (feuille '{sheet_name}'). "
                    f"Vérifiez que:\n"
                    f"- La feuille contient bien des données (au moins une ligne avec un numéro de compte)\n"
                    f"- {format_info}\n"
                    f"- Les numéros de compte ne sont pas vides et commencent par un chiffre\n"
                    f"- Les lignes vides ou sans numéro de compte sont ignorées"
                )
                print(f"❌ {error_msg}")
                raise Exception(error_msg)
                    
        except Exception as e:
            print(f"Erreur lors du traitement du fichier Excel: {str(e)}")
            raise e

        # Double vérification avant insertion
        if len(data) == 0:
            filename = balance.filename if hasattr(balance, 'filename') else 'fichier inconnu'
            raise Exception(f"Aucune donnée valide trouvée dans le fichier '{filename}'. Vérifiez le format du fichier Excel.")

        result = db.Balance.insert_one(
            {"id_client": id_client, "annee_balance": annee_auditee, "balance": data}
        )

        inserted_id = str(result.inserted_id)
        print(f"✅ Balance insérée en base avec {len(data)} lignes (ID: {inserted_id})")
        return inserted_id, data

    # ---------- Grouping ----------
    def create_grouping(self, balances_rapprochee, referentiel="syscohada"):
        try:
            grouping_path = os.path.join(os.path.dirname(__file__), "..", "grouping.json")
            print(f"📁 Chemin du fichier grouping.json: {grouping_path}")
            print(f"📁 Chemin absolu: {os.path.abspath(grouping_path)}")
            print(f"📁 Fichier existe? {os.path.exists(grouping_path)}")
            
            if not os.path.exists(grouping_path):
                raise FileNotFoundError(f"Le fichier grouping.json n'existe pas à l'emplacement: {grouping_path}")
            
            with open(grouping_path, 'r', encoding='utf-8') as file:
                result = json.load(file)
        except FileNotFoundError as e:
            print(f"❌ Erreur fichier non trouvé: {str(e)}")
            raise
        except json.JSONDecodeError as e:
            print(f"❌ Erreur de décodage JSON: {str(e)}")
            raise
        except Exception as e:
            print(f"❌ Erreur lors de l'ouverture du fichier grouping.json: {str(e)}")
            raise
        
        # Vérifier que le référentiel existe dans le fichier JSON
        if referentiel not in result:
            available_keys = list(result.keys())
            raise KeyError(f"Le référentiel '{referentiel}' n'existe pas dans grouping.json. Référentiels disponibles: {available_keys}")
        
        table_grouping = result[referentiel]
        print(f"✅ Référentiel '{referentiel}' chargé: {len(table_grouping)} groupes définis")

        # Vérifier que balances_rapprochee est une liste
        if not isinstance(balances_rapprochee, list):
            raise TypeError(f"balances_rapprochee doit être une liste, mais c'est un {type(balances_rapprochee)}")
        
        print(f"📊 Nombre de comptes dans balance_variation: {len(balances_rapprochee)}")
        
        # Créer un dictionnaire des groupes existants par préfixe
        groupes_existants = {group['compte']: group for group in table_grouping}
        
        # Identifier tous les préfixes de comptes dans les balances (2 premiers chiffres)
        prefixes_trouves = set()
        for item in balances_rapprochee:
            numero_compte = str(item.get('numero_compte', '')).strip()
            if numero_compte and len(numero_compte) >= 2:
                prefixe = numero_compte[:2]
                if prefixe.isdigit():
                    prefixes_trouves.add(prefixe)

        # Créer des groupes automatiques pour les préfixes non définis
        prefixes_manquants = prefixes_trouves - set(groupes_existants.keys())
        for prefixe in sorted(prefixes_manquants):
            # Déterminer la nature basée sur le préfixe
            if prefixe.startswith(('1', '2', '3', '4', '5')):
                nature = "bilan"
            else:
                nature = "pnl"
            
            nouveau_groupe = {
                "compte": prefixe,
                "nature": nature,
                "libelle": f"AUTRES - COMPTE {prefixe}"
            }
            table_grouping.append(nouveau_groupe)
            groupes_existants[prefixe] = nouveau_groupe

        # Traiter tous les groupes (définis + auto-créés)
        for group in table_grouping:
            nber_group = group['compte']
            group['solde_n'] = sum(
                item['solde_n'] for item in balances_rapprochee
                if str(item['numero_compte']).startswith(nber_group)
            )
            group['solde_n1'] = sum(
                item['solde_n1'] for item in balances_rapprochee
                if str(item['numero_compte']).startswith(nber_group)
            )
            group['variation'] = group['solde_n'] - group['solde_n1']

            if group['variation'] == 0:
                group['variation_percent'] = 0
            elif group['solde_n1'] == 0:
                group['variation_percent'] = 100
            else:
                group['variation_percent'] = (group['variation'] / group['solde_n1']) * 100

            # Ajouter le détail des comptes par groupe avec N et N-1
            # IMPORTANT: Tous les comptes doivent être regroupés par groupes
            comptes_du_groupe = []
            for item in balances_rapprochee:
                numero_compte_str = str(item.get('numero_compte', '')).strip()
                # Vérifier si le compte commence par le préfixe du groupe
                if numero_compte_str.startswith(str(nber_group)):
                    compte_detail = {
                        "numero_compte": numero_compte_str,
                        "libelle": item.get('libelle', ''),
                        "solde_n": item.get('solde_n', 0),
                        "solde_n1": item.get('solde_n1', 0),
                        "variation": item.get('solde_n', 0) - item.get('solde_n1', 0)
                    }
                    comptes_du_groupe.append(compte_detail)
            # Tri pour une lecture plus simple
            comptes_du_groupe.sort(key=lambda x: x["numero_compte"])
            # Toujours assigner la liste des comptes, même si elle est vide
            group['comptes'] = comptes_du_groupe
            group['comptes_detaille'] = comptes_du_groupe  # Alias pour compatibilité
            
            # Log pour vérifier le regroupement
            if len(comptes_du_groupe) > 0:
                print(f"✅ Groupe {nber_group} ({group.get('libelle', 'N/A')}): {len(comptes_du_groupe)} compte(s) regroupé(s)")
                # Afficher les premiers comptes pour vérification
                premiers_comptes = [c['numero_compte'] for c in comptes_du_groupe[:3]]
                print(f"   Exemples de comptes: {premiers_comptes}")
            else:
                print(f"⚠️ Groupe {nber_group} ({group.get('libelle', 'N/A')}): Aucun compte trouvé")
                # Debug: vérifier pourquoi aucun compte n'est trouvé
                comptes_test = [str(item.get('numero_compte', '')).strip() for item in balances_rapprochee[:10]]
                print(f"   Debug: Exemples de numéros de compte dans balance_variation: {comptes_test}")
                print(f"   Debug: Préfixe recherché: '{nber_group}' (type: {type(nber_group)})")

        # Vérifier que tous les comptes sont bien regroupés
        comptes_regroupes = set()
        for group in table_grouping:
            for compte in group.get('comptes', []):
                comptes_regroupes.add(compte.get('numero_compte', ''))
        
        # Identifier les comptes non regroupés
        tous_les_comptes = {str(item.get('numero_compte', '')) for item in balances_rapprochee}
        comptes_non_regroupes = tous_les_comptes - comptes_regroupes
        
        if comptes_non_regroupes:
            print(f"⚠️ {len(comptes_non_regroupes)} compte(s) non regroupé(s): {list(comptes_non_regroupes)[:10]}")
            # Essayer de regrouper les comptes manquants
            for compte_num in comptes_non_regroupes:
                if compte_num and len(compte_num) >= 2:
                    prefixe = compte_num[:2]
                    if prefixe.isdigit():
                        # Trouver le groupe correspondant
                        for group in table_grouping:
                            if group['compte'] == prefixe:
                                # Ajouter le compte manquant au groupe
                                compte_item = next((item for item in balances_rapprochee if str(item.get('numero_compte', '')) == compte_num), None)
                                if compte_item:
                                    compte_detail = {
                                        "numero_compte": str(compte_item.get('numero_compte', '')),
                                        "libelle": compte_item.get('libelle', ''),
                                        "solde_n": compte_item.get('solde_n', 0),
                                        "solde_n1": compte_item.get('solde_n1', 0),
                                        "variation": compte_item.get('solde_n', 0) - compte_item.get('solde_n1', 0)
                                    }
                                    group['comptes'].append(compte_detail)
                                    print(f"✅ Compte {compte_num} ajouté au groupe {prefixe}")
        
        # Trier par numéro de compte
        table_grouping.sort(key=lambda x: x['compte'])
        
        # Trier les comptes dans chaque groupe
        for group in table_grouping:
            if 'comptes' in group:
                group['comptes'].sort(key=lambda x: x.get('numero_compte', ''))
        
        print(f"📊 Grouping créé: {len(table_grouping)} groupes, {len(comptes_regroupes)} comptes regroupés sur {len(tous_les_comptes)} comptes totaux")
        
        return table_grouping

    # ---------- Variation N vs N-1 ----------
    def rapprochement_des_balances(self, balance_n, balance_n1):
        variation_des_balances = []
        idx_n1 = {str(item.get('numero_compte')): item for item in balance_n1}

        for bal in balance_n:
            numero = str(bal['numero_compte'])
            ligne = {
                'numero_compte': numero,
                'libelle': bal['libelle'],
                'solde_n': bal['solde_reel'],
                'solde_n1': idx_n1.get(numero, {}).get('solde_reel', 0)
            }
            variation_des_balances.append(ligne)

        return variation_des_balances

    # ---------- Charges ----------
    def total_charges(self, id_mission):
        balance = db.Mission1.find_one({"_id": ObjectId(id_mission)})['balance_variation']
        charges = sum(item['solde_n'] for item in balance if str(item['numero_compte']).startswith('6'))
        return abs(charges)

    # ---------- Benchmarks ----------
    def get_benchmarks(self, id_mission):
        try:
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            efi = mission['efi']

            bench = {}
            bench['total_assets'] = int(efi['actif'][-1]['net_solde_n'])
            bench['profit_before_tax'] = int(efi['pnl'][33]['net_solde_n']) + int(efi['pnl'][38]['net_solde_n'])
            bench['revenue'] = int(efi['pnl'][7]['net_solde_n'])
            bench['ebitda'] = int(efi['pnl'][23]['net_solde_n'])
            bench['expenses'] = self.total_charges(id_mission)
            return bench
        except Exception as e:
            print(f"An error there: {str(e)}")
            return None

    # ---------- Materiality ----------
    def save_materiality(self, id_mission, materialities):
        query = {"_id": ObjectId(id_mission)}
        verify = db.Mission1.find_one(query)

        if not verify:
            return 0

        # Vérifier si le benchmark existe déjà
        existing_materiality = verify.get('materiality', [])
        benchmark_exists = any(item.get('benchmark') == materialities['benchmark'] for item in existing_materiality)
        
        # Forcer des valeurs toujours positives pour les seuils (règle métier)
        try:
            materialities['materiality'] = abs(int(materialities.get('materiality', 0) or 0))
            materialities['performance_materiality'] = abs(int(materialities.get('performance_materiality', 0) or 0))
            materialities['trivial_misstatements'] = abs(int(materialities.get('trivial_misstatements', 0) or 0))
        except Exception:
            pass

        if benchmark_exists:
            # Mettre à jour le benchmark existant
            update = db.Mission1.update_one(
                {"_id": ObjectId(id_mission), "materiality.benchmark": materialities['benchmark']},
                {"$set": {"materiality.$": materialities}}
            )
            return update.modified_count
        else:
            # Ajouter un nouveau benchmark
            update = db.Mission1.update_one(query, {"$push": {"materiality": materialities}})
            return update.modified_count

    def validate_materiality(self, id_mission, benchmark):
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        materialities = mission['materiality']

        exist_choice = next((item for item in materialities if item.get('choice')), None)
        if not exist_choice:
            for item in materialities:
                if item['benchmark'] == benchmark:
                    item['choice'] = True
        else:
            exist_choice['choice'] = ""
            for item in materialities:
                if item['benchmark'] == benchmark:
                    item['choice'] = True

        result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"materiality": materialities}})
        return result.modified_count

    def get_materiality(self, id_mission):
        materiality = db.Mission1.find_one({"_id": ObjectId(id_mission)}, {"_id": 0, "materiality": 1})
        return materiality

    def materialite(self, id_mission):
        """
        Méthode pour récupérer et calculer les données de matérialité
        """
        try:
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "materiality": []}

            # Récupérer les données de matérialité existantes
            materiality = mission.get("materiality", [])
            
            # Si aucune donnée de matérialité, calculer les benchmarks
            if not materiality:
                print(f"📊 Calcul des benchmarks de matérialité pour la mission {id_mission}")
                materiality = self._calculate_materiality_benchmarks(mission)
                
                # Sauvegarder les calculs si des benchmarks ont été trouvés
                if materiality:
                    db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"materiality": materiality}})
                    print(f"✅ {len(materiality)} benchmark(s) sauvegardé(s) dans la base de données")

            if not materiality:
                # Vérifier la cause du problème pour donner un message plus explicite
                efi = mission.get("efi", {})
                if not efi:
                    message = "Impossible de calculer les seuils de matérialité. Les états financiers préliminaires n'ont pas été générés. Veuillez d'abord compléter l'étape 5 (États financiers préliminaires)."
                elif not isinstance(efi, dict) or not any(section in efi for section in ["actif", "passif", "pnl"]):
                    message = "Impossible de calculer les seuils de matérialité. Les états financiers sont incomplets ou mal formatés. Veuillez régénérer les états financiers préliminaires."
                else:
                    message = "Impossible de calculer les seuils de matérialité. Les données financières nécessaires (profit_before_tax, ebitda, revenue, total_assets, total_expenses) n'ont pas été trouvées dans les états financiers. Vérifiez que les balances contiennent les informations nécessaires."
                
                return {"ok": False, "message": message, "materiality": []}

            return {
                "ok": True,
                "message": "Données de matérialité récupérées avec succès",
                "materiality": materiality
            }

        except Exception as e:
            import traceback
            print(f"❌ Erreur dans materialite: {e}")
            traceback.print_exc()
            return {"ok": False, "message": f"Erreur lors du calcul de la matérialité: {str(e)}", "materiality": []}

    def _calculate_materiality_benchmarks(self, mission):
        """
        Calcule les benchmarks de matérialité basés sur les données financières
        """
        try:
            # Récupérer les données financières
            efi = mission.get("efi", {})
            if not efi:
                print("❌ Aucune donnée EFI trouvée dans la mission")
                print("💡 Conseil: Veuillez générer les états financiers préliminaires (Étape 5) avant de calculer les matérialités")
                return []

            # Vérifier que les données EFI sont structurées correctement
            if not isinstance(efi, dict):
                print(f"❌ Format de données EFI invalide: {type(efi)}")
                return []
            
            # Vérifier que les sections principales existent
            sections_presentes = [section for section in ["actif", "passif", "pnl"] if section in efi and isinstance(efi[section], list)]
            if not sections_presentes:
                print("❌ Aucune section EFI valide trouvée (actif, passif, pnl)")
                return []
            
            print(f"✅ Sections EFI trouvées: {sections_presentes}")
            
            benchmarks = []
            valeurs_trouvees = []
            
            # Benchmark 1: Profit Before Tax
            profit_before_tax = self._get_financial_value(efi, "profit_before_tax")
            if profit_before_tax and profit_before_tax != 0:
                # Multiplier par -1 car les débits sont en + et les crédits en - dans la Balance Générale
                profit_before_tax = profit_before_tax * -1
                materiality = profit_before_tax * 0.05  # 5% du bénéfice avant impôt
                # Le seuil de matérialité est toujours positif (valeur absolue)
                abs_materiality = abs(materiality)
                is_negative = materiality < 0
                
                benchmarks.append({
                    "benchmark": "profit_before_tax",
                    "factor": "5%",
                    "materiality": int(abs_materiality),  # Toujours positif
                    "performance_materiality": int(abs_materiality * 0.8),
                    "trivial_misstatements": int(abs_materiality * 0.05),
                    "choice": "",
                    "warning": "⚠️ ATTENTION: Seuil de matérialité négatif !" if is_negative else "",
                    "original_value": profit_before_tax
                })
                valeurs_trouvees.append(f"profit_before_tax: {profit_before_tax}")
            else:
                print(f"⚠️ Valeur profit_before_tax non trouvée ou nulle: {profit_before_tax}")

            # Benchmark 2: EBITDA
            ebitda = self._get_financial_value(efi, "ebitda")
            if ebitda and ebitda != 0:
                # Multiplier par -1 car les débits sont en + et les crédits en - dans la Balance Générale
                ebitda = ebitda * -1
                materiality = ebitda * 0.05  # 5% de l'EBITDA
                # Le seuil de matérialité est toujours positif (valeur absolue)
                abs_materiality = abs(materiality)
                is_negative = materiality < 0
                
                benchmarks.append({
                    "benchmark": "ebitda",
                    "factor": "5%",
                    "materiality": int(abs_materiality),  # Toujours positif
                    "performance_materiality": int(abs_materiality * 0.8),
                    "trivial_misstatements": int(abs_materiality * 0.05),
                    "choice": "",
                    "warning": "⚠️ ATTENTION: Seuil de matérialité négatif !" if is_negative else "",
                    "original_value": ebitda
                })
                valeurs_trouvees.append(f"ebitda: {ebitda}")
            else:
                print(f"⚠️ Valeur ebitda non trouvée ou nulle: {ebitda}")

            # Benchmark 3: Revenue
            revenue = self._get_financial_value(efi, "revenue")
            if revenue and revenue != 0:
                # Multiplier par -1 car les débits sont en + et les crédits en - dans la Balance Générale
                revenue = revenue * -1
                materiality = revenue * 0.01  # 1% du chiffre d'affaires
                # Le seuil de matérialité est toujours positif (valeur absolue)
                abs_materiality = abs(materiality)
                is_negative = materiality < 0
                
                benchmarks.append({
                    "benchmark": "revenue",
                    "factor": "1%",
                    "materiality": int(abs_materiality),  # Toujours positif
                    "performance_materiality": int(abs_materiality * 0.8),
                    "trivial_misstatements": int(abs_materiality * 0.05),
                    "choice": "",
                    "warning": "⚠️ ATTENTION: Seuil de matérialité négatif !" if is_negative else "",
                    "original_value": revenue
                })
                valeurs_trouvees.append(f"revenue: {revenue}")
            else:
                print(f"⚠️ Valeur revenue non trouvée ou nulle: {revenue}")

            # Benchmark 4: Total Assets
            total_assets = self._get_financial_value(efi, "total_assets")
            if total_assets and total_assets > 0:  # Les actifs ne peuvent pas être négatifs
                # Multiplier par -1 car les débits sont en + et les crédits en - dans la Balance Générale
                total_assets = total_assets * -1
                materiality = total_assets * 0.01  # 1% du total des actifs
                benchmarks.append({
                    "benchmark": "total_assets",
                    "factor": "1%",
                    "materiality": max(int(materiality), 0),
                    "performance_materiality": max(int(materiality * 0.8), 0),
                    "trivial_misstatements": max(int(materiality * 0.05), 0),
                    "choice": "",
                    "warning": "",
                    "original_value": total_assets
                })
                valeurs_trouvees.append(f"total_assets: {total_assets}")
            else:
                print(f"⚠️ Valeur total_assets non trouvée ou nulle: {total_assets}")

            # Benchmark 5: Total Expenses (Charges)
            total_expenses = self._get_financial_value(efi, "total_expenses")
            if total_expenses and total_expenses > 0:  # Les charges doivent toujours être positives
                # Multiplier par -1 car les débits sont en + et les crédits en - dans la Balance Générale
                total_expenses = total_expenses * -1
                materiality = total_expenses * 0.02  # 2% du total des charges
                benchmarks.append({
                    "benchmark": "total_expenses",
                    "factor": "2%",
                    "materiality": int(materiality),  # Toujours positif
                    "performance_materiality": int(materiality * 0.8),
                    "trivial_misstatements": int(materiality * 0.05),
                    "choice": "",
                    "warning": "",
                    "original_value": total_expenses
                })
                valeurs_trouvees.append(f"total_expenses: {total_expenses}")
            else:
                print(f"⚠️ Valeur total_expenses non trouvée ou nulle: {total_expenses}")

            if benchmarks:
                print(f"✅ {len(benchmarks)} benchmark(s) calculé(s) avec succès: {valeurs_trouvees}")
            else:
                print("❌ Aucun benchmark n'a pu être calculé. Vérifiez que les états financiers contiennent les données nécessaires.")
                print("💡 Les valeurs recherchées sont: profit_before_tax, ebitda, revenue, total_assets, total_expenses")

            return benchmarks

        except Exception as e:
            import traceback
            print(f"❌ Erreur lors du calcul des benchmarks: {e}")
            traceback.print_exc()
            return []

    def _get_financial_value(self, efi, key):
        """
        Récupère une valeur financière depuis les états financiers
        """
        try:
            # Mapping des clés vers les libellés français ou calculs
            key_mappings = {
                "profit_before_tax": ["bénéfice avant impôt", "résultat avant impôt", "profit", "bénéfice"],
                "ebitda": ["ebitda", "excédent brut d'exploitation", "ebe"],
                "revenue": ["chiffre d'affaires", "revenus", "recettes", "ventes", "ca"],
                "total_assets": ["total actif", "total de l'actif"],
                "total_expenses": ["total charges", "charges totales", "total des charges"]
            }
            
            # Chercher dans les différents états financiers
            for section in ["actif", "passif", "pnl"]:
                if section in efi and isinstance(efi[section], list):
                    for item in efi[section]:
                        ref = item.get("ref", "").lower()
                        libelle = item.get("libelle", "").lower()
                        net_solde = item.get("net_solde_n", 0)
                        
                        # Vérifier si la ref correspond
                        if ref == key or ref == key.replace("_", ""):
                            # Retourner la valeur brute (avec son signe) pour pouvoir multiplier par -1
                            return net_solde if net_solde else None
                        
                        # Vérifier si le libellé correspond aux mappings
                        if key in key_mappings:
                            for search_term in key_mappings[key]:
                                if search_term in libelle:
                                    # Retourner la valeur brute (avec son signe) pour pouvoir multiplier par -1
                                    return net_solde if net_solde else None
            
            # Calculs spécifiques si les valeurs directes ne sont pas trouvées
            if key == "total_assets" and "actif" in efi:
                # Calculer le total actif en sommant tous les actifs (avec leur signe)
                # Les actifs peuvent être négatifs dans la balance, donc on garde le signe
                total = sum(item.get("net_solde_n", 0) for item in efi.get("actif", []) if item.get("net_solde_n", 0))
                return total if total != 0 else None
                
            if key == "total_expenses" and "pnl" in efi:
                # Chercher les charges dans le compte de résultat (avec leur signe)
                charges_keywords = ["charge", "dépense", "coût"]
                total = 0
                for item in efi.get("pnl", []):
                    libelle = item.get("libelle", "").lower()
                    if any(kw in libelle for kw in charges_keywords):
                        solde = item.get("net_solde_n", 0)
                        # Garder le signe pour pouvoir multiplier par -1 ensuite
                        total += solde
                return abs(total) if total != 0 else None
            
            return None
        except Exception as e:
            print(f"Erreur dans _get_financial_value pour {key}: {e}")
            return None

    # ---------- Analyses grouping ----------
    def make_quantitative_analysis(self, id_mission):
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        grouping = mission['grouping']
        materiality = next((mat for mat in mission['materiality'] if mat.get('choice')), None)

        if materiality is not None:
            for item in grouping:
                value = False
                if int(item['solde_n']) >= int(materiality['materiality']):
                    value = True
                item['materiality'] = value

            result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})
            return result.modified_count
        else:
            return 0

    def make_qualitative_analysis(self, id_mission, significant):
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        grouping = mission['grouping']

        for item in grouping:
            value = next((elt['significant'] for elt in significant if elt.get('compte') == item.get('compte')), None)
            if value is None:
                item['significant'] = False
            else:
                item['significant'] = value

        result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})
        return result.modified_count

    def make_final_sm(self, id_mission):
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        grouping = mission['grouping']

        for item in grouping:
            mat = item.get('materiality', None)
            sign = item.get('significant', None)

            if (mat is not None) and (sign is not None):
                value = ""
                if mat is False and sign is True:
                    value = "non matériel significatif"
                elif mat is False and sign is False:
                    value = "non matériel non significatif"
                elif mat is True and sign is True:
                    value = "matériel significatif"
                elif mat is True and sign is False:
                    value = "matériel non significatif"
                else:
                    value = None
                item['mat_sign'] = value

        result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})
        return result.modified_count, grouping

    # ---------- Récup mission ----------
    def afficher_informations_missions(self, id_client):
        _id = id_client
        query = db.Mission1.find_one({"_id": _id})
        query['_id'] = str(query['_id'])
        return query

    # ---------- Production COMMENTAIRE ----------
    def prod_efi(self, balance_n, balance_n1, balance_variation):
        mapping_path = os.path.join(os.path.dirname(__file__), "..", "mapping_efi.json")
        with open(mapping_path, 'r', encoding='utf-8') as file:
            result = json.load(file)
        mapping = result['structure']

        # nettoyer les champs *_cpt en listes
        for mapp in mapping:
            mapp['brut_cpt'] = mapp['brut_cpt'].split(',') if mapp.get('brut_cpt') is not None else mapp.get('brut_cpt')
            mapp['amor_cpt'] = mapp['amor_cpt'].split(',') if mapp.get('amor_cpt') is not None else mapp.get('amor_cpt')
            mapp['net_cpt'] = mapp['net_cpt'].split(',') if mapp.get('net_cpt') is not None else mapp.get('net_cpt')
            mapp['brut_except_cpt'] = mapp['brut_except_cpt'].split(',') if mapp.get('brut_except_cpt') is not None else mapp.get('brut_except_cpt', [])
            mapp['amor_except_cpt'] = mapp['amor_except_cpt'].split(',') if mapp.get('amor_except_cpt') is not None else mapp.get('amor_except_cpt', [])
            mapp['net_except_cpt'] = mapp['net_except_cpt'].split(',') if mapp.get('net_except_cpt') is not None else mapp.get('net_except_cpt', [])

        datum = {}
        list_efi = ['actif', 'passif', 'pnl']
        for efi in list_efi:
            structure = []
            select_mapping = (elt for elt in mapping if elt['nature'] == efi)
            for data in select_mapping:
                row = {}
                if data.get('brut_cpt') and data.get('amor_cpt'):
                    # N
                    brut_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['brut_cpt']))
                    amor_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['amor_cpt']))

                    brut_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('brut_except_cpt', [])))
                    amor_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('amor_except_cpt', [])))

                    data['brut_solde_n'] = brut_solde_n + brut_except_n
                    data['amor_solde_n'] = amor_solde_n + amor_except_n
                    data['net_solde_n'] = data['brut_solde_n'] + data['amor_solde_n']

                    # N-1
                    brut_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['brut_cpt']))
                    amor_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['amor_cpt']))
                    net_except_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))
                    data['net_solde_n1'] = brut_n1 + amor_n1 + net_except_n1
                else:
                    net_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['net_cpt']))
                    net_solde_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['net_cpt']))

                    net_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))
                    net_except_n1_bis = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))

                    data['net_solde_n'] = net_solde_n + net_except_n
                    data['net_solde_n1'] = net_solde_n1 + net_except_n1_bis

                row['ref'] = data['ref']
                row['libelle'] = data['libelle']
                row['compte_to_be_used'] = str(data.get('brut_cpt')) + str(data.get('amor_cpt')) + str(data.get('net_cpt')) + str(data.get('brut_except_cpt')) + str(data.get('amor_except_cpt')) + str(data.get('net_except_cpt'))
                row['compte_to_be_used'] = row['compte_to_be_used'].replace('None', '')

                one = data.get('brut_cpt', []) or []
                two = data.get('amor_cpt', []) or []
                three = data.get('net_cpt', []) or []
                four = data.get('brut_except_cpt', []) or []
                five = data.get('amor_except_cpt', []) or []
                six = data.get('net_except_cpt', []) or []

                row['compte_to_be_used_off'] = list(set(one + two + three + four + five + six))

                row['brut_solde_n'] = data.get('brut_solde_n')
                row['amor_solde_n'] = data.get('amor_solde_n')
                row['net_solde_n'] = data.get('net_solde_n')
                row['net_solde_n1'] = data.get('net_solde_n1')

                structure.append(row)

            datum[efi] = structure

        return datum

    # ---------- Piste d'audit ----------
    def audit_trail(self, id_mission):
        # Créer un fichier Excel pour la piste d'audit
        wb = openpyxl.Workbook()
        sheet = wb.active

        columns = ['A', 'B', 'C', 'D', 'E']
        headers = ['Numéro compte', 'Solde n', 'Solde n-1', 'Grouping', 'Code COMMENTAIRE']

        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        balances = mission['balance_variation']
        actif = mission['efi']['actif']
        passif = mission['efi']['passif']
        pnl = mission['efi']['pnl']

        efi = actif + passif + pnl

        for i in range(len(columns)):
            sheet[columns[i] + '1'] = headers[i]

        for iteration, data in enumerate(balances):
            new_iteration = str(iteration + 2)
            sheet["A" + new_iteration] = data.get("numero_compte")
            sheet["B" + new_iteration] = data.get("solde_n")
            sheet["C" + new_iteration] = data.get("solde_n1")
            sheet["D" + new_iteration] = data.get("numero_compte")[0:2]

            list_code_efi = []
            for obj in efi:
                for elt in obj['compte_to_be_used_off']:
                    if data['numero_compte'].startswith(elt):
                        list_code_efi.append(obj['ref'])
            list_code_efi = list(set(list_code_efi))
            sheet["E" + new_iteration] = ','.join(list_code_efi)

        namefile = "piste_audit.xlsx"
        wb.save(namefile)

    # ---------- Extract grouping Excel ----------
    def extract_grouping(self, id_mission):
        # Créer un fichier Excel pour l'export du grouping
        wb = openpyxl.Workbook()
        sheet = wb.active

        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        headers = ['Numéro compte', 'Solde n', 'Solde n-1', 'Grouping', 'Variation', 'Variation %', 'Compte qualitatif', 'Compte quantitatif', 'Compte significatif']

        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        balances = mission['balance_variation']
        grouping = mission['grouping']
        materiality = next(item for item in mission['materiality'] if item['choice'] is True)

        for i in range(len(columns)):
            sheet[columns[i] + '1'] = headers[i]

        for iteration, data in enumerate(balances):
            new_iteration = str(iteration + 2)
            sheet["A" + new_iteration] = data.get("numero_compte")
            sheet["B" + new_iteration] = data.get("solde_n")
            sheet["C" + new_iteration] = data.get("solde_n1")

            value_grouping = data.get("numero_compte")[0:2]
            variation = data.get("solde_n") - data.get("solde_n1")

            if variation == 0:
                variation_percent = 0
            elif data.get("solde_n1") == 0:
                variation_percent = 100
            else:
                variation_percent = (variation / data.get("solde_n1")) * 100

            sheet["D" + new_iteration] = value_grouping
            sheet["E" + new_iteration] = variation
            sheet["F" + new_iteration] = variation_percent
            sheet["G" + new_iteration] = next(item['significant'] for item in grouping if item['compte'] == value_grouping)
            sheet["H" + new_iteration] = next(item['materiality'] for item in grouping if item['compte'] == value_grouping)
            sheet["I" + new_iteration] = next(item['mat_sign'] for item in grouping if item['compte'] == value_grouping)

        second_sheet = wb.create_sheet(title="Seuil de matérialité")
        second_headers = ['materiality', 'performance materiality', 'trivial misstatements']
        second_sheet["A1"] = second_headers[0]
        second_sheet["B1"] = second_headers[1]
        second_sheet["C1"] = second_headers[2]

        second_sheet["A2"] = materiality['materiality']
        second_sheet["B2"] = materiality['performance_materiality']
        second_sheet["C2"] = materiality['trivial_misstatements']

        excel_io = BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        return excel_io

    # ==============================
    #  CONTROLES — Cohérence & Intangibilité
    # ==============================
    def _load_balance(self, balance_id):
        bal = db.Balance.find_one({"_id": ObjectId(balance_id)})
        return bal["balance"] if bal else []

    def _get_sens_attendu_compte(self, numero_compte):
        """
        Détermine le sens attendu (Débit/Crédit) d'un compte selon le plan comptable SYSCOHADA
        Retourne: ('D', 'C', ou 'BOTH' pour les deux sens possibles)
        """
        if not numero_compte or len(numero_compte) < 2:
            return 'BOTH'
        
        prefixe = numero_compte[:2]
        classe = numero_compte[0] if numero_compte[0].isdigit() else None
        
        # Classe 1 - Capitaux : CRÉDITEUR (sens normal)
        if prefixe.startswith('10') or prefixe.startswith('11') or prefixe.startswith('12') or \
           prefixe.startswith('13') or prefixe.startswith('14') or prefixe.startswith('15') or \
           prefixe.startswith('16') or prefixe.startswith('17') or prefixe.startswith('18') or \
           prefixe.startswith('19'):
            # Exceptions : 
            # - 105 : peut être débiteur ou créditeur
            # - 12 (Report à nouveau) : peut être débiteur ou créditeur selon le résultat
            # - 109 (Associés non appelé) : peut être débiteur
            # - 129 (Report à nouveau débiteur) : peut être débiteur
            if numero_compte.startswith('105') or numero_compte.startswith('12') or numero_compte.startswith('109') or numero_compte.startswith('129'):
                return 'BOTH'  # Peut être débiteur ou créditeur selon le cas
            return 'C'  # Créditeur normal pour la classe 1
        
        # Classe 2 - Immobilisations : DÉBITEUR (sens normal)
        # Exceptions : 
        # - 28 (Amortissements) : CRÉDITEUR
        # - 29 (Dépréciations) : CRÉDITEUR
        if prefixe.startswith('2'):
            if prefixe.startswith('28') or prefixe.startswith('29'):
                return 'C'  # Amortissements et dépréciations sont créditeurs
            return 'D'  # Immobilisations brutes sont débitrices (sens normal)
        
        # Classe 3 - Stocks : DÉBITEUR (sens normal)
        # Exception : 
        # - 39 (Dépréciations des stocks) : CRÉDITEUR (le stock peut se déprécier)
        if prefixe.startswith('3'):
            if prefixe.startswith('39'):
                return 'C'  # Dépréciations des stocks sont créditrices
            return 'D'  # Stocks sont débitrices (sens normal)
        
        # Classe 4 - Tiers : Variable selon sous-classe
        # Règles détaillées :
        # - 40 (Fournisseurs) : CRÉDITEUR normal (dettes). Sens débiteur acceptable pour 4091 (avances), avoirs, trop-versé
        # - 41 (Clients) : DÉBITEUR normal (créances). Sens créditeur acceptable pour 4191 (avances), avoirs, trop-perçu
        # - 42 (Personnel) : généralement CRÉDITEUR (salaires à payer)
        # - 43 (Organismes sociaux) : CRÉDITEUR (charges sociales à payer)
        # - 44 (État) : TVA collectée créditeur, TVA déductible débiteur, impôts généralement créditeur, crédit TVA débiteur
        # - 45 (Groupe et associés) : variable selon la situation
        # - 46, 47 (Débiteurs et créditeurs divers) : selon la nature
        # - 48 (Comptes de régularisation) : variable selon le type
        # - 49 (Dépréciations tiers) : CRÉDITEUR
        if prefixe.startswith('4'):
            if prefixe.startswith('40'):  # Fournisseurs : CRÉDITEUR
                if numero_compte.startswith('4091'):  # Avances fournisseurs : exception
                    return 'BOTH'  # Peut être débiteur (avances, avoirs, trop-versé)
                return 'C'  # CRÉDITEUR (sens normal)
            elif prefixe.startswith('41'):  # Clients : DÉBITEUR
                if numero_compte.startswith('4191'):  # Avances clients : exception
                    return 'BOTH'  # Peut être créditeur (avances, avoirs, trop-perçu)
                return 'D'  # DÉBITEUR (sens normal)
            elif prefixe.startswith('42'):  # Personnel : généralement CRÉDITEUR
                return 'C'  # CRÉDITEUR (peut être soldé)
            elif prefixe.startswith('43'):  # Organismes sociaux : CRÉDITEUR
                return 'C'  # CRÉDITEUR
            elif prefixe.startswith('44'):  # État : variable selon le type
                # TVA collectée créditeur, TVA déductible débiteur, impôts généralement créditeur, crédit TVA débiteur
                return 'BOTH'  # Variable selon le sous-compte
            elif prefixe.startswith('45'):  # Groupe et associés : variable
                return 'BOTH'  # Variable selon la situation
            elif prefixe.startswith('46') or prefixe.startswith('47'):  # Débiteurs et créditeurs divers
                return 'BOTH'  # Selon la nature
            elif prefixe.startswith('48'):  # Comptes de régularisation
                return 'BOTH'  # Variable selon le type
            elif prefixe.startswith('49'):  # Dépréciations tiers : CRÉDITEUR
                return 'C'
            return 'BOTH'
        
        # Classe 5 - Financiers : DÉBITEUR (sens normal)
        # Exception : 
        # - 519 (Découverts bancaires) : peut être CRÉDITEUR
        if prefixe.startswith('5'):
            if prefixe.startswith('519'):  # Découverts bancaires : exception
                return 'BOTH'  # Peut être créditeur
            elif prefixe.startswith('58'):  # Virements internes
                return 'BOTH'  # Doit être soldé
            elif prefixe.startswith('59'):  # Dépréciations financiers
                return 'C'  # CRÉDITEUR
            return 'D'  # DÉBITEUR (sens normal)
        
        # Classe 6 - Charges : DÉBITEUR (sens normal)
        if prefixe.startswith('6'):
            return 'D'  # DÉBITEUR
        
        # Classe 7 - Produits : CRÉDITEUR (sens normal)
        if prefixe.startswith('7'):
            return 'C'  # CRÉDITEUR
        
        # Classe 8 - Résultats : selon le résultat
        if prefixe.startswith('8'):
            return 'BOTH'
        
        return 'BOTH'  # Par défaut, les deux sens sont possibles
    
    def _comptes_doivent_etre_soldes(self, numero_compte):
        """
        Vérifie si un compte doit être soldé en fin d'exercice selon les règles SYSCOHADA
        Retourne: (True/False, 'CRITIQUE'/'MOYENNE'/'FAIBLE', message)
        """
        if not numero_compte:
            return False, None, None
        
        # Comptes CRITIQUES qui doivent absolument être soldés
        comptes_critiques = {
            '471': ('CRITIQUE', 'Compte transitoire ou d\'attente (471) doit être soldé en fin d\'exercice. Tout solde indique des écritures en suspens non régularisées. Analyser ligne par ligne et régulariser chaque écriture.'),
            '58': ('CRITIQUE', 'Virements internes (58) doivent être soldés immédiatement. Utilisé pour les virements entre banques/caisses. Tout solde indique une erreur de lettrage ou un virement non comptabilisé des deux côtés.'),
        }
        
        # Comptes MOYENS à surveiller (fin d'exercice)
        comptes_moyens_fin_exercice = {
            '109': ('MOYENNE', 'Associés - capital souscrit non appelé (109) doit être soldé si tout le capital a été appelé. Si non soldé, vérifier la cohérence avec le capital social.'),
            '475': ('MOYENNE', 'Comptes de régularisation divers (475) doivent être soldés si toutes les régularisations ont été effectuées. Analyser les soldes résiduels.'),
        }
        
        # Comptes MOYENS à surveiller (en cours d'exercice)
        comptes_moyens_cours = {
            '422': ('MOYENNE', 'Personnel - Rémunérations dues (422) doivent être soldées lors de la paie suivante. Des soldes anciens peuvent indiquer des erreurs de lettrage ou des paiements non régularisés.'),
            '467': ('MOYENNE', 'Autres comptes débiteurs/créditeurs (467) - compte de passage temporaire. Ne doit pas avoir de solde ancien. Vérifier l\'ancienneté des montants non soldés.'),
            '468': ('MOYENNE', 'Charges à payer et produits à recevoir (468) doivent être soldés lors du paiement/encaissement effectif. Vérifier que les charges/produits provisionnés sont bien régularisés.'),
            '4387': ('MOYENNE', 'Organismes sociaux - Charges à payer (4387) doivent être soldées lors de la déclaration sociale suivante. Vérifier que les provisions sont bien régularisées.'),
            '4386': ('MOYENNE', 'Organismes sociaux - Autres charges à payer (4386) doivent être soldées lors du paiement effectif.'),
            '431': ('MOYENNE', 'Sécurité sociale (431) doit normalement être soldé chaque mois/trimestre lors du paiement. Un solde ancien indique un retard ou une erreur.'),
            '437': ('MOYENNE', 'Autres organismes sociaux (437) - même logique que 431. Normalement soldé chaque mois/trimestre lors du paiement.'),
        }
        
        # Comptes de TVA à solder à chaque déclaration
        comptes_tva = {
            '44551': ('CRITIQUE', 'TVA à décaisser (44551) doit être soldée lors du paiement de la TVA. Un solde ancien indique une TVA non payée → risque fiscal.'),
            '44567': ('MOYENNE', 'Crédit de TVA à reporter (44567) doit être soldé lorsque le crédit est imputé ou remboursé. Vérifier la cohérence avec les déclarations.'),
            '44558': ('MOYENNE', 'TVA à régulariser ou en attente (44558) - compte temporaire à solder rapidement. Tout solde doit être justifié précisément.'),
        }
        
        # Vérifier les comptes critiques
        for compte_key, (gravite, message) in comptes_critiques.items():
            if numero_compte.startswith(compte_key):
                return True, gravite, message
        
        # Vérifier les comptes TVA
        for compte_key, (gravite, message) in comptes_tva.items():
            if numero_compte.startswith(compte_key):
                return True, gravite, message
        
        # Vérifier les comptes moyens (fin d'exercice)
        for compte_key, (gravite, message) in comptes_moyens_fin_exercice.items():
            if numero_compte.startswith(compte_key):
                return True, gravite, message
        
        # Vérifier les comptes moyens (en cours)
        for compte_key, (gravite, message) in comptes_moyens_cours.items():
            if numero_compte.startswith(compte_key):
                return True, gravite, message
        
        # Comptes de gestion (classe 6 et 7) : doivent être soldés en fin d'exercice
        if numero_compte[0] in ['6', '7']:
            return True, 'MOYENNE', f'Compte de gestion (classe {numero_compte[0]}) doit être soldé en fin d\'exercice par le jeu des écritures de clôture (virement au compte de résultat 12). Aucun compte de classe {numero_compte[0]} ne doit avoir de solde à l\'ouverture de l\'exercice N+1. Si solde présent en début d\'exercice : erreur de clôture, oubli du virement au résultat, ou reports à nouveau incorrects.'
        
        return False, None, None

    def _generer_structure_vraisemblance(self, erreurs_detectees=None):
        """
        Génère la structure complète du contrôle de vraisemblance avec :
        - Résumé explicatif (objectif, principes généraux)
        - Tableau par classe (1 à 7) avec nature, sens normal, exceptions, anomalies détectées
        - Tableau des comptes à solder obligatoirement
        - Liste des contrôles essentiels à effectuer
        
        Args:
            erreurs_detectees: Liste des erreurs détectées pour cette balance (optionnel)
        """
        # Regrouper les erreurs de sens par classe pour affichage dans le tableau
        anomalies_par_classe = {}
        if erreurs_detectees:
            for erreur in erreurs_detectees:
                if erreur.get("type") == "signe" and erreur.get("numero_compte"):
                    numero_compte = erreur.get("numero_compte")
                    if numero_compte and numero_compte[0].isdigit():
                        classe = numero_compte[0]
                        if classe not in anomalies_par_classe:
                            anomalies_par_classe[classe] = []
                        # Extraire le message complet et le formater de manière détaillée
                        message_original = erreur.get("message", "")
                        
                        # Construire un message détaillé et structuré
                        message_detaille = message_original
                        
                        # Extraire les informations clés du message original
                        try:
                            # Chercher le sens attendu et le sens actuel
                            if "devrait être" in message_original and "mais le solde est" in message_original:
                                # Extraire le sens attendu
                                if "devrait être" in message_original:
                                    partie_attendu = message_original.split("devrait être")[1]
                                    sens_attendu = partie_attendu.split(",")[0].strip() if "," in partie_attendu else partie_attendu.split(".")[0].strip()
                                
                                # Extraire le sens actuel
                                if "mais le solde est" in message_original:
                                    partie_actuel = message_original.split("mais le solde est")[1]
                                    sens_actuel = partie_actuel.split("(")[0].strip()
                                
                                # Extraire le montant
                                montant = ""
                                if "(" in message_original and "FCFA" in message_original:
                                    try:
                                        montant_part = message_original.split("(")[1].split(")")[0]
                                        montant = montant_part.strip()
                                    except:
                                        pass
                                
                                # Construire le message détaillé
                                message_detaille = f"⚠️ ANOMALIE DÉTECTÉE : Le compte {numero_compte} présente un solde {sens_actuel.lower()}, alors que selon les règles SYSCOHADA pour la classe {classe}, il devrait être {sens_attendu.lower()}."
                                if montant:
                                    message_detaille += f" Le solde anormal s'élève à {montant}."
                                
                                # Ajouter des détails selon la classe
                                if classe == "3":
                                    message_detaille += " Un stock créditeur est ANORMAL et nécessite une investigation immédiate. Vérifiez les écritures d'inventaire, les erreurs de saisie ou les problèmes de valorisation."
                                elif classe == "5":
                                    if "53" in numero_compte:
                                        message_detaille += " Une caisse créditrice est IMPOSSIBLE physiquement. Il s'agit d'une erreur certaine qui doit être corrigée immédiatement."
                                    else:
                                        message_detaille += " Vérifiez la cohérence avec les relevés bancaires et les écritures de trésorerie."
                                elif classe == "6":
                                    message_detaille += " Un compte de charges créditeur indique généralement une erreur de comptabilisation ou un avoir mal enregistré. Vérifiez les écritures comptables et les avoirs clients."
                                elif classe == "7":
                                    message_detaille += " Un compte de produits débiteur indique généralement une erreur (annulation mal comptabilisée, avoir sur vente). Vérifiez les écritures d'annulation et les avoirs accordés."
                                elif classe == "2":
                                    message_detaille += " Un compte d'immobilisation créditeur est généralement anormal et peut indiquer une erreur ou une cession non comptabilisée correctement. Vérifiez les écritures de cession d'immobilisations."
                                else:
                                    message_detaille += " Vérifiez la nature du compte et les écritures comptables pour identifier l'origine de cette anomalie."
                        except Exception as e:
                            # En cas d'erreur d'extraction, utiliser le message original
                            message_detaille = message_original
                        
                        anomalies_par_classe[classe].append({
                            "compte": numero_compte,
                            "message": message_detaille
                        })
        structure = {
            "resume": {
                "objectif": "Le contrôle de vraisemblance vérifie la cohérence des soldes comptables selon les règles du plan comptable SYSCOHADA. Il permet de détecter les anomalies de sens, les comptes non soldés et les erreurs de comptabilisation.",
                "principes_generaux": [
                    "Chaque classe de compte a un sens normal de solde (débit ou crédit) selon sa nature",
                    "Certains comptes doivent obligatoirement être soldés en fin d'exercice",
                    "Des exceptions existent pour certains comptes spécifiques",
                    "Les anomalies détectées doivent être justifiées ou corrigées avant clôture"
                ]
            },
            "tableau_classes": [
                {
                    "classe": "1",
                    "nom": "Comptes de capitaux",
                    "nature": "Capital social, réserves, report à nouveau créditeur, subventions d'investissement, provisions réglementées, emprunts et dettes assimilées",
                    "sens_normal": "CRÉDITEUR",
                    "exceptions": [
                        "Report à nouveau débiteur (pertes antérieures) : débiteur accepté",
                        "Associés - capital souscrit non appelé (109) : débiteur (vient en diminution du capital)",
                        "Primes de remboursement des obligations : débiteur (charges à répartir)",
                        "Écarts de réévaluation (105) : peut être débiteur ou créditeur"
                    ],
                    "anomalies_detectees": anomalies_par_classe.get("1", [])
                },
                {
                    "classe": "2",
                    "nom": "Comptes d'immobilisations",
                    "nature": "Tous les comptes d'immobilisations brutes (20, 21, 23, etc.), immobilisations en cours, avances et acomptes sur immobilisations",
                    "sens_normal": "DÉBITEUR",
                    "exceptions": [
                        "Comptes d'amortissements (28x) : CRÉDITEUR",
                        "Comptes de dépréciations (29x) : CRÉDITEUR",
                        "Un compte d'immobilisation créditeur est généralement anormal (peut indiquer une erreur ou une cession non comptabilisée correctement)"
                    ],
                    "anomalies_detectees": anomalies_par_classe.get("2", [])
                },
                {
                    "classe": "3",
                    "nom": "Comptes de stocks",
                    "nature": "Tous les stocks (matières, marchandises, produits), en-cours de production",
                    "sens_normal": "DÉBITEUR",
                    "exceptions": [
                        "Comptes de dépréciation des stocks (39x) : CRÉDITEUR",
                        "Un stock créditeur est anormal et doit être investigué"
                    ],
                    "anomalies_detectees": anomalies_par_classe.get("3", [])
                },
                {
                    "classe": "4",
                    "nom": "Comptes de tiers",
                    "nature": "Comptes clients (41x), fournisseurs (40x), personnel (42x), sécurité sociale (43x), État et collectivités (44x), groupe et associés (45x), débiteurs et créditeurs divers (46x, 47x)",
                    "sens_normal": "Variable selon sous-classe",
                    "exceptions": [
                        "Comptes clients (41x) : DÉBITEUR normal (créances). Sens créditeur acceptable pour avances et acomptes reçus (4191), avoirs à établir, trop-perçu",
                        "Comptes fournisseurs (40x) : CRÉDITEUR normal (dettes). Sens débiteur acceptable pour avances et acomptes versés (4091), avoirs à obtenir, trop-versé",
                        "Personnel (42x) : généralement CRÉDITEUR (salaires à payer)",
                        "Sécurité sociale (43x) : CRÉDITEUR (charges sociales à payer)",
                        "État (44x) : TVA collectée créditeur, TVA déductible débiteur, impôts sur bénéfices généralement créditeur, crédit de TVA débiteur",
                        "Groupe et associés (45x) : variable selon la situation",
                        "Débiteurs et créditeurs divers (46x, 47x) : selon la nature",
                        "Dépréciations (49x) : CRÉDITEUR"
                    ],
                    "anomalies_detectees": anomalies_par_classe.get("4", [])
                },
                {
                    "classe": "5",
                    "nom": "Comptes financiers",
                    "nature": "Comptes de trésorerie active : Banques (512), Caisse (53x), Valeurs mobilières de placement (50x). Autres comptes financiers : Charges constatées d'avance (486), Produits constatés d'avance (487)",
                    "sens_normal": "DÉBITEUR",
                    "exceptions": [
                        "Concours bancaires courants (519) : CRÉDITEUR (découverts). Une banque créditrice indique un découvert",
                        "Caisse (53x) : TOUJOURS DÉBITEUR (un solde créditeur de caisse est impossible physiquement !)",
                        "Produits constatés d'avance (487) : CRÉDITEUR"
                    ],
                    "anomalies_detectees": anomalies_par_classe.get("5", [])
                },
                {
                    "classe": "6",
                    "nom": "Comptes de charges",
                    "nature": "Tous les comptes de charges (60 à 69) doivent être débiteurs en cours d'exercice",
                    "sens_normal": "DÉBITEUR",
                    "exceptions": [
                        "En fin d'exercice, après inventaire, certains comptes peuvent être soldés",
                        "Un compte de charges créditeur indique généralement une erreur de comptabilisation ou un avoir mal enregistré"
                    ],
                    "anomalies_detectees": anomalies_par_classe.get("6", [])
                },
                {
                    "classe": "7",
                    "nom": "Comptes de produits",
                    "nature": "Tous les comptes de produits (70 à 79) doivent être créditeurs en cours d'exercice",
                    "sens_normal": "CRÉDITEUR",
                    "exceptions": [
                        "En fin d'exercice, après inventaire, certains comptes peuvent être soldés",
                        "Un compte de produits débiteur indique généralement une erreur (annulation mal comptabilisée, avoir sur vente)"
                    ],
                    "anomalies_detectees": anomalies_par_classe.get("7", [])
                }
            ],
            "comptes_a_solder": self._generer_tableau_comptes_a_solder(),
            "controles_essentiels": [
                "Vérifier que tous les comptes de classe 6 et 7 sont soldés en fin d'exercice (virement au compte de résultat 12)",
                "Solder les comptes transitoires (471) - tout solde indique des écritures en suspens",
                "Solder les virements internes (58) - tout solde indique une erreur de lettrage",
                "Vérifier que la caisse (53) n'est jamais créditrice (impossible physiquement)",
                "Contrôler les stocks créditeurs (anormal - investigation obligatoire)",
                "Solder les comptes de régularisation (468, 475) en fin d'exercice",
                "Vérifier la cohérence des comptes de TVA (44551, 44567, 44558) avec les déclarations",
                "Contrôler les comptes de charges à payer et produits à recevoir (468) - doivent être soldés lors du paiement/encaissement",
                "Vérifier les comptes sociaux (431, 437) - normalement soldés chaque mois/trimestre",
                "Contrôler les comptes de rémunérations dues (422) - doivent être soldés lors de la paie suivante",
                "Contrôler la cohérence des amortissements (28) et dépréciations (29) avec les valeurs brutes",
                "Vérifier que les comptes de gestion (6 et 7) n'ont aucun solde à l'ouverture de l'exercice N+1",
                "Effectuer la concordance bancaire pour tous les comptes bancaires",
                "Vérifier la justification de tous les comptes débiteurs/créditeurs divers (467) avec soldes anciens"
            ]
        }
        
        return structure

    def _generer_tableau_comptes_a_solder(self):
        """
        Génère le tableau des comptes à solder obligatoirement avec numéro, moment, gravité
        """
        return [
            {
                "numero": "471",
                "libelle": "Comptes transitoires ou d'attente",
                "moment": "Fin d'exercice",
                "gravite": "CRITIQUE",
                "raison": "Tout solde indique des écritures en suspens non régularisées. Analyser ligne par ligne et régulariser chaque écriture."
            },
            {
                "numero": "58",
                "libelle": "Virements internes",
                "moment": "Immédiat",
                "gravite": "CRITIQUE",
                "raison": "Utilisé pour les virements entre banques/caisses. Tout solde indique une erreur de lettrage ou un virement non comptabilisé des deux côtés."
            },
            {
                "numero": "422",
                "libelle": "Personnel - Rémunérations dues",
                "moment": "Selon cycle paie",
                "gravite": "MOYENNE",
                "raison": "Doivent être soldées lors de la paie suivante. Des soldes anciens peuvent indiquer des erreurs de lettrage ou des paiements non régularisés."
            }
        ]

    def _verifier_presence_comptes_obligatoires(self, numeros_comptes_presents):
        """
        Vérifie la complétude par classe (1 à 6) selon SYSCOHADA.
        Retourne: (liste d'erreurs, liste détaillée des classes avec comptes présents/attendus).
        """
        erreurs = []
        classes_presence = []
        classes_attendues = {
            '1': {
                'libelle': "Classe 1 - Capitaux",
                'description': "Comptes 10 à 19 (capital, réserves, reports à nouveau, subventions, provisions réglementées, emprunts)",
                'exemples': ["101", "105", "12", "13", "16"]
            },
            '2': {
                'libelle': "Classe 2 - Immobilisations",
                'description': "Comptes 20 à 29 (immobilisations incorporelles, corporelles, financières, avances, amortissements, dépréciations)",
                'exemples': ["203", "213", "218", "281", "291"]
            },
            '3': {
                'libelle': "Classe 3 - Stocks",
                'description': "Comptes 30 à 39 (stocks de marchandises, matières, en-cours, produits, dépréciations)",
                'exemples': ["31", "32", "33", "37", "39"]
            },
            '4': {
                'libelle': "Classe 4 - Comptes de tiers",
                'description': "Comptes 40 à 49 (fournisseurs, clients, personnel, organismes sociaux, État, comptes divers, dépréciations)",
                'exemples': ["401", "4091", "411", "4191", "431", "445"]
            },
            '5': {
                'libelle': "Classe 5 - Comptes financiers",
                'description': "Comptes 50 à 59 (banques, caisse, VMP, virements internes, dépréciations)",
                'exemples': ["512", "53", "518", "519", "581"]
            },
            '6': {
                'libelle': "Classe 6 - Comptes de charges",
                'description': "Comptes 60 à 69 (achats, charges externes, impôts, charges de personnel, dotations, charges financières/exceptionnelles)",
                'exemples': ["601", "606", "62", "64", "68"]
            }
        }
        
        # Indexer les comptes présents par classe
        comptes_par_classe = {classe: [] for classe in classes_attendues.keys()}
        for numero in numeros_comptes_presents:
            if numero and numero[0] in comptes_par_classe:
                comptes_par_classe[numero[0]].append(numero)
        
        # Vérifier la présence de chaque classe
        for classe, infos in classes_attendues.items():
            comptes_classe = sorted(comptes_par_classe.get(classe, []))
            if comptes_classe:
                comptes_existants = ", ".join(comptes_classe)
                comptes_inexistants = infos['description']
                details = (
                    f"{len(comptes_classe)} compte(s) recensé(s) dans la classe {classe}. "
                    f"Exemples attendus : {', '.join(infos['exemples'])}."
                )
                status = "Présente"
            else:
                comptes_existants = "Aucun compte recensé dans la balance"
                comptes_inexistants = infos['description']
                details = (
                    f"{infos['libelle']} absente. Exemples attendus : {', '.join(infos['exemples'])}."
                )
                status = "Absente"
                message = (
                    f"{infos['libelle']} absente : aucun compte de cette classe n'est présent dans la balance. "
                    f"Attendu : {infos['description']}."
                )
                erreurs.append({
                    "type": "completude",
                    "numero_compte": "-",
                    "message": message,
                    "comptes_existants": comptes_existants,
                    "comptes_inexistants": comptes_inexistants,
                    "details": details,
                    "classe": infos['libelle']
                })

            classes_presence.append({
                "classe": infos['libelle'],
                "code_classe": classe,
                "status": status,
                "comptes_existants": comptes_existants,
                "comptes_inexistants": infos['description'],
                "details": details,
                "exemples": infos['exemples']
            })

        return erreurs, classes_presence

    def _valider_numerotation_syscohada(self, numero_compte):
        """
        Valide que la numérotation du compte suit le plan comptable SYSCOHADA
        Retourne: (True/False, message d'erreur si False)
        """
        if not numero_compte:
            return False, "Numéro de compte vide"
        
        # Nettoyer le numéro de compte (enlever espaces, caractères spéciaux)
        numero_clean = ''.join(c for c in numero_compte if c.isdigit())
        
        if not numero_clean:
            return False, f"Le numéro de compte '{numero_compte}' ne contient aucun chiffre valide"
        
        # Vérifier la longueur minimale (au moins 2 chiffres pour identifier la classe / sous-classe)
        if len(numero_clean) < 2:
            return False, f"Le numéro de compte '{numero_compte}' est trop court (minimum 2 chiffres requis)"
        
        # Vérifier que le premier chiffre correspond à une classe valide (1-9)
        premiere_classe = numero_clean[0]
        if premiere_classe not in ['1', '2', '3', '4', '5', '6', '7', '8', '9']:
            return False, f"Le numéro de compte '{numero_compte}' commence par '{premiere_classe}' qui n'est pas une classe valide (1-9 selon SYSCOHADA)"
        
        # Vérifier les plages de comptes selon SYSCOHADA
        # Les deux premiers chiffres doivent être dans la plage appropriée pour la classe
        if len(numero_clean) >= 2:
            deux_premiers = numero_clean[:2]
            try:
                deux_premiers_int = int(deux_premiers)
            except ValueError:
                return False, f"Le numéro de compte '{numero_compte}' contient des caractères non numériques invalides"
            
            classe = premiere_classe
            
            # Classe 1 : 10-19 (permet les sous-comptes comme 101, 102, etc.)
            if classe == '1' and not (10 <= deux_premiers_int <= 19):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 1 doit commencer par 10-19 selon SYSCOHADA"
            
            # Classe 2 : 20-29
            if classe == '2' and not (20 <= deux_premiers_int <= 29):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 2 doit commencer par 20-29 selon SYSCOHADA"
            
            # Classe 3 : 30-39
            if classe == '3' and not (30 <= deux_premiers_int <= 39):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 3 doit commencer par 30-39 selon SYSCOHADA"
            
            # Classe 4 : 40-49
            if classe == '4' and not (40 <= deux_premiers_int <= 49):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 4 doit commencer par 40-49 selon SYSCOHADA"
            
            # Classe 5 : 50-59
            if classe == '5' and not (50 <= deux_premiers_int <= 59):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 5 doit commencer par 50-59 selon SYSCOHADA"
            
            # Classe 6 : 60-69
            if classe == '6' and not (60 <= deux_premiers_int <= 69):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 6 doit commencer par 60-69 selon SYSCOHADA"
            
            # Classe 7 : 70-79
            if classe == '7' and not (70 <= deux_premiers_int <= 79):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 7 doit commencer par 70-79 selon SYSCOHADA"
            
            # Classe 8 : 80-89
            if classe == '8' and not (80 <= deux_premiers_int <= 89):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 8 doit commencer par 80-89 selon SYSCOHADA"
            
            # Classe 9 : 90-99
            if classe == '9' and not (90 <= deux_premiers_int <= 99):
                return False, f"Le numéro de compte '{numero_compte}' de la classe 9 doit commencer par 90-99 selon SYSCOHADA"
        
        return True, None

    def _coherence_checks_for_year(self, lines):
        report = {"equilibre_global": True}
        erreurs = []
        numeros_comptes_vus = set()
        comptes_dupliques = []
        numeros_comptes_valides = []  # Pour vérifier la présence des comptes obligatoires
        # Dictionnaire pour compter les occurrences de chaque compte
        occurrences_comptes = {}  # {numero_compte: [indices des lignes]}

        # ===== CONTRÔLES ARITHMÉTIQUES =====
        # IMPORTANT : Les deux contrôles arithmétiques sont ESSENTIELS et doivent être effectués pour TOUS les comptes
        
        # 1) PREMIER CONTRÔLE ARITHMÉTIQUE : Vérifier que le total des débits est strictement égal au total des crédits
        # Ce test vérifie l'équilibre global de la balance en additionnant TOUS les débits finaux et TOUS les crédits finaux
        
        # Vérifier que lines n'est pas vide
        if not lines or len(lines) == 0:
            erreurs.append({
                "type": "equilibre",
                "numero_compte": "-",
                "message": "Aucune ligne de balance à analyser. Le premier contrôle arithmétique ne peut pas être effectué."
            })
            report["equilibre_global"] = False
            return report
        
        # Compter le nombre total de lignes analysées (toutes les lignes, même sans numéro de compte)
        nb_comptes_analyses = len([x for x in lines if x])  # Toutes les lignes non vides
        
        # Calculer les totaux : somme de TOUS les débits finaux et crédits finaux de TOUTES les lignes
        # Ce calcul inclut TOUTES les lignes, même celles sans numéro de compte valide
        sum_deb_fin = 0
        sum_cre_fin = 0
        
        # PREMIER CONTRÔLE ARITHMÉTIQUE : TOUJOURS EXÉCUTÉ
        # Calculer la somme de tous les débits finaux et crédits finaux
        lignes_avec_valeurs = 0
        for x in lines:
            try:
                deb = int(x.get("debit_fin", 0) or 0)
                cre = int(x.get("credit_fin", 0) or 0)
                sum_deb_fin += deb
                sum_cre_fin += cre
                if deb > 0 or cre > 0:
                    lignes_avec_valeurs += 1
            except (ValueError, TypeError):
                # Ignorer les valeurs invalides mais continuer le calcul
                continue
        
        # Le premier contrôle est maintenant exécuté : sum_deb_fin et sum_cre_fin sont calculés
        # DEBUG : Afficher les totaux calculés et les premières lignes pour vérifier
        print(f"[PREMIER CONTRÔLE] ========================================")
        def format_espace(val):
            return '{:,.0f}'.format(val).replace(',', ' ')

        print(f"[PREMIER CONTRÔLE] RÉSULTATS EXACTS DU CALCUL :")
        print(f"[PREMIER CONTRÔLE] Total débits finaux : {format_espace(sum_deb_fin)} FCFA")
        print(f"[PREMIER CONTRÔLE] Total crédits finaux : {format_espace(sum_cre_fin)} FCFA")
        print(f"[PREMIER CONTRÔLE] ÉCART : {format_espace(abs(sum_deb_fin - sum_cre_fin))} FCFA")
        print(f"[PREMIER CONTRÔLE] Nombre de lignes analysées : {nb_comptes_analyses}")
        print(f"[PREMIER CONTRÔLE] Lignes avec valeurs > 0 : {lignes_avec_valeurs}")
        print(f"[PREMIER CONTRÔLE] ========================================")
        
        # Afficher les 5 premières lignes avec valeurs pour déboguer
        if len(lines) > 0:
            print(f"[PREMIER CONTRÔLE] Exemple de données (5 premières lignes avec valeurs) :")
            count = 0
            for i, x in enumerate(lines):
                if count >= 5:
                    break
                deb = x.get("debit_fin", 0) or 0
                cre = x.get("credit_fin", 0) or 0
                num = x.get("numero_compte", "N/A")
                lib = str(x.get("libelle", ""))[:30]
                if int(deb) > 0 or int(cre) > 0:
                    print(f"  Ligne {i+1}: compte={num}, libelle={lib}, debit_fin={deb:,}, credit_fin={cre:,}, solde={abs(int(deb)-int(cre)):,}")
                    count += 1
        
        # TOUJOURS exécuter le premier contrôle arithmétique
        # Vérifier si les totaux sont égaux
        if sum_deb_fin != sum_cre_fin:
            report["equilibre_global"] = False
            ecart = abs(sum_deb_fin - sum_cre_fin)
            
            # Identifier TOUS les comptes pour l'affichage
            comptes_anormaux = []
            for x in lines:
                numero_compte = str(x.get("numero_compte", "")).strip()
                if numero_compte and numero_compte != "None" and numero_compte.lower() != "nan":
                    df = int(x.get("debit_fin", 0) or 0)
                    cf = int(x.get("credit_fin", 0) or 0)
                    solde = abs(df - cf)
                    libelle = x.get("libelle", "") or x.get("libelle_compte", "") or "Sans libellé"
                    # Afficher TOUS les comptes, même ceux avec solde = 0
                    comptes_anormaux.append({
                        "compte": numero_compte,
                        "libelle": libelle,
                        "debit_fin": df,
                        "credit_fin": cf,
                        "solde": solde
                    })
            
            # Trier par solde décroissant - AFFICHER TOUS LES COMPTES
            comptes_anormaux.sort(key=lambda x: x["solde"], reverse=True)
            
            # Message d'alerte justifié et clair
            message = f"⚠️ ALERTE : Déséquilibre détecté dans la balance\n\n"
            message += f"🔴 ÉCART DÉTECTÉ : {ecart:,} FCFA\n\n"
            
            if sum_deb_fin > sum_cre_fin:
                message += f"Justification : Le total des DÉBITS ({format_espace(sum_deb_fin)} FCFA) est supérieur au total des CRÉDITS ({format_espace(sum_cre_fin)} FCFA).\n"
                message += f"Il manque {format_espace(ecart)} FCFA au crédit (ou il y a {format_espace(ecart)} FCFA en trop au débit).\n\n"
            else:
                message += f"Justification : Le total des CRÉDITS ({format_espace(sum_cre_fin)} FCFA) est supérieur au total des DÉBITS ({format_espace(sum_deb_fin)} FCFA).\n"
                message += f"Il manque {format_espace(ecart)} FCFA au débit (ou il y a {format_espace(ecart)} FCFA en trop au crédit).\n\n"
            
            message += f"Principe comptable : En comptabilité en partie double, le total des débits DOIT être STRICTEMENT ÉGAL au total des crédits.\n"
            message += f"Toute différence, même de 1 FCFA, indique FORCÉMENT une erreur dans les écritures comptables.\n\n"
            
            if comptes_anormaux:
                message += f"Comptes concernés ({len(comptes_anormaux)} compte(s)) :\n"
                for compte_info in comptes_anormaux:
                    message += f"• {compte_info['compte']} - {compte_info['libelle']} : Débit {compte_info['debit_fin']:,} | Crédit {compte_info['credit_fin']:,} | Solde {compte_info['solde']:,} FCFA\n"
            
            # Ajouter l'erreur d'équilibre à la liste des erreurs
            erreur_equilibre = {
                "type": "equilibre",
                "numero_compte": "-",
                "message": message
            }
            erreurs.append(erreur_equilibre)
            print(f"[PREMIER CONTRÔLE] Erreur d'équilibre ajoutée : type={erreur_equilibre['type']}, message_length={len(message)}")
            
            # Ajouter l'écart dans le rapport pour affichage dans le frontend
            report["ecart_equilibre"] = ecart
            report["total_debits"] = sum_deb_fin
            report["total_credits"] = sum_cre_fin
            
            # Créer la structure verification_equilibre même en cas de déséquilibre pour afficher les résultats exacts
            if sum_deb_fin > sum_cre_fin:
                explication = f"Le système a détecté un DÉSÉQUILIBRE : le total des débits ({format_espace(sum_deb_fin)} FCFA) est supérieur au total des crédits ({format_espace(sum_cre_fin)} FCFA). Il manque {format_espace(ecart)} FCFA au crédit (ou il y a {format_espace(ecart)} FCFA en trop au débit)."
            else:
                explication = f"Le système a détecté un DÉSÉQUILIBRE : le total des crédits ({format_espace(sum_cre_fin)} FCFA) est supérieur au total des débits ({format_espace(sum_deb_fin)} FCFA). Il manque {format_espace(ecart)} FCFA au débit (ou il y a {format_espace(ecart)} FCFA en trop au crédit)."
            
            report["verification_equilibre"] = {
                "statut": "ERREUR",
                "total_debits": sum_deb_fin,
                "total_credits": sum_cre_fin,
                "ecart": ecart,
                "nb_comptes_analyses": nb_comptes_analyses,
                "explication": explication
            }
            
            print(f"[PREMIER CONTRÔLE] Rapport mis à jour : equilibre_global=False, ecart={ecart:,} FCFA, total_debits={sum_deb_fin:,} FCFA, total_credits={sum_cre_fin:,} FCFA")
        else:
            # Les totaux sont égaux - équilibre OK
            report["equilibre_global"] = True
            print(f"[PREMIER CONTRÔLE] ✅ Équilibre OK : Total débits = Total crédits = {sum_deb_fin:,} FCFA")
            
            # Ajouter une information de vérification réussie dans le rapport
            report["verification_equilibre"] = {
                "statut": "OK",
                "total_debits": format_espace(sum_deb_fin),
                "total_credits": format_espace(sum_cre_fin),
                "ecart": 0,
                "nb_comptes_analyses": nb_comptes_analyses,
                "explication": f"Le système a vérifié que le total des débits ({format_espace(sum_deb_fin)} FCFA) est strictement égal au total des crédits ({format_espace(sum_cre_fin)} FCFA) en additionnant les colonnes 'Débit fin' et 'Crédit fin' de {nb_comptes_analyses} compte(s)."
            }
            report["ecart_equilibre"] = 0  # Pas d'écart si équilibré
            report["total_debits"] = sum_deb_fin
            report["total_credits"] = sum_cre_fin
        
        # IMPORTANT : Le premier contrôle arithmétique est TOUJOURS exécuté
        # Il vérifie que sum_deb_fin == sum_cre_fin
        # Si les totaux diffèrent, une erreur de type "equilibre" est ajoutée à la liste des erreurs
        # Même si les totaux sont égaux, le contrôle a été effectué et doit être documenté dans le rapport
        
        # S'assurer que les informations du premier contrôle sont TOUJOURS dans le rapport
        # (même si elles ont déjà été créées dans le bloc if/else ci-dessus)
        if "verification_equilibre" not in report:
            # Si le contrôle n'a pas créé verification_equilibre, le créer maintenant
            report["verification_equilibre"] = {
                "statut": "OK" if report.get("equilibre_global", True) else "ERREUR",
                "total_debits": sum_deb_fin,
                "total_credits": sum_cre_fin,
                "nb_comptes_analyses": nb_comptes_analyses,
                "explication": f"Le système a vérifié que le total des débits ({sum_deb_fin:,} FCFA) est strictement égal au total des crédits ({sum_cre_fin:,} FCFA) en additionnant les colonnes 'Débit fin' et 'Crédit fin' de {nb_comptes_analyses} compte(s)."
            }
        
        # Confirmer que le premier contrôle a été exécuté
        print(f"[PREMIER CONTRÔLE ARITHMÉTIQUE] Exécuté : {nb_comptes_analyses} ligne(s) analysée(s), Total débits = {sum_deb_fin:,} FCFA, Total crédits = {sum_cre_fin:,} FCFA, Équilibre = {'OK' if report.get('equilibre_global', True) else 'DÉSÉQUILIBRÉ'}")

        # 2) Détection des erreurs d'identité, de signe, arithmétiques et des comptes à solder
        nb_erreurs_identite = 0
        nb_erreurs_signe = 0
        nb_erreurs_arithmetique = 0
        nb_erreurs_comptes_soldes = 0
        
        # Compteurs pour le contrôle de vraisemblance
        nb_comptes_verifies_vraisemblance = 0
        nb_comptes_vraisemblance_ok = 0
        nb_comptes_erreur_signe = 0
        nb_comptes_erreur_soldes = 0
        
        # Compteurs pour le second contrôle arithmétique (formule solde d'ouverture + mouvements = solde de clôture)
        nb_comptes_verifies_formule = 0
        nb_comptes_formule_ok = 0
        nb_comptes_formule_erreur = 0
        
        print(f"[SECOND CONTRÔLE] Début de la vérification de la formule pour {len(lines)} ligne(s) de balance")
        
        # Première passe : collecter tous les numéros de comptes (on ne signale plus les doublons ici)
        for idx, x in enumerate(lines):
            numero_compte = str(x.get("numero_compte", "")).strip()
            if not numero_compte or numero_compte == "None" or numero_compte.lower() == "nan":
                continue
            numero_compte_normalise = numero_compte.strip()
            if numero_compte_normalise not in occurrences_comptes:
                occurrences_comptes[numero_compte_normalise] = []
            occurrences_comptes[numero_compte_normalise].append(idx)
        
        # Deuxième passe : vérifier les autres erreurs
        print(f"[SECOND CONTRÔLE] Début de la deuxième passe : vérification de la formule pour tous les comptes")
        for idx, x in enumerate(lines):
            numero_compte = str(x.get("numero_compte", "")).strip()
            if not numero_compte or numero_compte == "None" or numero_compte.lower() == "nan":
                # Pour les lignes sans numéro de compte, on utilise un identifiant temporaire
                numero_compte = f"LIGNE_SANS_NUMERO_{len(erreurs)}"
            
            # 2.1) Vérifier la numérotation SYSCOHADA (pour information, mais ne bloque pas)
            # On ne bloque plus sur la numérotation SYSCOHADA dans ce contrôle :
            # le contrôle de complétude/exhaustivité a été désactivé à la demande de l'utilisateur.
            is_valid, error_msg = self._valider_numerotation_syscohada(numero_compte)
            # On continue même si la validation SYSCOHADA échoue, car la vérification de la formule
            # est une vérification arithmétique fondamentale qui doit être faite pour TOUS les comptes
            
            # 2.2) Ajouter aux comptes valides pour vérification (si pas déjà en doublon)
            if numero_compte not in comptes_dupliques:
                if numero_compte not in numeros_comptes_vus:
                    numeros_comptes_vus.add(numero_compte)
                    if not numero_compte.startswith("LIGNE_SANS_NUMERO_"):
                        numeros_comptes_valides.append(numero_compte)
                
            di = int(x.get("debit_initial", 0) or 0)
            ci = int(x.get("credit_initial", 0) or 0)
            df = int(x.get("debit_fin", 0) or 0)
            cf = int(x.get("credit_fin", 0) or 0)
            
            # Calcul des soldes
            solde_initial = di - ci  # Solde d'ouverture
            solde_fin = df - cf  # Solde de clôture
            
            # ===== SECOND CONTRÔLE ARITHMÉTIQUE =====
            # IMPORTANT : Cette vérification doit être effectuée pour TOUS les comptes,
            # indépendamment de la validation SYSCOHADA, car c'est une vérification arithmétique fondamentale
            # 
            # Ce test vérifie la cohérence des soldes de CHAQUE compte individuellement :
            # Vérifier que : Solde de clôture = Solde d'ouverture + Mouvements de période
            # Formule : Solde de clôture = Solde d'ouverture + (Mouvement débit - Mouvement crédit)
            # Où :
            #   - Solde d'ouverture = Débit initial - Crédit initial
            #   - Solde de clôture = Débit fin - Crédit fin
            #   - Mouvements = (Débit fin - Débit initial) - (Crédit fin - Crédit initial)
            
            # Compter ce compte comme vérifié
            nb_comptes_verifies_formule += 1
            
            # Log pour tous les comptes (pour debug)
            if numero_compte == "46210010":
                print(f"[SECOND CONTRÔLE] ⚠️  COMPTE 46210010 DÉTECTÉ - Début de la vérification")
            
            # Récupérer les mouvements explicites s'ils sont fournis
            # Les mouvements peuvent être stockés sous différents noms selon le format d'import
            # IMPORTANT : Vérifier si les colonnes de mouvements EXISTENT dans les données
            # (même si elles sont vides/0, si elles existent, on doit les utiliser)
            # Cela permet de détecter les erreurs où mouvements=0 mais solde final != solde initial
            
            # Vérifier si les clés de mouvements existent dans le dictionnaire
            debit_explicite_present = "mouvement_debit" in x or "debit_mvt" in x
            credit_explicite_present = "mouvement_credit" in x or "credit_mvt" in x
            
            # Récupérer les valeurs (même si elles sont 0 ou None)
            if debit_explicite_present:
                mouvement_debit_explicite = x.get("mouvement_debit") or x.get("debit_mvt") or 0
                try:
                    mouvement_debit_explicite = int(mouvement_debit_explicite or 0)
                except (ValueError, TypeError):
                    mouvement_debit_explicite = 0
            else:
                mouvement_debit_explicite = 0
            
            if credit_explicite_present:
                mouvement_credit_explicite = x.get("mouvement_credit") or x.get("credit_mvt") or 0
                try:
                    mouvement_credit_explicite = int(mouvement_credit_explicite or 0)
                except (ValueError, TypeError):
                    mouvement_credit_explicite = 0
            else:
                mouvement_credit_explicite = 0
            
            # Calculer les mouvements à partir des débits/crédits (pour vérification)
            mouvement_debit_calcule = df - di
            mouvement_credit_calcule = cf - ci
            
            # IMPORTANT : Si les mouvements explicites sont PRÉSENTS dans les données (même à 0),
            # on DOIT les utiliser pour la vérification. Sinon, on utilise les mouvements calculés.
            # Cela permet de détecter les erreurs où les mouvements sont à 0 mais le solde final n'est pas à 0.
            if debit_explicite_present or credit_explicite_present:
                # Utiliser les mouvements explicites (même s'ils sont à 0)
                mouvement_debit = mouvement_debit_explicite
                mouvement_credit = mouvement_credit_explicite
                source_mouvements = "explicites"
            else:
                # Utiliser les mouvements calculés (quand les colonnes de mouvements n'existent pas)
                mouvement_debit = mouvement_debit_calcule
                mouvement_credit = mouvement_credit_calcule
                source_mouvements = "calculés"
            
            # Calculer le solde de clôture attendu selon la formule
            # Solde de clôture = Solde d'ouverture + (Mouvement débit - Mouvement crédit)
            solde_cloture_attendu = solde_initial + (mouvement_debit - mouvement_credit)
            
            # Vérifier la cohérence avec une tolérance STRICTE de 0 FCFA (pas d'arrondi accepté)
            # En comptabilité, la formule doit être STRICTEMENT respectée
            ecart = abs(solde_fin - solde_cloture_attendu)
            
            # Log détaillé pour TOUS les comptes avec écart > 0 (même très petit)
            if ecart > 0:
                print(f"[SECOND CONTRÔLE] Compte {numero_compte}:")
                print(f"  Solde initial: {solde_initial:,} (Débit init {di:,} - Crédit init {ci:,})")
                print(f"  Mouvements: Débit {mouvement_debit:,} | Crédit {mouvement_credit:,} | Net {mouvement_debit - mouvement_credit:,} (source: {source_mouvements})")
                print(f"  Solde attendu: {solde_cloture_attendu:,} FCFA")
                print(f"  Solde réel: {solde_fin:,} FCFA (Débit fin {df:,} - Crédit fin {cf:,})")
                print(f"  ÉCART: {ecart:,} FCFA")
                print(f"  Erreur détectée: {'OUI' if ecart > 0 else 'NON'}")
            
            # Tolérance STRICTE : tout écart > 0 est une erreur
            if ecart > 0:
                nb_comptes_formule_erreur += 1
                nb_erreurs_arithmetique += 1
                print(f"[SECOND CONTRÔLE] Erreur détectée pour compte {numero_compte}: écart={ecart:,} FCFA (solde_fin={solde_fin:,}, attendu={solde_cloture_attendu:,})")
                
                # Récupérer le libellé du compte
                libelle = x.get("libelle", "") or x.get("libelle_compte", "") or "Sans libellé"
                
                # Message d'alerte justifié et clair
                compte_display = "[SANS NUMÉRO]" if numero_compte.startswith("LIGNE_SANS_NUMERO_") else numero_compte
                message = f"⚠️ ALERTE : Formule non respectée pour le compte {compte_display}\n"
                message += f"Libellé : {libelle}\n\n"
                message += f"🔴 ÉCART DÉTECTÉ : {ecart:,} FCFA\n\n"
                message += f"Formule attendue : Solde de clôture = Solde d'ouverture + Mouvements de période\n"
                message += f"• Solde attendu : {solde_cloture_attendu:,} FCFA\n"
                message += f"• Solde réel : {solde_fin:,} FCFA\n"
                message += f"• Écart : {ecart:,} FCFA\n\n"
                message += f"Détails des valeurs :\n"
                message += f"• Solde d'ouverture : {solde_initial:,} FCFA (Débit initial {di:,} - Crédit initial {ci:,})\n"
                message += f"• Mouvements : Débit {mouvement_debit:,} | Crédit {mouvement_credit:,} | Net {mouvement_debit - mouvement_credit:,} FCFA\n"
                message += f"• Solde de clôture : {solde_fin:,} FCFA (Débit fin {df:,} - Crédit fin {cf:,})\n\n"
                message += f"Justification : La formule comptable fondamentale n'est pas respectée. "
                message += f"Le solde de clôture devrait être {solde_cloture_attendu:,} FCFA mais il est {solde_fin:,} FCFA. "
                message += f"Cet écart de {ecart:,} FCFA indique une erreur dans les écritures comptables de ce compte."
                
                erreurs.append({
                    "type": "arithmetique",
                    "numero_compte": numero_compte,
                    "message": message
                })
            else:
                # La formule est respectée pour ce compte
                nb_comptes_formule_ok += 1
                # Log pour le compte 46210010 même si pas d'erreur
                if numero_compte == "46210010":
                    print(f"[SECOND CONTRÔLE] Compte 46210010: Formule respectée (écart={ecart:.2f} <= tolérance 0.01)")
            # ===== FIN DU SECOND CONTRÔLE ARITHMÉTIQUE =====
            
            # ===== CONTRÔLE DE VRAISEMBLANCE =====
            # Ce contrôle vérifie la cohérence des soldes selon les règles comptables SYSCOHADA
            # 1. Vérification du sens du solde (débit/crédit) selon la classe de compte
            # 2. Vérification des cas critiques (caisse créditrice, stock créditeur)
            # 3. Vérification des comptes qui doivent être soldés en fin d'exercice
            
            # Compter ce compte comme vérifié pour le contrôle de vraisemblance
            nb_comptes_verifies_vraisemblance += 1
            
            # Variable pour suivre si ce compte a des erreurs de vraisemblance
            compte_erreur_vraisemblance = False
            
            # Récupérer le solde réel pour les autres vérifications (vraisemblance)
            solde_reel = int(x.get("solde_reel", 0) or 0)

            # Vérifier le sens du solde selon la classe de compte
            sign = x.get("sign_solde")
            expect_sign = "D" if solde_reel >= 0 else "C"
            
            # Vérifier le sens attendu selon le plan comptable SYSCOHADA
            sens_attendu = self._get_sens_attendu_compte(numero_compte)
            
            # Vérifications spéciales pour cas critiques
            classe = numero_compte[0] if numero_compte[0].isdigit() else '?'
            
            # CAS CRITIQUE 1 : Caisse créditrice (impossible physiquement)
            if numero_compte.startswith('53') and solde_reel < 0:
                nb_erreurs_signe += 1
                nb_comptes_erreur_signe += 1
                compte_erreur_vraisemblance = True
                erreurs.append({
                    "type": "signe",
                    "numero_compte": numero_compte,
                    "message": f"ERREUR CRITIQUE : Compte de caisse {numero_compte} avec solde CRÉDITEUR ({abs(solde_reel):,} FCFA). C'est IMPOSSIBLE physiquement ! Une caisse ne peut pas avoir un solde créditeur. Vérifiez immédiatement les écritures comptables - erreur certaine."
                })
                continue  # Ne pas continuer avec les autres vérifications
            
            # CAS CRITIQUE 2 : Stock créditeur (anormal - investigation obligatoire)
            if classe == '3' and not numero_compte.startswith('39') and solde_reel < 0:
                nb_erreurs_signe += 1
                nb_comptes_erreur_signe += 1
                compte_erreur_vraisemblance = True
                erreurs.append({
                    "type": "signe",
                    "numero_compte": numero_compte,
                    "message": f"ANOMALIE CRITIQUE : Compte de stock {numero_compte} avec solde CRÉDITEUR ({abs(solde_reel):,} FCFA). Un stock créditeur est ANORMAL et doit être investigué. Vérifiez les écritures d'inventaire, les erreurs de saisie ou les problèmes de valorisation."
                })
                continue  # Ne pas continuer avec les autres vérifications
            
            if sens_attendu != 'BOTH':
                # Le compte a un sens attendu spécifique selon sa classe
                sens_classe = {
                    '1': 'CRÉDITEUR',
                    '2': 'DÉBITEUR',
                    '3': 'DÉBITEUR',
                    '4': 'Variable selon sous-classe',
                    '5': 'DÉBITEUR',
                    '6': 'DÉBITEUR',
                    '7': 'CRÉDITEUR',
                    '8': 'Variable'
                }
                sens_classe_attendu = sens_classe.get(classe, 'Variable')
                
                if sens_attendu == 'D' and solde_reel < 0:
                    # Compte devrait être débiteur mais est créditeur
                    nb_erreurs_signe += 1
                    nb_comptes_erreur_signe += 1
                    compte_erreur_vraisemblance = True
                    # Messages spécifiques selon le type de compte
                    if classe == '3':
                        message = f"ANOMALIE CRITIQUE : Compte de stock {numero_compte} avec solde CRÉDITEUR ({abs(solde_reel):,} FCFA). Un stock créditeur est ANORMAL et doit être investigué. Vérifiez les écritures d'inventaire, les erreurs de saisie ou les problèmes de valorisation."
                    elif classe == '6':
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe} - Charges). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être DÉBITEUR, mais le solde est CRÉDITEUR ({abs(solde_reel):,} FCFA). Un compte de charges créditeur indique généralement une erreur de comptabilisation ou un avoir mal enregistré. Charge créditrice : Généralement anormal, sauf cas spécifiques d'ajustements. Vérifiez les écritures comptables."
                    elif classe == '2':
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe} - Immobilisations). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être DÉBITEUR, mais le solde est CRÉDITEUR ({abs(solde_reel):,} FCFA). Un compte d'immobilisation créditeur est généralement anormal (peut indiquer une erreur ou une cession non comptabilisée correctement). Vérifiez les écritures comptables."
                    else:
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe}). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être DÉBITEUR, mais le solde est CRÉDITEUR ({abs(solde_reel):,} FCFA). Vérifiez la nature du compte et les écritures comptables."
                    
                    erreurs.append({
                        "type": "signe",
                        "numero_compte": numero_compte,
                        "message": message
                    })
                elif sens_attendu == 'C' and solde_reel > 0:
                    # Compte devrait être créditeur mais est débiteur
                    nb_erreurs_signe += 1
                    nb_comptes_erreur_signe += 1
                    compte_erreur_vraisemblance = True
                    # Messages spécifiques selon le type de compte
                    if classe == '1':
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe} - Capitaux). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être CRÉDITEUR, mais le solde est DÉBITEUR ({solde_reel:,} FCFA). Vérifiez la nature du compte (report à nouveau débiteur, associés non appelé, primes de remboursement) et les écritures comptables."
                    elif classe == '2':
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe} - Immobilisations). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être CRÉDITEUR, mais le solde est DÉBITEUR ({solde_reel:,} FCFA). Un compte d'immobilisation créditeur est généralement anormal (peut indiquer une erreur ou une cession non comptabilisée correctement). Vérifiez les écritures comptables."
                    elif classe == '6':
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe} - Charges). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être DÉBITEUR, mais le solde est CRÉDITEUR ({abs(solde_reel):,} FCFA). Un compte de charges créditeur indique généralement une erreur de comptabilisation ou un avoir mal enregistré. Vérifiez les écritures comptables."
                    elif classe == '7':
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe} - Produits). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être CRÉDITEUR, mais le solde est DÉBITEUR ({solde_reel:,} FCFA). Un compte de produits débiteur indique généralement une erreur (annulation mal comptabilisée, avoir sur vente). Produit débiteur : Généralement anormal, sauf cas spécifiques d'ajustements. Vérifiez les écritures comptables."
                    else:
                        message = f"Anomalie de sens pour le compte {numero_compte} (classe {classe}). Le sens normal de la classe {classe} est {sens_classe_attendu}. Ce compte devrait être CRÉDITEUR, mais le solde est DÉBITEUR ({solde_reel:,} FCFA). Vérifiez la nature du compte et les écritures comptables."
                    
                    erreurs.append({
                        "type": "signe",
                        "numero_compte": numero_compte,
                        "message": message
                    })
            
            # Vérifier aussi le signe enregistré
            if sign not in ("D", "C") or sign != expect_sign:
                nb_erreurs_signe += 1
                nb_comptes_erreur_signe += 1
                compte_erreur_vraisemblance = True
                erreurs.append({
                    "type": "signe",
                    "numero_compte": numero_compte,
                    "message": f"Erreur de signe pour le compte {numero_compte}. Solde réel: {solde_reel:,}, Signe attendu: {expect_sign}, Signe trouvé: {sign if sign else 'N/A'}."
                })
            
            # Vérifier si le compte doit être soldé
            doit_etre_solde, gravite, message_solde = self._comptes_doivent_etre_soldes(numero_compte)
            if doit_etre_solde and abs(solde_fin) > 0.01:
                nb_erreurs_comptes_soldes += 1
                nb_comptes_erreur_soldes += 1
                compte_erreur_vraisemblance = True
                erreurs.append({
                    "type": "compte_non_solde",
                    "numero_compte": numero_compte,
                    "gravite": gravite,
                    "message": f"[{gravite}] {message_solde} Compte {numero_compte} a un solde de {solde_fin:,} FCFA. Vérifiez et régularisez."
                })
            
            # Si aucune erreur de vraisemblance détectée, compter comme OK
            if not compte_erreur_vraisemblance:
                nb_comptes_vraisemblance_ok += 1
            # ===== FIN DU CONTRÔLE DE VRAISEMBLANCE =====

        # 3) Vérifier la présence des comptes obligatoires (désactivé : contrôle de complétude retiré)
        # erreurs_presence, classes_presence = self._verifier_presence_comptes_obligatoires(numeros_comptes_valides)
        # if erreurs_presence:
        #     erreurs.extend(erreurs_presence)
        # report["classes_presence"] = classes_presence

        # Ajouter les statistiques de vérification de la formule (second contrôle arithmétique)
        if nb_comptes_verifies_formule > 0:
            statut_formule = "OK" if nb_comptes_formule_erreur == 0 else "ERREUR"
            
            if nb_comptes_formule_erreur == 0:
                explication = f"✅ Le système a vérifié la formule 'Solde de clôture = Solde d'ouverture + Mouvements de période' pour {nb_comptes_verifies_formule} compte(s). Tous les comptes ({nb_comptes_formule_ok}) respectent correctement la formule comptable fondamentale."
            else:
                explication = f"⚠️ Le système a vérifié la formule 'Solde de clôture = Solde d'ouverture + Mouvements de période' pour {nb_comptes_verifies_formule} compte(s). {nb_comptes_formule_ok} compte(s) respectent la formule, mais {nb_comptes_formule_erreur} compte(s) présentent des ERREURS. Les comptes en erreur sont listés ci-dessous avec les détails de l'écart détecté."
            
            report["verification_formule"] = {
                "statut": statut_formule,
                "nb_comptes_verifies": nb_comptes_verifies_formule,
                "nb_comptes_ok": nb_comptes_formule_ok,
                "nb_comptes_erreur": nb_comptes_formule_erreur,
                "explication": explication
            }
            
            print(f"[SECOND CONTRÔLE] Rapport créé : statut={statut_formule}, vérifiés={nb_comptes_verifies_formule}, OK={nb_comptes_formule_ok}, Erreurs={nb_comptes_formule_erreur}")

        # Ajouter les statistiques de vérification de vraisemblance avec structure complète
        # IMPORTANT : La structure (tableaux des classes, comptes à solder) est la même pour toutes les balances
        # car ce sont des règles générales SYSCOHADA. Cependant, les statistiques (erreurs détectées) 
        # sont calculées individuellement pour chaque balance et doivent être différentes.
        if nb_comptes_verifies_vraisemblance > 0:
            statut_vraisemblance = "OK" if (nb_comptes_erreur_signe == 0 and nb_comptes_erreur_soldes == 0) else "ERREUR"
            
            if nb_comptes_erreur_signe == 0 and nb_comptes_erreur_soldes == 0:
                explication = f"✅ Le système a vérifié la vraisemblance des soldes pour {nb_comptes_verifies_vraisemblance} compte(s). Tous les comptes ({nb_comptes_vraisemblance_ok}) respectent les règles comptables SYSCOHADA (sens des soldes et comptes à solder)."
            else:
                explication = f"⚠️ Le système a vérifié la vraisemblance des soldes pour {nb_comptes_verifies_vraisemblance} compte(s). {nb_comptes_vraisemblance_ok} compte(s) sont conformes, mais {nb_comptes_erreur_signe} compte(s) présentent des anomalies de sens et {nb_comptes_erreur_soldes} compte(s) devraient être soldés. Les comptes en erreur sont listés ci-dessous."
            
            # Générer la structure complète du contrôle de vraisemblance avec les erreurs détectées pour cette balance
            # NOTE : La structure de base (règles générales) est identique pour toutes les balances,
            # mais les anomalies détectées sont spécifiques à cette balance
            structure_vraisemblance = self._generer_structure_vraisemblance(erreurs)
            
            report["verification_vraisemblance"] = {
                "statut": statut_vraisemblance,
                "nb_comptes_verifies": nb_comptes_verifies_vraisemblance,
                "nb_comptes_ok": nb_comptes_vraisemblance_ok,
                "nb_comptes_erreur_signe": nb_comptes_erreur_signe,
                "nb_comptes_erreur_soldes": nb_comptes_erreur_soldes,
                "explication": explication,
                "structure": structure_vraisemblance
            }
            
            print(f"[CONTRÔLE VRAISEMBLANCE] Rapport créé pour cette balance : statut={statut_vraisemblance}, vérifiés={nb_comptes_verifies_vraisemblance}, OK={nb_comptes_vraisemblance_ok}, Erreurs signe={nb_comptes_erreur_signe}, Erreurs soldes={nb_comptes_erreur_soldes}")
            print(f"[CONTRÔLE VRAISEMBLANCE] ⚠️ NOTE : La structure (tableaux des classes, comptes à solder) est identique pour toutes les balances car ce sont des règles générales SYSCOHADA.")
            print(f"[CONTRÔLE VRAISEMBLANCE] ⚠️ Les statistiques ci-dessus sont spécifiques à cette balance et peuvent différer d'une balance à l'autre.")

        # Récap totaux pour UI
        report["totaux"] = {
            "debit_fin": sum_deb_fin,
            "credit_fin": sum_cre_fin,
            "nb_erreurs_identite": nb_erreurs_identite,
            "nb_erreurs_signe": nb_erreurs_signe,
            "nb_erreurs_arithmetique": nb_erreurs_arithmetique,
            "nb_erreurs_comptes_soldes": nb_erreurs_comptes_soldes,
            "nb_erreurs_total": nb_erreurs_identite + nb_erreurs_signe + nb_erreurs_arithmetique + nb_erreurs_comptes_soldes,
            "nb_comptes_dupliques": len(comptes_dupliques),
            "nb_comptes_verifies_formule": nb_comptes_verifies_formule,
            "nb_comptes_formule_ok": nb_comptes_formule_ok,
            "nb_comptes_formule_erreur": nb_comptes_formule_erreur
        }
        
        # Détails des erreurs avec numéros de comptes
        report["erreurs"] = erreurs
        
        # DEBUG : Afficher le résumé final
        print(f"\n{'='*60}")
        print(f"[RAPPORT FINAL] RÉSUMÉ DU CONTRÔLE DE COHÉRENCE")
        print(f"{'='*60}")
        print(f"Nombre total d'erreurs détectées : {len(erreurs)}")
        print(f"Nombre de comptes vérifiés (formule) : {nb_comptes_verifies_formule}")
        print(f"  - Comptes OK : {nb_comptes_formule_ok}")
        print(f"  - Comptes en erreur : {nb_comptes_formule_erreur}")
        print(f"\nErreurs par type :")
        erreurs_par_type = {}
        for e in erreurs:
            t = e.get("type", "autre")
            erreurs_par_type[t] = erreurs_par_type.get(t, 0) + 1
        for t, count in erreurs_par_type.items():
            print(f"  - {t}: {count}")
        print(f"\nEquilibre global : {report.get('equilibre_global', 'N/A')}")
        print(f"{'='*60}\n")

        return report

    def controle_coherence(self, id_mission):
        print(f"=== DÉBUT contrôle_coherence pour mission {id_mission} ===")
        print(f"⚠️ IMPORTANT : Le contrôle utilise les balances DÉJÀ IMPORTÉES dans la base de données")
        print(f"⚠️ Les modifications du code s'appliquent automatiquement aux balances existantes")
        
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        print(f"Mission trouvée : {mission is not None}")
        
        if not mission:
            print(f"ERREUR : Mission {id_mission} non trouvée")
            return {"error": "Mission non trouvée"}
        
        bal_ids = mission.get("balances", [])
        print(f"Balances trouvées : {len(bal_ids)} - IDs: {bal_ids}")
        
        if not bal_ids:
            print("ERREUR : Aucune balance trouvée pour cette mission")
            return {"error": "Aucune balance trouvée"}
        
        out = {}

        # Récupérer l'année auditée de la mission
        annee_auditee = int(mission.get("annee_auditee", 0))

        # IMPORTANT : Le contrôle recalcule TOUJOURS avec le code actuel
        # Il utilise les balances DÉJÀ IMPORTÉES - PAS BESOIN DE RÉIMPORTER
        for idx, bal_id in enumerate(bal_ids):
            # Calculer l'année de chaque balance
            annee_balance = annee_auditee - idx
            print(f"Traitement balance {idx} (année {annee_balance}) : {bal_id}")
            
            # Charger les données depuis la base de données (balances déjà importées)
            lines = self._load_balance(bal_id)
            print(f"Lignes de balance chargées depuis la base : {len(lines)}")
            
            # Exécuter le contrôle avec le code actuel (inclut toutes les modifications récentes)
            rapport = self._coherence_checks_for_year(lines)
            rapport["annee"] = annee_balance
            
            out[str(annee_balance)] = rapport

        print(f"Rapport final généré avec le code actuel : {out}")

        # Sauvegarder le nouveau rapport (avec les modifications appliquées)
        db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"controle_coherence": out}})
        print("=== FIN contrôle_coherence ===")
        print(f"✅ Le contrôle a été recalculé avec le code actuel. Les balances n'ont PAS besoin d'être réimportées.")
        return out

    def _index_by_compte(self, lines):
        # Filtrer les lignes qui ont un numero_compte valide (non vide, non None)
        index = {}
        lignes_ignorees = 0
        lignes_ignorees_none = 0
        lignes_ignorees_vide = 0
        
        if not lines:
            print("[INDEX_BY_COMPTE] liste de lignes vide")
            return index
        
        print(f"[INDEX_BY_COMPTE] traitement de {len(lines)} lignes")
        
        for idx, x in enumerate(lines):
            if not x:
                lignes_ignorees += 1
                continue
                
            num_compte = x.get("numero_compte")
            
            # Debug pour les premières lignes
            if idx < 3:
                print(f"  📋 Ligne {idx}: numero_compte={num_compte} (type: {type(num_compte)}, bool: {bool(num_compte)})")
            
            # Gérer le cas où num_compte pourrait être 0 (qui est False mais valide comme numéro)
            if num_compte is None:
                lignes_ignorees_none += 1
                continue
            
            # Convertir en string et nettoyer (inclure "0" comme numéro valide)
            num_str = str(num_compte).strip()
            
            if not num_str or num_str == "None" or num_str.lower() == "nan":
                lignes_ignorees_vide += 1
                continue
                
            # Ajouter au index
            index[num_str] = x
        
        print(f"[INDEX_BY_COMPTE] {len(index)} comptes indexes")
        if lignes_ignorees > 0:
            print(f"   ⚠️  Lignes None ignorées: {lignes_ignorees}")
        if lignes_ignorees_none > 0:
            print(f"   ⚠️  Numéros None ignorés: {lignes_ignorees_none}")
        if lignes_ignorees_vide > 0:
            print(f"   ⚠️  Numéros vides ignorés: {lignes_ignorees_vide}")
        
        return index

    def controle_intangibilite(self, id_mission):
        try:
            print(f"[CONTROLE_INTANGIBILITE] ========== DEBUT ==========")
            print(f"[CONTROLE_INTANGIBILITE] Mission ID: {id_mission}")
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                print(f"[CONTROLE_INTANGIBILITE] ERREUR: Mission non trouvee")
                return {"ok": False, "message": "Mission non trouvée.", "comptes": []}
                
            bal_ids = mission.get("balances", [])
            if len(bal_ids) < 2:
                print(f"[CONTROLE_INTANGIBILITE] ERREUR: Pas assez de balances ({len(bal_ids)})")
                return {"ok": False, "message": "Il faut au moins N et N-1.", "comptes": []}
            
            print(f"[CONTROLE_INTANGIBILITE] Nombre de balances: {len(bal_ids)}")
            print(f"[CONTROLE_INTANGIBILITE] Balance IDs: {bal_ids}")

            # Récupérer l'année auditée pour déterminer les périodes
            annee_auditee = int(mission.get("annee_auditee", 0)) if mission.get("annee_auditee") else 0
            
            # Charger les balances avec leurs métadonnées
            balance_docs = []
            for bal_id in bal_ids:
                bal_doc = db.Balance.find_one({"_id": ObjectId(bal_id)})
                if bal_doc:
                    # Utiliser annee_balance au lieu de periode
                    annee_balance = bal_doc.get("annee_balance") or bal_doc.get("periode") or ""
                    balance_docs.append({
                        "id": bal_id,
                        "annee": annee_balance,
                        "data": bal_doc.get("balance", [])
                    })
            
            # Si aucune année n'est trouvée, utiliser l'ordre des balances et calculer les années
            if all(not bd.get("annee") for bd in balance_docs) and annee_auditee:
                # Calculer les années basées sur annee_auditee
                for idx, bd in enumerate(balance_docs):
                    bd["annee"] = annee_auditee - idx
            
            # Trier les balances par année (du plus ancien au plus récent)
            balance_docs.sort(key=lambda x: x.get("annee", 0))
            
            # Identifier N et N-1
            if len(balance_docs) < 2:
                return {"ok": False, "message": "Il faut au moins N et N-1.", "periodes": {}, "comptes": []}
            
            bal_N = balance_docs[-1]["data"]  # Dernière année = N
            bal_N1 = balance_docs[-2]["data"]  # Avant-dernière = N-1
            periode_N = str(balance_docs[-1]["annee"]) if balance_docs[-1]["annee"] else "N"
            periode_N1 = str(balance_docs[-2]["annee"]) if balance_docs[-2]["annee"] else "N-1"
            
            # Vérifier que les balances contiennent des données
            if not bal_N or len(bal_N) == 0:
                return {"ok": False, "message": f"La balance N ({periode_N}) ne contient aucune donnée.", "periodes": {"N": periode_N, "N-1": periode_N1}, "comptes": []}
            if not bal_N1 or len(bal_N1) == 0:
                return {"ok": False, "message": f"La balance N-1 ({periode_N1}) ne contient aucune donnée.", "periodes": {"N": periode_N, "N-1": periode_N1}, "comptes": []}
            
            print(f"📊 Lignes brutes dans balance N: {len(bal_N)}")
            print(f"📊 Lignes brutes dans balance N-1: {len(bal_N1)}")
            
            # Vérifier les premières lignes pour debug
            if len(bal_N) > 0:
                print(f"📋 Exemple ligne N (première): {list(bal_N[0].keys()) if bal_N[0] else 'vide'}")
                if bal_N[0] and 'numero_compte' in bal_N[0]:
                    num_compte = bal_N[0].get('numero_compte')
                    print(f"📋 Premier numero_compte N: '{num_compte}' (type: {type(num_compte)}, value: {repr(num_compte)})")
                else:
                    print(f"📋 Première ligne N ne contient pas 'numero_compte': {bal_N[0].keys() if bal_N[0] else 'ligne vide'}")
            
            if len(bal_N1) > 0:
                if bal_N1[0] and 'numero_compte' in bal_N1[0]:
                    num_compte = bal_N1[0].get('numero_compte')
                    print(f"📋 Premier numero_compte N-1: '{num_compte}' (type: {type(num_compte)}, value: {repr(num_compte)})")
            
            print(f"[CONTROLE_INTANGIBILITE] Indexation des comptes...")
            print(f"[CONTROLE_INTANGIBILITE] Avant indexation: bal_N contient {len(bal_N)} lignes, bal_N1 contient {len(bal_N1)} lignes")
            
            idxN = self._index_by_compte(bal_N)
            idxN1 = self._index_by_compte(bal_N1)
            
            print(f"[CONTROLE_INTANGIBILITE] APRES indexation:")
            print(f"[CONTROLE_INTANGIBILITE]    - Comptes indexes dans N: {len(idxN)}")
            print(f"[CONTROLE_INTANGIBILITE]    - Comptes indexes dans N-1: {len(idxN1)}")
            print(f"[CONTROLE_INTANGIBILITE] idxN type: {type(idxN)}, idxN1 type: {type(idxN1)}")

            # Limiter aux classes 1 à 5 uniquement
            def _is_class_1_to_5(account_number):
                try:
                    s = str(account_number).strip()
                    return len(s) > 0 and s[0] in ("1", "2", "3", "4", "5")
                except Exception:
                    return False

            idxN = {k: v for k, v in idxN.items() if _is_class_1_to_5(k)}
            idxN1 = {k: v for k, v in idxN1.items() if _is_class_1_to_5(k)}

            print(f"[CONTROLE_INTANGIBILITE] Filtrage classes 1-5:")
            print(f"[CONTROLE_INTANGIBILITE]    - Comptes retenus en N: {len(idxN)}")
            print(f"[CONTROLE_INTANGIBILITE]    - Comptes retenus en N-1: {len(idxN1)}")
            
            # DIAGNOSTIC: Vérifier si l'indexation a échoué inopinément
            if len(bal_N) > 0 and len(idxN) == 0:
                print(f"⚠️  PROBLÈME: bal_N contient {len(bal_N)} lignes mais idxN est vide!")
                print(f"   - Type de bal_N[0]: {type(bal_N[0])}")
                if len(bal_N) > 0 and isinstance(bal_N[0], dict):
                    print(f"   - Clés de la première ligne: {list(bal_N[0].keys())}")
                    print(f"   - numero_compte de la première ligne: {bal_N[0].get('numero_compte')}")
            
            if len(bal_N1) > 0 and len(idxN1) == 0:
                print(f"⚠️  PROBLÈME: bal_N1 contient {len(bal_N1)} lignes mais idxN1 est vide!")
                print(f"   - Type de bal_N1[0]: {type(bal_N1[0])}")
                if len(bal_N1) > 0 and isinstance(bal_N1[0], dict):
                    print(f"   - Clés de la première ligne: {list(bal_N1[0].keys())}")
                    print(f"   - numero_compte de la première ligne: {bal_N1[0].get('numero_compte')}")
            
            if len(idxN) > 0:
                print(f"📋 Exemple compte indexé N: {list(idxN.keys())[:3]}")
            if len(idxN1) > 0:
                print(f"📋 Exemple compte indexé N-1: {list(idxN1.keys())[:3]}")
            
            # Vérifier que les index contiennent des comptes
            # MODIFICATION: Ne pas retourner si au moins une balance a des comptes
            # On peut faire le contrôle même si une seule balance a des comptes
            if len(idxN) == 0 and len(idxN1) == 0:
                print("❌ Aucun compte trouvé dans les balances")
                print(f"   - Balance N: {len(bal_N)} lignes brutes, mais 0 comptes indexés")
                print(f"   - Balance N-1: {len(bal_N1)} lignes brutes, mais 0 comptes indexés")
                print(f"   - Vérifiez que les lignes contiennent bien un champ 'numero_compte' avec une valeur non vide")
                print(f"   - DEBUG: Type de bal_N[0] si existe: {type(bal_N[0]) if bal_N and len(bal_N) > 0 else 'N/A'}")
                print(f"   - DEBUG: Type de bal_N1[0] si existe: {type(bal_N1[0]) if bal_N1 and len(bal_N1) > 0 else 'N/A'}")
                
                # Analyser pourquoi aucun compte n'a été trouvé
                if len(bal_N) > 0:
                    premiere_ligne = bal_N[0]
                    print(f"   - Exemple ligne N: {premiere_ligne}")
                    if 'numero_compte' not in premiere_ligne:
                        print(f"   - ⚠️  Le champ 'numero_compte' est absent de la première ligne")
                    elif premiere_ligne.get('numero_compte') is None:
                        print(f"   - ⚠️  Le champ 'numero_compte' est None")
                    elif str(premiere_ligne.get('numero_compte')).strip() == "":
                        print(f"   - ⚠️  Le champ 'numero_compte' est une chaîne vide")
                
                return {
                    "ok": False, 
                    "message": f"Aucun compte trouvé dans les balances N et N-1. Les balances contiennent {len(bal_N)} et {len(bal_N1)} lignes brutes, mais aucun numéro de compte valide n'a été détecté. Vérifiez que les balances contiennent bien un champ 'numero_compte' avec des valeurs non vides.", 
                    "periodes": {"N": periode_N, "N-1": periode_N1}, 
                    "comptes": []
                }

            # Créer une liste de tous les comptes (sans filtrage sur les classes)
            tous_comptes = []
            
            print(f"🔍 Début du traitement: {len(idxN)} comptes en N, {len(idxN1)} comptes en N-1")
            
            if len(idxN) == 0:
                print(f"⚠️  ATTENTION: idxN est vide même si bal_N contient {len(bal_N)} lignes")
            
            if len(idxN1) == 0:
                print(f"⚠️  ATTENTION: idxN1 est vide même si bal_N1 contient {len(bal_N1)} lignes")
            
            comptes_ajoutes = 0
            comptes_erreur = 0
            
            # 1. Traiter tous les comptes présents en N
            for num, ln in idxN.items():
                try:
                    # Pour l'ouverture N, utiliser debit_initial et credit_initial de N
                    di = float(ln.get("debit_initial", 0) or 0)
                    ci = float(ln.get("credit_initial", 0) or 0)
                    ouvN = di - ci

                    prev = idxN1.get(num)
                    if prev:
                        # Pour la clôture N-1, utiliser debit_fin et credit_fin de N-1
                        df = float(prev.get("debit_fin", 0) or 0)
                        cf = float(prev.get("credit_fin", 0) or 0)
                        clotN1 = df - cf
                        ecart = clotN1 - ouvN  # Écart = Clôture N-1 - Ouverture N
                        
                        # Afficher les 5 premiers comptes pour debug
                        if len(tous_comptes) < 5:
                            print(f"  📋 Compte {num}: ouvN={ouvN} (di={di}-ci={ci}), clotN1={clotN1} (df={df}-cf={cf}), ecart={ecart}")
                        
                        # Ajouter tous les comptes, pas seulement ceux avec des écarts
                        tous_comptes.append({
                            "numero_compte": num,
                            "libelle": ln.get("libelle", ""),
                            "ouverture_n": ouvN,
                            "cloture_n1": clotN1,
                            "ecart": ecart,
                            "status": "ecart" if ecart != 0 else "ok",
                            "message": f"Clôture N-1 {clotN1} ≠ Ouverture N {ouvN}" if ecart != 0 else f"Clôture N-1 {clotN1} = Ouverture N {ouvN}",
                            "justification": f"Écart de {ecart} entre la clôture de l'exercice N-1 ({clotN1}) et l'ouverture de l'exercice N ({ouvN})." if ecart != 0 else "Aucun écart détecté.",
                            "conclusion_audit": "Écart significatif détecté - Nécessite une justification et une documentation des causes de cette variation." if ecart != 0 else "Aucune anomalie détectée."
                        })
                        comptes_ajoutes += 1
                    else:
                        # Compte nouveau (présent en N, absent en N-1)
                        ecart = -ouvN  # Écart = 0 - Ouverture N (car absent en N-1)
                        tous_comptes.append({
                            "numero_compte": num,
                            "libelle": ln.get("libelle", ""),
                            "ouverture_n": ouvN,
                            "cloture_n1": None,
                            "ecart": ecart,
                            "status": "nouveau",
                            "message": "Compte présent en N mais absent en N-1",
                            "justification": f"Le compte {num} est présent dans l'exercice N avec un solde d'ouverture de {ouvN}, mais n'existait pas dans l'exercice N-1. Cela peut indiquer une création de compte, un reclassement ou une erreur de saisie.",
                            "conclusion_audit": "Compte nouvellement créé ou reclassé - Vérifier la légitimité de cette création et documenter les raisons."
                        })
                        comptes_ajoutes += 1
                except Exception as e:
                    comptes_erreur += 1
                    print(f"⚠️  Erreur lors du traitement du compte {num}: {e}")
                    if comptes_erreur <= 3:
                        import traceback
                        traceback.print_exc()
                    continue
            
            print(f"📊 Après traitement des comptes N: {comptes_ajoutes} comptes ajoutés, {comptes_erreur} erreurs")
            
            # 2. Ajouter les comptes présents en N-1 mais absents en N
            for num, ln in idxN1.items():
                try:
                    # Si le compte n'existe pas en N, c'est un écart
                    if num not in idxN:
                        df = float(ln.get("debit_fin", 0) or 0)
                        cf = float(ln.get("credit_fin", 0) or 0)
                        clotN1 = df - cf
                        
                        tous_comptes.append({
                            "numero_compte": num,
                            "libelle": ln.get("libelle", ""),
                            "ouverture_n": None,
                            "cloture_n1": clotN1,
                            "ecart": clotN1,  # Écart = Clôture N-1 - 0 (car absent en N)
                            "status": "supprime",
                            "message": "Compte présent en N-1 mais absent en N",
                            "justification": f"Le compte {num} était présent dans l'exercice N-1 avec un solde de clôture de {clotN1}, mais n'existe plus dans l'exercice N. Cela peut indiquer une suppression de compte, un reclassement ou une erreur de saisie.",
                            "conclusion_audit": "Compte supprimé ou reclassé - Vérifier la légitimité de cette suppression et documenter les raisons."
                        })
                        comptes_ajoutes += 1
                except Exception as e:
                    comptes_erreur += 1
                    print(f"⚠️  Erreur lors du traitement du compte N-1 {num}: {e}")
                    continue
            
            print(f"📊 Après traitement des comptes N-1: {len(tous_comptes)} comptes au total")

            # Trier par numéro de compte
            tous_comptes.sort(key=lambda x: x["numero_compte"])
            
            # Compter les écarts
            ecarts_count = len([c for c in tous_comptes if c["status"] in ["ecart", "nouveau", "supprime"]])
            
            print(f"📊 Résumé final: {len(tous_comptes)} comptes traités")
            print(f"   - Comptes OK: {len([c for c in tous_comptes if c['status'] == 'ok'])}")
            print(f"   - Comptes avec écart: {len([c for c in tous_comptes if c['status'] == 'ecart'])}")
            print(f"   - Comptes nouveaux: {len([c for c in tous_comptes if c['status'] == 'nouveau'])}")
            print(f"   - Comptes supprimés: {len([c for c in tous_comptes if c['status'] == 'supprime'])}")
            
            # Si aucun compte n'a été trouvé, retourner un message d'aide
            if len(tous_comptes) == 0:
                print("⚠️  Aucun compte trouvé dans les balances")
                return {
                    "ok": False,
                    "message": "Aucun compte trouvé dans les balances N et N-1. Vérifiez que les balances contiennent bien des données avec des numéros de compte valides.",
                    "periodes": {"N": periode_N, "N-1": periode_N1},
                    "total_comptes": 0,
                    "ecarts_count": 0,
                    "comptes": []
                }
            
            report = {
                "ok": ecarts_count == 0, 
                "total_comptes": len(tous_comptes),
                "ecarts_count": ecarts_count,
                "periodes": {
                    "N": periode_N,
                    "N-1": periode_N1
                },
                "comptes": tous_comptes
            }
            
            print(f"✅ Rapport généré: {len(tous_comptes)} comptes, {ecarts_count} écarts")
            print(f"📋 Clés du rapport: {list(report.keys())}")
            print(f"📋 Nombre de comptes dans le rapport: {len(report.get('comptes', []))}")
            
            # CRITIQUE: Vérifier que tous_comptes n'est pas vide avant de sauvegarder
            if len(tous_comptes) == 0:
                print(f"❌ ERREUR CRITIQUE: tous_comptes est vide avant sauvegarde!")
                print(f"   - idxN avait {len(idxN)} comptes")
                print(f"   - idxN1 avait {len(idxN1)} comptes")
                print(f"   - Vérifiez les logs ci-dessus pour comprendre pourquoi tous_comptes est vide")
                # Ne pas sauvegarder un rapport vide, retourner une erreur explicite
                return {
                    "ok": False,
                    "message": f"Erreur lors du traitement des comptes: {len(idxN)} comptes indexés en N et {len(idxN1)} en N-1, mais aucun compte n'a été ajouté à la liste finale. Vérifiez les logs du serveur.",
                    "periodes": {"N": periode_N, "N-1": periode_N1},
                    "total_comptes": 0,
                    "ecarts_count": 0,
                    "comptes": []
                }
            
            db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"controle_intangibilite": report}})
            print(f"[CONTROLE_INTANGIBILITE] Rapport sauvegarde dans la base de donnees")
            print(f"[CONTROLE_INTANGIBILITE] Rapport retourne: total_comptes={report.get('total_comptes')}, ecarts={report.get('ecarts_count')}")
            print(f"[CONTROLE_INTANGIBILITE] ========== FIN (SUCCES) ==========")
            return report
            
        except Exception as e:
            print(f"[CONTROLE_INTANGIBILITE] ERREUR EXCEPTION: {str(e)}")
            import traceback
            traceback.print_exc()
            print(f"[CONTROLE_INTANGIBILITE] ========== FIN (ERREUR) ==========")
            return {"ok": False, "message": f"Erreur lors du contrôle d'intangibilité: {str(e)}", "comptes": []}

    def classement_bilan(self, id_mission):
        try:
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "classement": []}

            balances = mission.get("balances", [])
            if len(balances) < 2:
                return {"ok": False, "message": "Balances manquantes", "classement": []}

            balance_n = db.Balance.find_one({"_id": ObjectId(balances[0])}, {"balance": 1})
            balance_n1 = db.Balance.find_one({"_id": ObjectId(balances[1])}, {"balance": 1})

            if not balance_n or not balance_n1:
                return {"ok": False, "message": "Balances introuvables", "classement": []}

            # Charger le fichier de grouping pour obtenir la structure de classement
            grouping_path = os.path.join(os.path.dirname(__file__), "..", "grouping.json")
            with open(grouping_path, 'r', encoding='utf-8') as file:
                grouping_data = json.load(file)
            
            # Utiliser le référentiel syscohada par défaut
            referentiel = mission.get("referentiel", "syscohada")
            table_grouping = grouping_data.get(referentiel, grouping_data.get("syscohada", []))

            # Récupérer les données de balance
            balance_n_data = balance_n.get("balance", [])
            balance_n1_data = balance_n1.get("balance", [])

            # Créer un index pour la balance N-1
            balance_n1_index = {str(item.get("numero_compte", "")): item for item in balance_n1_data}

            # Identifier tous les préfixes de comptes dans les balances (2 premiers chiffres)
            prefixes_trouves = set()
            for item_n in balance_n_data:
                numero_compte = str(item_n.get("numero_compte", "")).strip()
                if numero_compte and len(numero_compte) >= 2:
                    prefixe = numero_compte[:2]
                    if prefixe.isdigit():
                        prefixes_trouves.add(prefixe)

            # Créer un dictionnaire des groupes existants par préfixe
            groupes_existants = {group['compte']: group for group in table_grouping}

            # Créer des groupes automatiques pour les préfixes non définis
            prefixes_manquants = prefixes_trouves - set(groupes_existants.keys())
            for prefixe in sorted(prefixes_manquants):
                # Déterminer la nature basée sur le préfixe
                if prefixe.startswith(('1', '2', '3', '4', '5')):
                    nature = "bilan"
                else:
                    nature = "pnl"
                
                nouveau_groupe = {
                    "compte": prefixe,
                    "nature": nature,
                    "libelle": f"AUTRES - COMPTE {prefixe}"
                }
                table_grouping.append(nouveau_groupe)

            # Trier les groupes par numéro de compte
            table_grouping.sort(key=lambda x: x['compte'])

            classement = []
            for group in table_grouping:
                compte_prefix = str(group.get("compte", ""))
                nature = group.get("nature", "bilan")
                libelle = group.get("libelle", "")

                # Calculer les totaux pour ce groupe
                solde_n = 0
                solde_n1 = 0
                comptes_detaille = []

                for item_n in balance_n_data:
                    numero_compte = str(item_n.get("numero_compte", ""))
                    if numero_compte.startswith(compte_prefix):
                        item_n1 = balance_n1_index.get(numero_compte, {})
                        
                        # Calculer les soldes
                        solde_item_n = (item_n.get("debit_fin", 0) or 0) - (item_n.get("credit_fin", 0) or 0)
                        solde_item_n1 = (item_n1.get("debit_fin", 0) or 0) - (item_n1.get("credit_fin", 0) or 0)
                        
                        solde_n += solde_item_n
                        solde_n1 += solde_item_n1

                        comptes_detaille.append({
                            "numero_compte": numero_compte,
                            "libelle": item_n.get("libelle", ""),
                            "solde_n": solde_item_n,
                            "solde_n1": solde_item_n1,
                            "variation": solde_item_n - solde_item_n1
                        })

                # Calculer la variation
                variation = solde_n - solde_n1
                variation_percent = 0
                if solde_n1 != 0:
                    variation_percent = (variation / abs(solde_n1)) * 100

                classement.append({
                    "compte": compte_prefix,
                    "libelle": libelle,
                    "nature": nature,
                    "solde_n": solde_n,
                    "solde_n1": solde_n1,
                    "variation": variation,
                    "variation_percent": variation_percent,
                    "comptes_detaille": comptes_detaille
                })

            # Sauvegarder le classement dans la mission
            report = {
                "ok": True,
                "message": f"Classement généré avec {len(classement)} groupes",
                "classement": classement,
                "referentiel": referentiel
            }
            db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"classement_bilan": report}})
            return report

        except Exception as e:
            return {"ok": False, "message": str(e), "classement": []}

    def etats_financiers_preliminaires(self, id_mission):
        try:
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "efi": {}}

            balances = mission.get("balances", [])
            if len(balances) < 2:
                return {"ok": False, "message": "Balances manquantes", "efi": {}}

            balance_n = db.Balance.find_one({"_id": ObjectId(balances[0])}, {"balance": 1})
            balance_n1 = db.Balance.find_one({"_id": ObjectId(balances[1])}, {"balance": 1})

            if not balance_n or not balance_n1:
                return {"ok": False, "message": "Balances introuvables", "efi": {}}

            # Charger le fichier de mapping EFI
            mapping_path = os.path.join(os.path.dirname(__file__), "..", "mapping_efi.json")
            with open(mapping_path, 'r', encoding='utf-8') as file:
                mapping_data = json.load(file)
            
            mapping = mapping_data.get("structure", [])

            # Récupérer les données de balance
            balance_n_data = balance_n.get("balance", [])
            balance_n1_data = balance_n1.get("balance", [])

            # Générer les états financiers en utilisant la méthode existante prod_efi
            efi_data = self.prod_efi(balance_n_data, balance_n1_data, [])

            # prod_efi retourne déjà un dictionnaire organisé par nature (actif, passif, pnl)
            # Utiliser directement les données organisées
            efi_organized = {
                "actif": efi_data.get("actif", []),
                "passif": efi_data.get("passif", []),
                "pnl": efi_data.get("pnl", [])
            }
            
            # Log pour vérifier le nombre de lignes générées
            print(f"📊 États financiers générés - Actif: {len(efi_organized.get('actif', []))} lignes, Passif: {len(efi_organized.get('passif', []))} lignes, PNL: {len(efi_organized.get('pnl', []))} lignes")

            # Sauvegarder les états financiers dans la mission
            report = {
                "ok": True,
                "message": f"États financiers générés avec succès",
                "efi": efi_organized,
                "annee_auditee": mission.get("annee_auditee", [])[0] if mission.get("annee_auditee") else None
            }
            db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"etats_financiers_preliminaires": report}})
            return report

        except Exception as e:
            return {"ok": False, "message": str(e), "efi": {}}

    def analyse_quantitative(self, id_mission):
        try:
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "analyse": []}

            # Récupérer les données de grouping
            grouping = mission.get("grouping", [])
            if not grouping:
                return {"ok": False, "message": "Données de grouping manquantes", "analyse": []}

            # Récupérer les données de matérialité
            materiality = mission.get("materiality", [])
            selected_materiality = next((mat for mat in materiality if mat.get('choice')), None)

            if not selected_materiality:
                return {"ok": False, "message": "Aucune matérialité sélectionnée", "analyse": []}

            # Analyser chaque compte du grouping
            analyse_data = []
            for item in grouping:
                solde_n = abs(item.get('solde_n', 0))
                solde_n1 = abs(item.get('solde_n1', 0))
                materiality_threshold = selected_materiality.get('materiality', 0)
                
                # Déterminer si le compte est significatif
                is_significant = solde_n >= materiality_threshold
                
                # Calculer le pourcentage par rapport au seuil
                percentage_of_threshold = 0
                if materiality_threshold > 0:
                    percentage_of_threshold = (solde_n / materiality_threshold) * 100

                analyse_data.append({
                    "compte": item.get('compte', ''),
                    "libelle": item.get('libelle', ''),
                    "solde_n": solde_n,
                    "solde_n1": solde_n1,
                    "variation": solde_n - solde_n1,
                    "materiality_threshold": materiality_threshold,
                    "is_significant": is_significant,
                    "percentage_of_threshold": percentage_of_threshold,
                    "status": "À tester" if is_significant else "Ne pas tester"
                })

            # Trier par solde décroissant
            analyse_data.sort(key=lambda x: x['solde_n'], reverse=True)

            # Statistiques
            total_accounts = len(analyse_data)
            significant_accounts = len([a for a in analyse_data if a['is_significant']])
            total_significant_amount = sum([a['solde_n'] for a in analyse_data if a['is_significant']])

            report = {
                "ok": True,
                "message": f"Analyse quantitative générée avec succès",
                "analyse": analyse_data,
                "statistics": {
                    "total_accounts": total_accounts,
                    "significant_accounts": significant_accounts,
                    "non_significant_accounts": total_accounts - significant_accounts,
                    "total_significant_amount": total_significant_amount,
                    "materiality_threshold": selected_materiality.get('materiality', 0),
                    "materiality_benchmark": selected_materiality.get('benchmark', '')
                }
            }

            # Sauvegarder l'analyse dans la mission
            db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"analyse_quantitative": report}})
            return report

        except Exception as e:
            return {"ok": False, "message": str(e), "analyse": []}

    def analyse_qualitative(self, id_mission):
        try:
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "analyse": []}

            # Récupérer les données de grouping
            grouping = mission.get("grouping", [])
            if not grouping:
                return {"ok": False, "message": "Données de grouping manquantes", "analyse": []}

            # Questions Q1-Q8 pour l'analyse qualitative
            questions = [
                "Volume d'activité, complexité et homogénéité des transactions enregistrées, existence de transactions significatives inhabituelles ou anormales dans le COTABD",
                "Changements identifiés dans le COTABD et détermination si un ou de nouveaux risque(s) sont apparus du fait de changement au sein de l'entité ou de son environnement (économique, légal, réglementaire, normatif ou méthodes comptables)",
                "Sensibilité de l'entité aux anomalies issues de fraudes (Si oui, le risque est obligatoirement Significant)",
                "Niveau de complexité des normes, règles, méthodes comptables, notes annexes, estimations ou jugements liées aux comptes ou aux notes annexes",
                "Exposition du COTABD à des pertes (charges ou dépréciations)",
                "Probabilité que des passifs éventuels significatifs (procès, contentieux, litiges etc…) puissent être issus des transactions enregistrées dans le COTABD",
                "Existence de transactions avec des parties liées dans le COTABD",
                "Niveau de contrôle interne et fiabilité des systèmes d'information liés aux comptes"
            ]

            # Récupérer les réponses existantes si disponibles
            qualitative_responses = mission.get("qualitative_responses", {})
            
            # Préparer les données d'analyse
            analyse_data = []
            for item in grouping:
                compte = item.get('compte', '')
                libelle = item.get('libelle', '')
                solde_n = abs(item.get('solde_n', 0))
                solde_n1 = abs(item.get('solde_n1', 0))
                
                # Récupérer les réponses pour ce compte (initialiser à False si pas de réponses)
                compte_responses = qualitative_responses.get(compte, {})
                
                # S'assurer que toutes les questions Q1-Q8 sont initialisées
                for q in range(1, len(questions) + 1):
                    if f'Q{q}' not in compte_responses:
                        compte_responses[f'Q{q}'] = False
                
                # Calculer le score qualitatif
                total_questions = len(questions)
                positive_responses = sum(1 for q in range(1, total_questions + 1) if compte_responses.get(f'Q{q}', False))
                qualitative_score = (positive_responses / total_questions) * 100
                
                # Déterminer si le compte est significatif qualitativement (logique à 3 niveaux)
                if qualitative_score >= 50:
                    is_qualitatively_significant = True
                elif qualitative_score >= 25:
                    is_qualitatively_significant = False  # À évaluer mais pas significatif
                else:
                    is_qualitatively_significant = False
                
                # Préparer les réponses détaillées
                responses_detail = []
                for i, question in enumerate(questions, 1):
                    responses_detail.append({
                        "question_id": f"Q{i}",
                        "question_text": question,
                        "response": compte_responses.get(f'Q{i}', False),
                        "is_positive": compte_responses.get(f'Q{i}', False)
                    })

                analyse_data.append({
                    "compte": compte,
                    "libelle": libelle,
                    "solde_n": solde_n,
                    "solde_n1": solde_n1,
                    "variation": solde_n - solde_n1,
                    "qualitative_score": qualitative_score,
                    "positive_responses": positive_responses,
                    "total_questions": total_questions,
                    "is_qualitatively_significant": is_qualitatively_significant,
                    "responses_detail": responses_detail,
                    "status": "À tester" if is_qualitatively_significant else "Ne pas tester"
                })

            # Trier par score qualitatif décroissant
            analyse_data.sort(key=lambda x: x['qualitative_score'], reverse=True)

            # Statistiques
            total_accounts = len(analyse_data)
            significant_accounts = len([a for a in analyse_data if a['is_qualitatively_significant']])
            total_positive_responses = sum([a['positive_responses'] for a in analyse_data])

            report = {
                "ok": True,
                "message": f"Analyse qualitative générée avec succès",
                "analyse": analyse_data,
                "questions": questions,
                "statistics": {
                    "total_accounts": total_accounts,
                    "significant_accounts": significant_accounts,
                    "non_significant_accounts": total_accounts - significant_accounts,
                    "total_positive_responses": total_positive_responses,
                    "average_score": sum([a['qualitative_score'] for a in analyse_data]) / total_accounts if total_accounts > 0 else 0
                }
            }

            # Sauvegarder l'analyse dans la mission
            db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"analyse_qualitative": report}})
            return report

        except Exception as e:
            return {"ok": False, "message": str(e), "analyse": []}

    def synthese_comptes_significatifs(self, id_mission):
        try:
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "synthese": []}

            # Récupérer les données de grouping
            grouping = mission.get("grouping", [])
            if not grouping:
                return {"ok": False, "message": "Données de grouping manquantes", "synthese": []}

            # Récupérer les analyses quantitative et qualitative
            analyse_quantitative = mission.get("analyse_quantitative", {})
            analyse_qualitative = mission.get("analyse_qualitative", {})
            
            # Créer un index des analyses pour un accès rapide
            quant_index = {}
            if analyse_quantitative.get("analyse"):
                for item in analyse_quantitative["analyse"]:
                    quant_index[item["compte"]] = item
            
            qual_index = {}
            if analyse_qualitative.get("analyse"):
                for item in analyse_qualitative["analyse"]:
                    qual_index[item["compte"]] = item

            # Préparer les données de synthèse
            synthese_data = []
            for item in grouping:
                compte = item.get('compte', '')
                libelle = item.get('libelle', '')
                solde_n = abs(item.get('solde_n', 0))
                solde_n1 = abs(item.get('solde_n1', 0))
                
                # Récupérer les données quantitatives
                quant_data = quant_index.get(compte, {})
                is_quantitatively_significant = quant_data.get('is_significant', False)
                materiality_threshold = quant_data.get('materiality_threshold', 0)
                percentage_of_threshold = quant_data.get('percentage_of_threshold', 0)
                
                # Récupérer les données qualitatives
                qual_data = qual_index.get(compte, {})
                is_qualitatively_significant = qual_data.get('is_qualitatively_significant', False)
                qualitative_score = qual_data.get('qualitative_score', 0)
                positive_responses = qual_data.get('positive_responses', 0)
                
                # Déterminer le statut final
                final_status = self.determine_final_status(
                    is_quantitatively_significant, 
                    is_qualitatively_significant
                )
                
                # Calculer le niveau de risque
                risk_level = self.calculate_risk_level(
                    is_quantitatively_significant,
                    is_qualitatively_significant,
                    percentage_of_threshold,
                    qualitative_score
                )

                synthese_data.append({
                    "compte": compte,
                    "libelle": libelle,
                    "solde_n": solde_n,
                    "solde_n1": solde_n1,
                    "variation": solde_n - solde_n1,
                    "is_quantitatively_significant": is_quantitatively_significant,
                    "is_qualitatively_significant": is_qualitatively_significant,
                    "materiality_threshold": materiality_threshold,
                    "percentage_of_threshold": percentage_of_threshold,
                    "qualitative_score": qualitative_score,
                    "positive_responses": positive_responses,
                    "final_status": final_status,
                    "risk_level": risk_level,
                    "recommendation": self.get_recommendation(final_status, risk_level)
                })

            # Trier par niveau de risque décroissant
            risk_order = {"Très élevé": 4, "Élevé": 3, "Modéré": 2, "Faible": 1}
            synthese_data.sort(key=lambda x: risk_order.get(x['risk_level'], 0), reverse=True)

            # Statistiques
            total_accounts = len(synthese_data)
            accounts_to_test = len([a for a in synthese_data if a['final_status'] == "À tester"])
            accounts_not_to_test = len([a for a in synthese_data if a['final_status'] == "Ne pas tester"])
            
            risk_distribution = {}
            for item in synthese_data:
                risk = item['risk_level']
                risk_distribution[risk] = risk_distribution.get(risk, 0) + 1

            report = {
                "ok": True,
                "message": f"Synthèse des comptes significatifs générée avec succès",
                "synthese": synthese_data,
                "statistics": {
                    "total_accounts": total_accounts,
                    "accounts_to_test": accounts_to_test,
                    "accounts_not_to_test": accounts_not_to_test,
                    "risk_distribution": risk_distribution,
                    "materiality_threshold": materiality_threshold,
                    "total_amount_to_test": sum([a['solde_n'] for a in synthese_data if a['final_status'] == "À tester"])
                }
            }

            # Sauvegarder la synthèse dans la mission
            db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"synthese_comptes_significatifs": report}})
            return report

        except Exception as e:
            return {"ok": False, "message": str(e), "synthese": []}

    def determine_final_status(self, is_quantitatively_significant, is_qualitatively_significant):
        """Détermine le statut final d'un compte basé sur les analyses quantitative et qualitative"""
        if is_quantitatively_significant and is_qualitatively_significant:
            return "À tester (Quantitatif + Qualitatif)"
        elif is_quantitatively_significant:
            return "À tester (Quantitatif)"
        elif is_qualitatively_significant:
            return "À tester (Qualitatif)"
        else:
            return "Ne pas tester"

    def calculate_risk_level(self, is_quantitatively_significant, is_qualitatively_significant, percentage_of_threshold, qualitative_score):
        """Calcule le niveau de risque d'un compte"""
        if is_quantitatively_significant and is_qualitatively_significant:
            if percentage_of_threshold >= 200 or qualitative_score >= 75:
                return "Très élevé"
            elif percentage_of_threshold >= 150 or qualitative_score >= 50:
                return "Élevé"
            else:
                return "Modéré"
        elif is_quantitatively_significant:
            if percentage_of_threshold >= 200:
                return "Très élevé"
            elif percentage_of_threshold >= 150:
                return "Élevé"
            else:
                return "Modéré"
        elif is_qualitatively_significant:
            if qualitative_score >= 75:
                return "Élevé"
            elif qualitative_score >= 50:
                return "Modéré"
            else:
                return "Faible"
        else:
            return "Faible"

    def get_recommendation(self, final_status, risk_level):
        """Génère une recommandation basée sur le statut final et le niveau de risque"""
        if final_status.startswith("À tester"):
            if risk_level == "Très élevé":
                return "Tests d'audit approfondis obligatoires"
            elif risk_level == "Élevé":
                return "Tests d'audit substantiels recommandés"
            elif risk_level == "Modéré":
                return "Tests d'audit de base suffisants"
            else:
                return "Tests d'audit minimaux"
        else:
            return "Aucun test d'audit requis"

    def presentation_comptes_significatifs(self, id_mission):
        try:
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "presentation": []}

            # Récupérer les données de grouping
            grouping = mission.get("grouping", [])
            if not grouping:
                return {"ok": False, "message": "Données de grouping manquantes", "presentation": []}

            # Récupérer les analyses quantitative et qualitative
            analyse_quantitative = mission.get("analyse_quantitative", {})
            analyse_qualitative = mission.get("analyse_qualitative", {})
            
            # Vérifier que les analyses sont disponibles
            if not analyse_quantitative.get("analyse") or not analyse_qualitative.get("analyse"):
                return {"ok": False, "message": "Analyses quantitative et qualitative requises. Veuillez d'abord exécuter les étapes 7 et 8.", "presentation": []}
            
            # Créer un index des analyses pour un accès rapide
            quant_index = {}
            for item in analyse_quantitative["analyse"]:
                quant_index[item["compte"]] = item
            
            qual_index = {}
            for item in analyse_qualitative["analyse"]:
                qual_index[item["compte"]] = item

            # Préparer les données de présentation
            presentation_data = []
            for item in grouping:
                compte = item.get('compte', '')
                libelle = item.get('libelle', '')
                solde_n = abs(item.get('solde_n', 0))
                solde_n1 = abs(item.get('solde_n1', 0))
                variation = solde_n - solde_n1
                
                # Récupérer les données quantitatives
                quant_data = quant_index.get(compte, {})
                is_quantitatively_significant = quant_data.get('is_significant', False)
                materiality_threshold = quant_data.get('materiality_threshold', 0)
                percentage_of_threshold = quant_data.get('percentage_of_threshold', 0)
                
                # Récupérer les données qualitatives
                qual_data = qual_index.get(compte, {})
                is_qualitatively_significant = qual_data.get('is_qualitatively_significant', False)
                qualitative_score = qual_data.get('qualitative_score', 0)
                positive_responses = qual_data.get('positive_responses', 0)
                
                # Calculer le pourcentage de variation
                variation_percent = 0
                if solde_n1 != 0:
                    variation_percent = (variation / abs(solde_n1)) * 100
                
                # Déterminer le statut de significativité
                if is_quantitatively_significant and is_qualitatively_significant:
                    significativite_status = "Significatif (Quantitatif + Qualitatif)"
                    priorite = "Haute"
                elif is_quantitatively_significant:
                    significativite_status = "Significatif (Quantitatif)"
                    priorite = "Moyenne"
                elif is_qualitatively_significant:
                    significativite_status = "Significatif (Qualitatif)"
                    priorite = "Moyenne"
                else:
                    significativite_status = "Non significatif"
                    priorite = "Faible"
                
                # Générer une recommandation d'audit
                if significativite_status.startswith("Significatif"):
                    if priorite == "Haute":
                        recommandation_audit = "Tests d'audit approfondis obligatoires"
                    else:
                        recommandation_audit = "Tests d'audit substantiels recommandés"
                else:
                    recommandation_audit = "Tests d'audit minimaux ou aucun test"

                presentation_data.append({
                    "compte": compte,
                    "libelle": libelle,
                    "solde_n": solde_n,
                    "solde_n1": solde_n1,
                    "variation": variation,
                    "variation_percent": variation_percent,
                    "is_quantitatively_significant": is_quantitatively_significant,
                    "is_qualitatively_significant": is_qualitatively_significant,
                    "materiality_threshold": materiality_threshold,
                    "percentage_of_threshold": percentage_of_threshold,
                    "qualitative_score": qualitative_score,
                    "positive_responses": positive_responses,
                    "significativite_status": significativite_status,
                    "priorite": priorite,
                    "recommandation_audit": recommandation_audit
                })

            # Trier par priorité et montant décroissant
            priorite_order = {"Haute": 3, "Moyenne": 2, "Faible": 1}
            presentation_data.sort(key=lambda x: (priorite_order.get(x['priorite'], 0), x['solde_n']), reverse=True)

            # Statistiques
            total_accounts = len(presentation_data)
            significant_accounts = len([a for a in presentation_data if a['significativite_status'].startswith("Significatif")])
            non_significant_accounts = total_accounts - significant_accounts
            high_priority_accounts = len([a for a in presentation_data if a['priorite'] == "Haute"])
            total_significant_amount = sum([a['solde_n'] for a in presentation_data if a['significativite_status'].startswith("Significatif")])

            report = {
                "ok": True,
                "message": f"Présentation des comptes significatifs générée avec succès",
                "presentation": presentation_data,
                "statistics": {
                    "total_accounts": total_accounts,
                    "significant_accounts": significant_accounts,
                    "non_significant_accounts": non_significant_accounts,
                    "high_priority_accounts": high_priority_accounts,
                    "total_significant_amount": total_significant_amount,
                    "materiality_threshold": materiality_threshold
                }
            }

            # Sauvegarder la présentation dans la mission
            db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"presentation_comptes_significatifs": report}})
            return report

        except Exception as e:
            return {"ok": False, "message": str(e), "presentation": []}

    def revue_analytique_finale(self, id_mission):
        try:
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                return {"ok": False, "message": "Mission non trouvée", "revue": []}

            # Récupérer les données de grouping
            grouping = mission.get("grouping", [])
            if not grouping:
                return {"ok": False, "message": "Données de grouping manquantes", "revue": []}

            # Récupérer les analyses précédentes (optionnelles)
            synthese = mission.get("synthese_comptes_significatifs", {})
            analyse_quantitative = mission.get("analyse_quantitative", {})
            analyse_qualitative = mission.get("analyse_qualitative", {})
            
            # Créer un index des analyses pour un accès rapide
            synthese_index = {}
            if synthese and synthese.get("synthese"):
                for item in synthese["synthese"]:
                    synthese_index[item["compte"]] = item
            
            # Créer un index des analyses quantitatives si disponibles
            quant_index = {}
            if analyse_quantitative and analyse_quantitative.get("analyse"):
                for item in analyse_quantitative["analyse"]:
                    quant_index[item["compte"]] = item
            
            # Créer un index des analyses qualitatives si disponibles
            qual_index = {}
            if analyse_qualitative and analyse_qualitative.get("analyse"):
                for item in analyse_qualitative["analyse"]:
                    qual_index[item["compte"]] = item

            # Liste des tableaux de revue analytique
            list_tableaux = [
                "CR", "BI", "ERE", "MA", "APE", "TSE", "IT", "PE", "ADP", "ACE",
                "RF", "PA", "IC", "AAVI", "IF", "ST", "AC", "TR", "CP", "DF",
                "DCRA", "FCR", "DFS", "AD", "COMMENTAIRE"
            ]

            # Préparer les données de revue analytique
            revue_data = []
            for item in grouping:
                compte = item.get('compte', '')
                libelle = item.get('libelle', '')
                solde_n = abs(item.get('solde_n', 0))
                solde_n1 = abs(item.get('solde_n1', 0))
                variation = solde_n - solde_n1
                
                # Récupérer les données de synthèse si disponibles
                synthese_data = synthese_index.get(compte, {})
                final_status = synthese_data.get('final_status', 'Non évalué')
                risk_level = synthese_data.get('risk_level', 'Non évalué')
                recommendation = synthese_data.get('recommendation', 'Non évalué')
                
                # Si pas de synthèse, essayer d'utiliser les analyses individuelles
                if final_status == 'Non évalué':
                    # Vérifier les analyses quantitatives
                    quant_data = quant_index.get(compte, {})
                    is_quantitatively_significant = quant_data.get('is_significant', False)
                    
                    # Vérifier les analyses qualitatives
                    qual_data = qual_index.get(compte, {})
                    is_qualitatively_significant = qual_data.get('is_qualitatively_significant', False)
                    
                    # Déterminer le statut basé sur les analyses disponibles
                    if is_quantitatively_significant and is_qualitatively_significant:
                        final_status = 'À tester (Quantitatif + Qualitatif)'
                        risk_level = 'Élevé'
                        recommendation = 'Tests d\'audit substantiels recommandés'
                    elif is_quantitatively_significant:
                        final_status = 'À tester (Quantitatif)'
                        risk_level = 'Modéré'
                        recommendation = 'Tests d\'audit de base suffisants'
                    elif is_qualitatively_significant:
                        final_status = 'À tester (Qualitatif)'
                        risk_level = 'Modéré'
                        recommendation = 'Tests d\'audit de base suffisants'
                    else:
                        # Aucune analyse disponible, utiliser les données de grouping
                        # Suppression du statut "À évaluer" : considérer comme non significatif
                        final_status = 'Ne pas tester'
                        risk_level = 'Faible'
                        recommendation = 'Aucun test d\'audit requis'
                
                # Calculer le pourcentage de variation
                variation_percent = 0
                if solde_n1 != 0:
                    variation_percent = (variation / abs(solde_n1)) * 100
                
                # Générer un commentaire automatique basé sur l'analyse
                commentaire_auto = self.generate_automatic_comment(
                    compte, solde_n, solde_n1, variation_percent, 
                    final_status, risk_level, list_tableaux
                )
                
                # Déterminer le statut de validation
                validation_status = self.determine_validation_status(
                    final_status, risk_level, variation_percent
                )

                revue_data.append({
                    "compte": compte,
                    "libelle": libelle,
                    "solde_n": solde_n,
                    "solde_n1": solde_n1,
                    "variation": variation,
                    "variation_percent": variation_percent,
                    "final_status": final_status,
                    "risk_level": risk_level,
                    "recommendation": recommendation,
                    "commentaire_auto": commentaire_auto,
                    "commentaire_perso": "",  # À remplir par l'auditeur
                    "validation_status": validation_status,
                    "is_validated": False  # À valider par l'auditeur
                })

            # Trier par niveau de risque décroissant
            risk_order = {"Très élevé": 5, "Élevé": 4, "Modéré": 3, "Faible": 2, "À déterminer": 1, "Non évalué": 0}
            revue_data.sort(key=lambda x: risk_order.get(x['risk_level'], 0), reverse=True)

            # Statistiques de validation
            total_accounts = len(revue_data)
            accounts_to_validate = len([a for a in revue_data if a['final_status'].startswith("À tester")])
            accounts_validated = len([a for a in revue_data if a['is_validated']])
            accounts_pending = accounts_to_validate - accounts_validated

            # Calculer le pourcentage de validation
            validation_percentage = (accounts_validated / accounts_to_validate * 100) if accounts_to_validate > 0 else 100

            # Déterminer quelles analyses sont disponibles
            analyses_disponibles = []
            if synthese and synthese.get("synthese"):
                analyses_disponibles.append("Synthèse")
            if analyse_quantitative and analyse_quantitative.get("analyse"):
                analyses_disponibles.append("Quantitative")
            if analyse_qualitative and analyse_qualitative.get("analyse"):
                analyses_disponibles.append("Qualitative")
            
            if analyses_disponibles:
                message = f"Revue analytique finale générée avec succès (Analyses disponibles: {', '.join(analyses_disponibles)})"
            else:
                message = "Revue analytique finale générée avec succès (Évaluation basique basée sur les données de grouping)"

            report = {
                "ok": True,
                "message": message,
                "revue": revue_data,
                "analyses_disponibles": analyses_disponibles,
                "statistics": {
                    "total_accounts": total_accounts,
                    "accounts_to_validate": accounts_to_validate,
                    "accounts_validated": accounts_validated,
                    "accounts_pending": accounts_pending,
                    "validation_percentage": validation_percentage,
                    "audit_status": "En cours" if accounts_pending > 0 else "Validé"
                }
            }

            # Sauvegarder la revue analytique dans la mission
            db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"revue_analytique_finale": report}})
            return report

        except Exception as e:
            return {"ok": False, "message": str(e), "revue": []}

    def generate_automatic_comment(self, compte, solde_n, solde_n1, variation_percent, final_status, risk_level, list_tableaux):
        """Génère un commentaire automatique basé sur l'analyse"""
        # Déterminer le type de compte
        compte_type = "Compte général"
        for tableau in list_tableaux:
            if str(compte).startswith(str(tableau)):
                compte_type = f"Compte {tableau}"
                break
        
        # Commentaire basé sur la variation
        if abs(variation_percent) > 50:
            variation_comment = f"Variation significative de {variation_percent:.1f}%"
        elif abs(variation_percent) > 20:
            variation_comment = f"Variation notable de {variation_percent:.1f}%"
        else:
            variation_comment = f"Variation normale de {variation_percent:.1f}%"
        
        # Commentaire basé sur le statut
        if final_status.startswith("À tester"):
            status_comment = f"Compte significatif nécessitant des tests d'audit"
        elif final_status == "Ne pas tester":
            status_comment = f"Compte non significatif"
        else:
            status_comment = f"Compte non significatif"
        
        # Commentaire basé sur le niveau de risque
        risk_comment = f"Niveau de risque: {risk_level}"
        
        return f"{compte_type}. {variation_comment}. {status_comment}. {risk_comment}."

    def determine_validation_status(self, final_status, risk_level, variation_percent):
        """Détermine le statut de validation d'un compte"""
        if final_status.startswith("À tester"):
            if risk_level == "Très élevé":
                return "Validation obligatoire"
            elif risk_level == "Élevé":
                return "Validation recommandée"
            elif abs(variation_percent) > 30:
                return "Validation recommandée"
            else:
                return "Validation optionnelle"
        else:
            return "Validation non requise"


        return materiality



    # ---------- Analyses grouping ----------

    def make_quantitative_analysis(self, id_mission):

        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})

        grouping = mission['grouping']

        materiality = next((mat for mat in mission['materiality'] if mat.get('choice')), None)



        if materiality is not None:

            for item in grouping:

                value = False

                if int(item['solde_n']) >= int(materiality['materiality']):

                    value = True

                item['materiality'] = value



            result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})

            return result.modified_count

        else:

            return 0



    def make_qualitative_analysis(self, id_mission, significant):

        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})

        grouping = mission['grouping']



        for item in grouping:

            value = next((elt['significant'] for elt in significant if elt.get('compte') == item.get('compte')), None)

            if value is None:

                item['significant'] = False

            else:

                item['significant'] = value



        result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})

        return result.modified_count



    def make_final_sm(self, id_mission):

        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})

        grouping = mission['grouping']



        for item in grouping:

            mat = item.get('materiality', None)

            sign = item.get('significant', None)



            if (mat is not None) and (sign is not None):

                value = ""

                if mat is False and sign is True:

                    value = "non matériel significatif"

                elif mat is False and sign is False:

                    value = "non matériel non significatif"

                elif mat is True and sign is True:

                    value = "matériel significatif"

                elif mat is True and sign is False:

                    value = "matériel non significatif"

                else:

                    value = None

                item['mat_sign'] = value



        result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})

        return result.modified_count, grouping



    # ---------- Récup mission ----------

    def afficher_informations_missions(self, id_client):

        _id = id_client

        query = db.Mission1.find_one({"_id": _id})

        query['_id'] = str(query['_id'])

        return query



    # ---------- Production COMMENTAIRE ----------

    def prod_efi(self, balance_n, balance_n1, balance_variation):

        mapping_path = os.path.join(os.path.dirname(__file__), "..", "mapping_efi.json")

        with open(mapping_path, 'r', encoding='utf-8') as file:

            result = json.load(file)

        mapping = result['structure']



        # nettoyer les champs *_cpt en listes

        for mapp in mapping:

            mapp['brut_cpt'] = mapp['brut_cpt'].split(',') if mapp.get('brut_cpt') is not None else mapp.get('brut_cpt')

            mapp['amor_cpt'] = mapp['amor_cpt'].split(',') if mapp.get('amor_cpt') is not None else mapp.get('amor_cpt')

            mapp['net_cpt'] = mapp['net_cpt'].split(',') if mapp.get('net_cpt') is not None else mapp.get('net_cpt')

            mapp['brut_except_cpt'] = mapp['brut_except_cpt'].split(',') if mapp.get('brut_except_cpt') is not None else mapp.get('brut_except_cpt', [])

            mapp['amor_except_cpt'] = mapp['amor_except_cpt'].split(',') if mapp.get('amor_except_cpt') is not None else mapp.get('amor_except_cpt', [])

            mapp['net_except_cpt'] = mapp['net_except_cpt'].split(',') if mapp.get('net_except_cpt') is not None else mapp.get('net_except_cpt', [])



        datum = {}

        list_efi = ['actif', 'passif', 'pnl']

        for efi in list_efi:

            structure = []

            select_mapping = [elt for elt in mapping if elt['nature'] == efi]
            
            # Log pour vérifier le nombre de lignes dans le mapping
            print(f"📊 Mapping {efi}: {len(select_mapping)} lignes trouvées")

            for idx, data in enumerate(select_mapping):

                row = {}
                
                ref = data.get('ref', '')
                libelle = data.get('libelle', '')
                
                # Log pour debug si ref est vide
                if not ref:
                    print(f"⚠️ Ligne {idx+1} sans ref dans {efi}: {libelle[:50]}")

                # Vérifier si brut_cpt ou net_cpt existe (au moins un des deux doit être défini)
                brut_cpt = data.get('brut_cpt')
                amor_cpt = data.get('amor_cpt')
                net_cpt = data.get('net_cpt')
                
                # brut_cpt et amor_cpt peuvent être None ou une liste (après split)
                has_brut_amor = brut_cpt and amor_cpt
                if isinstance(brut_cpt, list):
                    has_brut_amor = has_brut_amor and len(brut_cpt) > 0
                if isinstance(amor_cpt, list):
                    has_brut_amor = has_brut_amor and len(amor_cpt) > 0
                
                # net_cpt peut être None ou une liste (après split)
                has_net = net_cpt is not None
                if isinstance(net_cpt, list):
                    has_net = has_net and len(net_cpt) > 0
                
                # TOUJOURS inclure la ligne, même sans comptes
                if not has_brut_amor and not has_net:
                    # Si aucune définition de compte n'existe, on crée quand même la ligne avec des valeurs nulles
                    row['ref'] = data.get('ref') or ''
                    row['libelle'] = data.get('libelle') or ''
                    row['note'] = data.get('note')  # Inclure la note si elle existe
                    row['brut_solde_n'] = 0
                    row['amor_solde_n'] = None
                    row['net_solde_n'] = 0
                    row['net_solde_n1'] = 0
                    row['compte_to_be_used'] = ''
                    row['compte_to_be_used_off'] = []
                    structure.append(row)
                    print(f"✅ Ligne {idx+1} ({ref}) ajoutée sans comptes")
                    continue

                if has_brut_amor:

                    # N

                    brut_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['brut_cpt']))

                    amor_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['amor_cpt']))



                    brut_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('brut_except_cpt', [])))

                    amor_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('amor_except_cpt', [])))



                    data['brut_solde_n'] = brut_solde_n + brut_except_n

                    data['amor_solde_n'] = amor_solde_n + amor_except_n

                    data['net_solde_n'] = data['brut_solde_n'] + data['amor_solde_n']



                    # N-1

                    brut_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['brut_cpt']))

                    amor_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['amor_cpt']))

                    net_except_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))

                    data['net_solde_n1'] = brut_n1 + amor_n1 + net_except_n1

                else:

                    net_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['net_cpt']))

                    net_solde_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['net_cpt']))



                    net_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))

                    net_except_n1_bis = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))



                    data['net_solde_n'] = net_solde_n + net_except_n

                    data['net_solde_n1'] = net_solde_n1 + net_except_n1_bis



                row['ref'] = data.get('ref') or ''

                row['libelle'] = data.get('libelle') or ''
                
                row['note'] = data.get('note')  # Inclure la note si elle existe

                row['compte_to_be_used'] = str(data.get('brut_cpt')) + str(data.get('amor_cpt')) + str(data.get('net_cpt')) + str(data.get('brut_except_cpt')) + str(data.get('amor_except_cpt')) + str(data.get('net_except_cpt'))

                row['compte_to_be_used'] = row['compte_to_be_used'].replace('None', '')



                one = data.get('brut_cpt', []) or []

                two = data.get('amor_cpt', []) or []

                three = data.get('net_cpt', []) or []

                four = data.get('brut_except_cpt', []) or []

                five = data.get('amor_except_cpt', []) or []

                six = data.get('net_except_cpt', []) or []



                row['compte_to_be_used_off'] = list(set(one + two + three + four + five + six))



                row['brut_solde_n'] = data.get('brut_solde_n')

                row['amor_solde_n'] = data.get('amor_solde_n')

                row['net_solde_n'] = data.get('net_solde_n')

                row['net_solde_n1'] = data.get('net_solde_n1')



                structure.append(row)
                print(f"✅ Ligne {idx+1} ({ref}) ajoutée avec comptes")

            # Log pour vérifier le nombre de lignes générées
            print(f"📊 Structure {efi} générée: {len(structure)} lignes sur {len(select_mapping)} attendues")
            
            # Log des références pour vérifier
            refs = [row.get('ref', '') for row in structure]
            print(f"📊 Références {efi} ({len(refs)}): {', '.join(refs[:30])}{'...' if len(refs) > 30 else ''}")
            
            # Vérifier si toutes les lignes sont incluses
            if len(structure) < len(select_mapping):
                missing_refs = [data.get('ref', 'NO_REF') for data in select_mapping if data.get('ref', '') not in refs]
                print(f"⚠️ Lignes manquantes dans {efi}: {missing_refs}")

            datum[efi] = structure



        return datum



    # ---------- Piste d'audit ----------

    def audit_trail(self, id_mission):

        # Créer un fichier Excel pour la piste d'audit
        wb = openpyxl.Workbook()

        sheet = wb.active



        columns = ['A', 'B', 'C', 'D', 'E']

        headers = ['Numéro compte', 'Solde n', 'Solde n-1', 'Grouping', 'Code COMMENTAIRE']



        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})

        balances = mission['balance_variation']

        actif = mission['efi']['actif']

        passif = mission['efi']['passif']

        pnl = mission['efi']['pnl']



        efi = actif + passif + pnl



        for i in range(len(columns)):

            sheet[columns[i] + '1'] = headers[i]



        for iteration, data in enumerate(balances):

            new_iteration = str(iteration + 2)

            sheet["A" + new_iteration] = data.get("numero_compte")

            sheet["B" + new_iteration] = data.get("solde_n")

            sheet["C" + new_iteration] = data.get("solde_n1")

            sheet["D" + new_iteration] = data.get("numero_compte")[0:2]



            list_code_efi = []

            for obj in efi:

                for elt in obj['compte_to_be_used_off']:

                    if data['numero_compte'].startswith(elt):

                        list_code_efi.append(obj['ref'])

            list_code_efi = list(set(list_code_efi))

            sheet["E" + new_iteration] = ','.join(list_code_efi)



        namefile = "piste_audit.xlsx"

        wb.save(namefile)



    # ---------- Extract grouping Excel ----------

    def extract_grouping(self, id_mission):

        # Créer un fichier Excel pour l'export du grouping
        wb = openpyxl.Workbook()

        sheet = wb.active



        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

        headers = ['Numéro compte', 'Solde n', 'Solde n-1', 'Grouping', 'Variation', 'Variation %', 'Compte qualitatif', 'Compte quantitatif', 'Compte significatif']



        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})

        balances = mission['balance_variation']

        grouping = mission['grouping']

        materiality = next(item for item in mission['materiality'] if item['choice'] is True)



        for i in range(len(columns)):

            sheet[columns[i] + '1'] = headers[i]



        for iteration, data in enumerate(balances):

            new_iteration = str(iteration + 2)

            sheet["A" + new_iteration] = data.get("numero_compte")

            sheet["B" + new_iteration] = data.get("solde_n")

            sheet["C" + new_iteration] = data.get("solde_n1")



            value_grouping = data.get("numero_compte")[0:2]

            variation = data.get("solde_n") - data.get("solde_n1")



            if variation == 0:

                variation_percent = 0

            elif data.get("solde_n1") == 0:

                variation_percent = 100

            else:

                variation_percent = (variation / data.get("solde_n1")) * 100



            sheet["D" + new_iteration] = value_grouping

            sheet["E" + new_iteration] = variation

            sheet["F" + new_iteration] = variation_percent

            sheet["G" + new_iteration] = next(item['significant'] for item in grouping if item['compte'] == value_grouping)

            sheet["H" + new_iteration] = next(item['materiality'] for item in grouping if item['compte'] == value_grouping)

            sheet["I" + new_iteration] = next(item['mat_sign'] for item in grouping if item['compte'] == value_grouping)



        second_sheet = wb.create_sheet(title="Seuil de matérialité")

        second_headers = ['materiality', 'performance materiality', 'trivial misstatements']

        second_sheet["A1"] = second_headers[0]

        second_sheet["B1"] = second_headers[1]

        second_sheet["C1"] = second_headers[2]



        second_sheet["A2"] = materiality['materiality']

        second_sheet["B2"] = materiality['performance_materiality']

        second_sheet["C2"] = materiality['trivial_misstatements']



        excel_io = BytesIO()

        wb.save(excel_io)

        excel_io.seek(0)

        return excel_io





