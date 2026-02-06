#!/usr/bin/env python3
"""
Script pour tester la détection de l'ordre des colonnes
"""

import sys
import os
import openpyxl

def test_detection_colonnes(fichier_excel):
    """Teste la détection de l'ordre des colonnes dans un fichier Excel"""
    
    if not os.path.exists(fichier_excel):
        print(f"❌ Fichier non trouvé: {fichier_excel}")
        return
    
    print(f"\n{'='*60}")
    print(f"🔍 TEST DE DÉTECTION DES COLONNES")
    print(f"Fichier: {os.path.basename(fichier_excel)}")
    print(f"{'='*60}\n")
    
    try:
        workbook = openpyxl.load_workbook(fichier_excel, data_only=True)
        sheet = workbook.active
        
        print(f"📊 Feuille: {sheet.title}")
        print(f"   Nombre de colonnes: {sheet.max_column}")
        print(f"   Nombre de lignes: {sheet.max_row}\n")
        
        # Trouver la ligne d'en-tête
        header_row = 1
        for row_idx in range(1, min(5, sheet.max_row + 1)):
            row_data = [sheet.cell(row=row_idx, column=col).value for col in range(1, min(9, sheet.max_column + 1))]
            if any('compte' in str(cell).lower() for cell in row_data if cell):
                header_row = row_idx
                print(f"📝 En-tête détecté à la ligne {row_idx}")
                break
        
        # Lire les en-têtes
        header_row_data = [sheet.cell(row=header_row, column=col).value for col in range(1, min(9, sheet.max_column + 1))]
        
        print(f"\n📋 EN-TÊTES DÉTECTÉS:")
        for i, header in enumerate(header_row_data, start=1):
            print(f"   Colonne {i}: '{header}'")
        
        if len(header_row_data) >= 8:
            col7_header = str(header_row_data[6] if len(header_row_data) > 6 else '').strip()
            col8_header = str(header_row_data[7] if len(header_row_data) > 7 else '').strip()
            col7_val = col7_header.lower()
            col8_val = col8_header.lower()
            
            print(f"\n🔍 ANALYSE DES COLONNES 7 ET 8:")
            print(f"   Colonne 7: '{col7_header}' -> '{col7_val}'")
            print(f"   Colonne 8: '{col8_header}' -> '{col8_val}'")
            
            has_solde_credit_col7 = (
                ('solde' in col7_val and 'credit' in col7_val) or
                ('solde' in col7_val and 'crédit' in col7_val) or
                (col7_val and 'credit' in col7_val and 'debit' not in col7_val and 'débit' not in col7_val)
            )
            
            has_solde_debit_col8 = (
                ('solde' in col8_val and 'debit' in col8_val) or
                ('solde' in col8_val and 'débit' in col8_val) or
                (col8_val and 'debit' in col8_val and 'credit' not in col8_val and 'crédit' not in col8_val)
            )
            
            print(f"\n✅ RÉSULTAT DE LA DÉTECTION:")
            if has_solde_credit_col7 and has_solde_debit_col8:
                print(f"   ✅ COLONNES INVERSÉES DÉTECTÉES")
                print(f"   Mapping: credit_fin = colonne 7, debit_fin = colonne 8")
            else:
                print(f"   ✅ FORMAT STANDARD DÉTECTÉ")
                print(f"   Mapping: debit_fin = colonne 7, credit_fin = colonne 8")
            
            # Afficher quelques exemples de données
            print(f"\n📊 EXEMPLES DE DONNÉES (3 premières lignes):")
            for row_idx in range(header_row + 1, min(header_row + 4, sheet.max_row + 1)):
                row = [sheet.cell(row=row_idx, column=col).value for col in range(1, min(9, sheet.max_column + 1))]
                if len(row) >= 8 and row[0]:
                    compte = row[0]
                    col7_val_data = row[6] or 0
                    col8_val_data = row[7] or 0
                    print(f"   Ligne {row_idx}: Compte={compte}, Col7={col7_val_data}, Col8={col8_val_data}")
        
    except Exception as e:
        print(f"❌ Erreur: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python test_detection_colonnes.py <fichier_excel>")
        sys.exit(1)
    
    test_detection_colonnes(sys.argv[1])



