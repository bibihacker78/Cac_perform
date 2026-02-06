#!/usr/bin/env python3
"""
Script pour analyser un fichier Excel et identifier les problèmes avant l'import
Peut être utilisé pour diagnostiquer pourquoi les balances sont vides après l'import
"""

import sys
import os
import openpyxl
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def analyser_fichier_excel(chemin_fichier):
    """Analyse un fichier Excel et identifie les problèmes potentiels"""
    
    print("=" * 80)
    print(f"📊 ANALYSE DU FICHIER EXCEL")
    print(f"Fichier: {os.path.basename(chemin_fichier)}")
    print("=" * 80)
    
    if not os.path.exists(chemin_fichier):
        print(f"❌ Fichier non trouvé: {chemin_fichier}")
        return
    
    try:
        # Charger le fichier Excel
        print(f"\n📂 Chargement du fichier...")
        workbook = openpyxl.load_workbook(chemin_fichier, data_only=True)
        
        print(f"✅ Fichier chargé")
        print(f"   Nombre de feuilles: {len(workbook.sheetnames)}")
        print(f"   Feuilles disponibles: {', '.join(workbook.sheetnames)}")
        
        # Détecter la feuille à utiliser
        sheet = None
        sheet_name = None
        
        accepted_sheet_names = [
            'Balance_des_comptes',
            'BALANCE_2023',
            'BALANCE__2024',
            'Sage',
            'Sheet1',
            'Balance',
            'Comptes',
        ]
        
        for name in accepted_sheet_names:
            if name in workbook.sheetnames:
                sheet = workbook[name]
                sheet_name = name
                print(f"\n✅ Feuille détectée: '{sheet_name}'")
                break
        
        if sheet is None and workbook.sheetnames:
            sheet = workbook[workbook.sheetnames[0]]
            sheet_name = workbook.sheetnames[0]
            print(f"\n⚠️  Utilisation de la première feuille: '{sheet_name}'")
        
        if sheet is None:
            print(f"❌ Aucune feuille trouvée dans le fichier")
            return
        
        print(f"\n📊 Analyse de la feuille '{sheet_name}'")
        print("-" * 80)
        print(f"   Lignes totales: {sheet.max_row}")
        print(f"   Colonnes totales: {sheet.max_column}")
        
        # Analyser les premières lignes pour détecter le format
        print(f"\n📋 Analyse des premières lignes...")
        
        # Chercher la ligne d'en-tête
        header_row = None
        for row_idx in range(1, min(10, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(10, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row_idx, column=col).value
                row_data.append(cell_value)
            
            # Vérifier si c'est un en-tête
            if any(cell and 'compte' in str(cell).lower() for cell in row_data if cell):
                header_row = row_idx
                print(f"   ✅ En-tête détecté à la ligne {row_idx}")
                print(f"      Valeurs: {[str(c)[:20] if c else '' for c in row_data[:8]]}")
                break
        
        if header_row is None:
            print(f"   ⚠️  Aucun en-tête détecté, traitement à partir de la ligne 2")
            header_row = 1
        
        # Analyser les lignes de données
        print(f"\n📊 Analyse des lignes de données (à partir de la ligne {header_row + 1})...")
        print("-" * 80)
        
        lignes_analysees = 0
        lignes_valides = 0
        lignes_vides = 0
        lignes_avec_problemes = []
        
        # Analyser les 50 premières lignes de données
        max_lignes_analyse = min(50, sheet.max_row - header_row)
        
        for row_idx in range(header_row + 1, header_row + max_lignes_analyse + 1):
            row_data = []
            for col in range(1, min(10, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row_idx, column=col).value
                row_data.append(cell_value)
            
            lignes_analysees += 1
            
            # Vérifier si la ligne est vide
            if all(cell is None or str(cell).strip() == '' for cell in row_data[:3]):
                lignes_vides += 1
                if lignes_vides <= 3:
                    print(f"   Ligne {row_idx}: ⚠️  Vide")
                continue
            
            # Extraire les données importantes
            numero_compte = row_data[0] if len(row_data) > 0 else None
            libelle = row_data[1] if len(row_data) > 1 else None
            
            # Problèmes possibles
            problemes = []
            
            if numero_compte is None:
                problemes.append("Numéro compte = None")
            elif str(numero_compte).strip() == "":
                problemes.append("Numéro compte vide")
            
            if len(row_data) < 6:
                problemes.append(f"Moins de 6 colonnes ({len(row_data)} colonnes)")
            
            # Format balance simple nécessite 6 colonnes minimum
            if len(row_data) >= 6:
                try:
                    # Tester la conversion en nombres
                    debit_init = float(row_data[2] or 0) if row_data[2] is not None else 0
                    credit_init = float(row_data[3] or 0) if row_data[3] is not None else 0
                    debit_fin = float(row_data[4] or 0) if row_data[4] is not None else 0
                    credit_fin = float(row_data[5] or 0) if row_data[5] is not None else 0
                except:
                    problemes.append("Valeurs numériques invalides")
            
            if problemes:
                lignes_avec_problemes.append({
                    'ligne': row_idx,
                    'numero': numero_compte,
                    'libelle': libelle,
                    'problemes': problemes
                })
            else:
                lignes_valides += 1
                if lignes_valides <= 5:
                    print(f"   Ligne {row_idx}: ✅ Valide - Compte: {numero_compte}, Libellé: {str(libelle)[:30] if libelle else 'N/A'}")
        
        # Résumé
        print(f"\n" + "=" * 80)
        print(f"📊 RÉSUMÉ DE L'ANALYSE")
        print("=" * 80)
        print(f"   Lignes analysées: {lignes_analysees}")
        print(f"   ✅ Lignes valides: {lignes_valides}")
        print(f"   ⚠️  Lignes vides: {lignes_vides}")
        print(f"   ❌ Lignes avec problèmes: {len(lignes_avec_problemes)}")
        
        if lignes_valides == 0 and lignes_analysees > 0:
            print(f"\n❌ PROBLÈME CRITIQUE:")
            print(f"   Aucune ligne valide trouvée!")
            
            if lignes_avec_problemes:
                print(f"\n   Exemples de problèmes trouvés:")
                for prob in lignes_avec_problemes[:5]:
                    print(f"      Ligne {prob['ligne']}: {', '.join(prob['problemes'])}")
                    print(f"         Numéro: {repr(prob['numero'])}, Libellé: {repr(prob['libelle'])}")
            
            print(f"\n💡 SOLUTIONS:")
            print(f"   1. Vérifiez que la première colonne contient des numéros de compte")
            print(f"   2. Vérifiez que vous avez au moins 6 colonnes (format balance simple)")
            print(f"   3. Vérifiez qu'il n'y a pas de lignes complètement vides avant les données")
            
        elif lignes_valides > 0:
            print(f"\n✅ Le fichier semble valide")
            print(f"   {lignes_valides} lignes valides détectées")
            
            # Estimer le nombre total de lignes valides
            if lignes_analysees < sheet.max_row - header_row:
                total_estime = lignes_valides * (sheet.max_row - header_row) / lignes_analysees
                print(f"   Estimation totale: ~{int(total_estime)} lignes valides")
            
            if lignes_avec_problemes:
                print(f"\n   ⚠️  Quelques problèmes détectés sur {len(lignes_avec_problemes)} lignes")
                print(f"   Mais le fichier devrait pouvoir être importé")
        
        # Format détecté
        print(f"\n📋 FORMAT DÉTECTÉ:")
        if len(row_data) >= 6:
            print(f"   ✅ Format balance simple (6 colonnes)")
            print(f"      Colonne 1: Numéro de compte")
            print(f"      Colonne 2: Libellé")
            print(f"      Colonne 3: Débit initial")
            print(f"      Colonne 4: Crédit initial")
            print(f"      Colonne 5: Débit final")
            print(f"      Colonne 6: Crédit final")
        elif len(row_data) >= 8:
            print(f"   ✅ Format standard (8 colonnes)")
        else:
            print(f"   ⚠️  Format non standard ({len(row_data)} colonnes)")
        
    except Exception as e:
        print(f"❌ Erreur lors de l'analyse: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python analyser_fichier_excel.py <chemin_vers_fichier.xlsx>")
        print("\nExemple:")
        print("  python analyser_fichier_excel.py ../docs/Balance_2024.xlsx")
        print("  python analyser_fichier_excel.py \"C:/Mes Documents/Balance.xlsx\"")
        sys.exit(1)
    
    chemin_fichier = sys.argv[1]
    analyser_fichier_excel(chemin_fichier)









